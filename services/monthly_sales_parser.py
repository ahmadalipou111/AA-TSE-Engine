from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


def normalize_text(value: Any) -> str:
    """Normalize Persian/Arabic text for reliable comparisons."""
    if value is None:
        return ""

    text = str(value)
    replacements = {
        "ي": "ی",
        "ى": "ی",
        "ك": "ک",
        "ۀ": "ه",
        "ة": "ه",
        "ؤ": "و",
        "إ": "ا",
        "أ": "ا",
        "ٱ": "ا",
        "ـ": "",
        "\u200c": " ",
        "\u200f": "",
        "\u200e": "",
        "\ufeff": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def compact_text(value: Any) -> str:
    """Return normalized text without spaces and punctuation."""
    text = normalize_text(value).lower()
    return re.sub(r"[^0-9a-zآ-ی]+", "", text)


def normalize_digits(value: Any) -> str:
    """Convert Persian and Arabic digits to English digits."""
    if value is None:
        return ""

    return str(value).translate(
        str.maketrans(
            "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
            "01234567890123456789",
        )
    )


def numeric_value(value: Any) -> int | float | None:
    """Convert an Excel value or formatted numeric text to a number."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value

    text = normalize_digits(normalize_text(value))
    text = text.replace(",", "").replace("٬", "").replace(" ", "")
    text = text.replace("−", "-")
    if text in {"", "-", "—"}:
        return None

    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


@dataclass(frozen=True)
class MonthlySalesValues:
    """The ten highlighted monthly-sales values from one CODAL workbook."""

    symbol: str
    report_file: str
    sheet_name: str
    report_month: str | None

    export_previous_original: int | float | None
    export_previous_corrected: int | float | None
    export_current_month: int | float | None
    export_ytd_current: int | float | None
    export_ytd_prior_year: int | float | None

    total_previous_original: int | float | None
    total_previous_corrected: int | float | None
    total_current_month: int | float | None
    total_ytd_current: int | float | None
    total_ytd_prior_year: int | float | None

    status: str = "OK"
    note: str = ""


def merged_anchor_value(worksheet: Any, row: int, column: int) -> Any:
    """Return the visible value when a cell belongs to a merged range."""
    value = worksheet.cell(row, column).value
    if value is not None:
        return value

    coordinate = worksheet.cell(row, column).coordinate
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return worksheet.cell(
                merged_range.min_row,
                merged_range.min_col,
            ).value
    return None


def find_row(
    worksheet: Any,
    exact_labels: Iterable[str],
    start_row: int = 1,
    end_row: int | None = None,
) -> int | None:
    """Find the first row containing one of the exact normalized labels."""
    targets = {compact_text(label) for label in exact_labels}
    final_row = min(end_row or worksheet.max_row, worksheet.max_row)

    for row in range(start_row, final_row + 1):
        for column in range(1, min(worksheet.max_column, 8) + 1):
            if compact_text(worksheet.cell(row, column).value) in targets:
                return row
    return None


def find_sales_header_row(worksheet: Any) -> int | None:
    """Locate the production-and-sales period header."""
    for row in range(1, min(worksheet.max_row, 50) + 1):
        values = [
            normalize_text(merged_anchor_value(worksheet, row, column))
            for column in range(1, worksheet.max_column + 1)
        ]
        joined = " | ".join(value for value in values if value)

        if "دوره یک ماهه" not in joined and "دوره یکماهه" not in joined:
            continue
        if "از ابتدای سال مالی" not in joined:
            continue

        next_row = " | ".join(
            normalize_text(worksheet.cell(row + 1, column).value)
            for column in range(1, worksheet.max_column + 1)
            if row < worksheet.max_row
            and worksheet.cell(row + 1, column).value is not None
        )
        if "مبلغ فروش" in next_row:
            return row

    return None


def detect_period_columns(worksheet: Any, header_row: int) -> dict[str, int]:
    """Identify the five requested amount columns without fixed coordinates."""
    amount_columns = [
        column
        for column in range(1, worksheet.max_column + 1)
        if "مبلغ فروش"
        in normalize_text(worksheet.cell(header_row + 1, column).value)
    ]

    period_columns: dict[str, int] = {}
    ytd_columns: list[int] = []

    for column in amount_columns:
        header = normalize_text(
            merged_anchor_value(worksheet, header_row, column)
        )

        if "اصلاحات" in header and "اصلاح شده" not in header:
            continue
        if "اصلاح شده" in header:
            period_columns["previous_corrected"] = column
        elif "دوره یک ماهه" in header or "دوره یکماهه" in header:
            period_columns["current_month"] = column
        elif "از ابتدای سال مالی" in header:
            ytd_columns.append(column)

    if len(ytd_columns) >= 1:
        period_columns["previous_original"] = ytd_columns[0]
    if len(ytd_columns) >= 2:
        period_columns["ytd_current"] = ytd_columns[1]
    if len(ytd_columns) >= 3:
        period_columns["ytd_prior_year"] = ytd_columns[2]

    required = {
        "previous_original",
        "previous_corrected",
        "current_month",
        "ytd_current",
        "ytd_prior_year",
    }
    missing = required.difference(period_columns)
    if missing:
        raise ValueError(
            "Could not identify these sales columns: "
            + ", ".join(sorted(missing))
        )

    return period_columns


def extract_report_month(
    worksheet: Any,
    header_row: int,
    current_month_column: int,
) -> str | None:
    """Extract a Jalali report date such as 1405/04/31 from the header."""
    text = normalize_digits(
        normalize_text(
            merged_anchor_value(
                worksheet,
                header_row,
                current_month_column,
            )
        )
    )
    match = re.search(r"(1[34]\d{2}/\d{2}/\d{2})", text)
    return match.group(1) if match else None


def parse_monthly_sales_workbook(
    report_path: str | Path,
    symbol: str,
) -> MonthlySalesValues:
    """Extract the highlighted export and total-sales values from one report."""
    path = Path(report_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Monthly report not found: {path}")

    workbook = load_workbook(path, read_only=False, data_only=True)
    last_error = "Production and sales table was not found."

    try:
        for worksheet in workbook.worksheets:
            try:
                header_row = find_sales_header_row(worksheet)
                if header_row is None:
                    continue

                columns = detect_period_columns(worksheet, header_row)
                purchase_row = find_row(
                    worksheet,
                    ["خرید مواد اولیه"],
                    start_row=header_row + 1,
                )
                section_end = (purchase_row - 1) if purchase_row else None

                export_row = find_row(
                    worksheet,
                    ["جمع فروش صادراتی"],
                    start_row=header_row + 1,
                    end_row=section_end,
                )
                total_row = find_row(
                    worksheet,
                    ["جمع"],
                    start_row=header_row + 1,
                    end_row=section_end,
                )

                if export_row is None:
                    raise ValueError("Row 'جمع فروش صادراتی' was not found.")
                if total_row is None:
                    raise ValueError("Grand-total row 'جمع' was not found.")

                def value_at(row: int, key: str) -> int | float | None:
                    return numeric_value(
                        worksheet.cell(row, columns[key]).value
                    )

                return MonthlySalesValues(
                    symbol=normalize_text(symbol),
                    report_file=path.name,
                    sheet_name=worksheet.title,
                    report_month=extract_report_month(
                        worksheet,
                        header_row,
                        columns["current_month"],
                    ),
                    export_previous_original=value_at(
                        export_row, "previous_original"
                    ),
                    export_previous_corrected=value_at(
                        export_row, "previous_corrected"
                    ),
                    export_current_month=value_at(
                        export_row, "current_month"
                    ),
                    export_ytd_current=value_at(
                        export_row, "ytd_current"
                    ),
                    export_ytd_prior_year=value_at(
                        export_row, "ytd_prior_year"
                    ),
                    total_previous_original=value_at(
                        total_row, "previous_original"
                    ),
                    total_previous_corrected=value_at(
                        total_row, "previous_corrected"
                    ),
                    total_current_month=value_at(
                        total_row, "current_month"
                    ),
                    total_ytd_current=value_at(
                        total_row, "ytd_current"
                    ),
                    total_ytd_prior_year=value_at(
                        total_row, "ytd_prior_year"
                    ),
                )
            except Exception as exc:
                last_error = f"{worksheet.title}: {exc}"

        return MonthlySalesValues(
            symbol=normalize_text(symbol),
            report_file=path.name,
            sheet_name="",
            report_month=None,
            export_previous_original=None,
            export_previous_corrected=None,
            export_current_month=None,
            export_ytd_current=None,
            export_ytd_prior_year=None,
            total_previous_original=None,
            total_previous_corrected=None,
            total_current_month=None,
            total_ytd_current=None,
            total_ytd_prior_year=None,
            status="ERROR",
            note=last_error,
        )
    finally:
        workbook.close()
