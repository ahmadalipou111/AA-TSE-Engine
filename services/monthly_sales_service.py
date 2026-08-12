from __future__ import annotations

import re
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import requests
from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.monthly_sales_parser import (
    MonthlySalesValues,
    parse_monthly_sales_workbook,
)


def normalize_digits(value: Any) -> str:
    """Convert Persian and Arabic digits to English digits."""

    if value is None:
        return ""

    translation = str.maketrans(
        "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
        "01234567890123456789",
    )

    return str(value).translate(translation).strip()

def normalize_text(value: Any) -> str:
    """
    Normalize Persian/Arabic text for reliable symbol matching.

    Handles:
    - Arabic/Persian variants of Yeh and Kaf
    - Unicode presentation forms
    - zero-width and direction-control characters
    - non-breaking spaces
    - Persian and Arabic digits
    """

    if value is None:
        return ""

    text = unicodedata.normalize("NFKC", str(value))

    replacements = {
        "ي": "ی",
        "ى": "ی",
        "ئ": "ی",
        "ك": "ک",
        "ۀ": "ه",
        "ة": "ه",
        "ؤ": "و",
        "\u00a0": " ",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    # Remove all invisible formatting and direction-control characters.
    text = "".join(
        character
        for character in text
        if unicodedata.category(character) != "Cf"
    )

    text = normalize_digits(text)
    text = re.sub(r"\s+", " ", text)

    return text.strip()

def safe_filename(value: str) -> str:
    """Create a filename safe for macOS and Windows."""

    value = normalize_text(value)
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = re.sub(r"\s+", "_", value)

    return value.strip("._ ") or "report"


def first_non_empty(
    dictionary: dict[str, Any],
    *keys: str,
) -> Any:
    """Return the first non-empty dictionary value."""

    for key in keys:
        value = dictionary.get(key)

        if value is not None and str(value).strip():
            return value

    return ""


@dataclass(frozen=True)
class MonthlySalesReport:
    """Normalized CODAL monthly-sales announcement."""

    symbol: str
    company: str
    title: str
    report_date: str
    publish_date: str
    publish_time: str
    code: str
    excel_link: str
    announcement_link: str
    pdf_link: str
    attachment_link: str
    source_page: int

    @property
    def unique_key(self) -> tuple[str, str, str, str]:
        """Return a stable key used for duplicate removal."""

        return (
            normalize_text(self.symbol),
            normalize_text(self.title),
            normalize_digits(self.publish_date),
            self.excel_link,
        )


@dataclass(frozen=True)
class ParsedMonthlySalesReport:
    """Connect an announcement, local file and extracted values."""

    report: MonthlySalesReport
    local_file: Path | None
    values: MonthlySalesValues


class MonthlySalesService:
    """
    Retrieve, download and parse CODAL monthly-sales reports.

    Pipeline:

    1. Read AA-TSE symbols.
    2. Retrieve category-3 CODAL announcements.
    3. Keep only AA-TSE symbols.
    4. Download original Excel files automatically.
    5. Parse the selected monthly-sales values.
    6. Produce consolidated output workbooks.
    """

    CEMENT_SHEET = "Cement"
    MOPFRA_SHEET = "MOPFRA"

    CEMENT_SYMBOL_COLUMN = 17
    MOPFRA_SYMBOL_COLUMN = 18

    DATA_START_ROW = 5

    def __init__(
        self,
        codal_api: Any,
        project_root: str | Path | None = None,
        request_delay: float = 0.15,
        request_timeout: int = 60,
    ) -> None:
        self.codal_api = codal_api

        self.project_root = (
            Path(project_root).expanduser().resolve()
            if project_root
            else Path(__file__).resolve().parent.parent
        )

        self.output_dir = self.project_root / "output"
        self.download_dir = (
            self.output_dir / "monthly_sales_downloads"
        )

        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.download_dir.mkdir(parents=True, exist_ok=True)

        self.request_delay = max(0.0, request_delay)
        self.request_timeout = request_timeout

        self.session = requests.Session()

        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X) "
                    "AppleWebKit/537.36 "
                    "(KHTML, like Gecko) "
                    "Chrome/131.0 Safari/537.36"
                ),
                "Accept": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet,"
                    "application/vnd.ms-excel,"
                    "application/octet-stream,*/*"
                ),
            }
        )

    # ------------------------------------------------------------------
    # AA-TSE symbols
    # ------------------------------------------------------------------

    def read_aa_tse_symbols(
        self,
        workbook_path: str | Path,
    ) -> dict[str, list[str]]:
        """Read valid symbols from Cement and MOPFRA worksheets."""

        path = Path(workbook_path).expanduser().resolve()

        if not path.exists():
            raise FileNotFoundError(
                f"AA-TSE workbook not found: {path}"
            )

        logger.info("Reading symbols from {}", path.name)

        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
        )

        try:
            cement_symbols = self._read_sheet_symbols(
                workbook=workbook,
                sheet_name=self.CEMENT_SHEET,
                symbol_column=self.CEMENT_SYMBOL_COLUMN,
            )

            mopfra_symbols = self._read_sheet_symbols(
                workbook=workbook,
                sheet_name=self.MOPFRA_SHEET,
                symbol_column=self.MOPFRA_SYMBOL_COLUMN,
            )
        finally:
            workbook.close()

        all_symbols = sorted(
            set(cement_symbols).union(mopfra_symbols)
        )

        logger.success(
            "Symbols loaded | Cement={} | MOPFRA={} | Unique={}",
            len(cement_symbols),
            len(mopfra_symbols),
            len(all_symbols),
        )

        return {
            "cement": cement_symbols,
            "mopfra": mopfra_symbols,
            "all": all_symbols,
        }

    def _read_sheet_symbols(
        self,
        workbook: Any,
        sheet_name: str,
        symbol_column: int,
    ) -> list[str]:
        """Read genuine stock symbols from one AA-TSE worksheet."""

        if sheet_name not in workbook.sheetnames:
            raise KeyError(
                f"Worksheet '{sheet_name}' was not found."
            )

        worksheet = workbook[sheet_name]
        symbols: set[str] = set()

        for row_number in range(
            self.DATA_START_ROW,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_number,
                column=symbol_column,
            ).value

            symbol = normalize_text(value)
            symbol = re.sub(r"\s+", "", symbol)

            if self._looks_like_symbol(symbol):
                symbols.add(symbol)

        return sorted(symbols)

    @staticmethod
    def _looks_like_symbol(value: str) -> bool:
        """
        Accept only genuine Persian stock symbols.

        This rejects formulas, numbers, percentages, English control
        letters and configuration values located below the company list.
        """

        if not value:
            return False

        if value.startswith("="):
            return False

        compact_value = re.sub(r"\s+", "", value)

        # Genuine TSE symbols in the current AA-TSE universe consist
        # only of Persian letters.
        if not re.fullmatch(r"[آ-ی]+", compact_value):
            return False

        return 2 <= len(compact_value) <= 20

    # ------------------------------------------------------------------
    # CODAL announcements
    # ------------------------------------------------------------------

    def fetch_monthly_reports(
        self,
        date_start: str | None = None,
        date_end: str | None = None,
        max_pages: int | None = 50,
    ) -> list[MonthlySalesReport]:
        """Retrieve category-3 announcements from BRSAPI."""

        reports: list[MonthlySalesReport] = []
        page = 1
        total_pages: int | None = None

        while True:
            logger.info(
                "Fetching monthly reports | page={} | {} to {}",
                page,
                date_start or "-",
                date_end or "-",
            )

            response = self.codal_api.get_monthly_sales_reports(
                date_start=date_start,
                date_end=date_end,
                page=page,
            )

            if not isinstance(response, dict):
                raise RuntimeError(
                    "Unexpected BRSAPI response type."
                )

            announcements = response.get("announcement", [])

            if not isinstance(announcements, list):
                raise RuntimeError(
                    "BRSAPI field 'announcement' is not a list."
                )

            if total_pages is None:
                total_pages = self._to_int(
                    response.get("count_page")
                )

                logger.info(
                    "Available API pages: {}",
                    total_pages,
                )

            if not announcements:
                break

            for raw_report in announcements:
                parsed_report = self._parse_report(
                    raw_report=raw_report,
                    source_page=page,
                )

                if parsed_report is not None:
                    reports.append(parsed_report)

            if max_pages is not None and page >= max_pages:
                logger.warning(
                    "Maximum API page limit reached: {}",
                    max_pages,
                )
                break

            if total_pages and page >= total_pages:
                break

            page += 1

            if self.request_delay:
                time.sleep(self.request_delay)

        unique_reports = self._deduplicate(reports)

        logger.success(
            "Monthly reports received | Raw={} | Unique={}",
            len(reports),
            len(unique_reports),
        )

        return unique_reports

    @staticmethod
    def _to_int(value: Any) -> int:
        """Convert a value to integer without raising an error."""

        try:
            return int(normalize_digits(value))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _parse_report(
        raw_report: Any,
        source_page: int,
    ) -> MonthlySalesReport | None:
        """Convert one API dictionary into MonthlySalesReport."""

        if not isinstance(raw_report, dict):
            return None

        symbol = normalize_text(
            first_non_empty(
                raw_report,
                "l18",
                "symbol",
                "ticker",
            )
        )

        symbol = re.sub(r"\s+", "", symbol)

        if not symbol:
            return None

        excel_link = str(
            first_non_empty(
                raw_report,
                "link_excel",
                "excel",
                "excel_url",
                "ExcelUrl",
            )
        ).strip()

        announcement_link = str(
            first_non_empty(
                raw_report,
                "link",
                "url",
                "announcement_url",
            )
        ).strip()

        pdf_link = str(
            first_non_empty(
                raw_report,
                "link_pdf",
                "pdf",
                "pdf_url",
            )
        ).strip()

        attachment_link = str(
            first_non_empty(
                raw_report,
                "link_attachment",
                "attachment",
                "attachment_url",
            )
        ).strip()

        return MonthlySalesReport(
            symbol=symbol,
            company=normalize_text(
                first_non_empty(
                    raw_report,
                    "l30",
                    "company",
                    "company_name",
                )
            ),
            title=normalize_text(
                first_non_empty(
                    raw_report,
                    "title",
                    "subject",
                )
            ),
            report_date=normalize_digits(
                first_non_empty(
                    raw_report,
                    "date_title",
                    "report_date",
                )
            ),
            publish_date=normalize_digits(
                first_non_empty(
                    raw_report,
                    "date_publish",
                    "publish_date",
                )
            ),
            publish_time=normalize_digits(
                first_non_empty(
                    raw_report,
                    "time_publish",
                    "publish_time",
                )
            ),
            code=normalize_text(
                first_non_empty(
                    raw_report,
                    "code",
                    "letter_code",
                )
            ),
            excel_link=excel_link,
            announcement_link=announcement_link,
            pdf_link=pdf_link,
            attachment_link=attachment_link,
            source_page=source_page,
        )

    @staticmethod
    def _deduplicate(
        reports: Iterable[MonthlySalesReport],
    ) -> list[MonthlySalesReport]:
        """Remove duplicate reports."""

        seen: set[tuple[str, str, str, str]] = set()
        result: list[MonthlySalesReport] = []

        for report in reports:
            if report.unique_key in seen:
                continue

            seen.add(report.unique_key)
            result.append(report)

        return result

    def filter_reports_for_symbols(
        self,
        reports: Iterable[MonthlySalesReport],
        symbols: Iterable[str],
        target_period: str | None = None,
    ) -> list[MonthlySalesReport]:
        """Keep reports belonging to AA-TSE symbols."""

        symbol_set = {
            re.sub(r"\s+", "", normalize_text(symbol))
            for symbol in symbols
            if normalize_text(symbol)
        }

        normalized_period = (
            normalize_digits(target_period)
            if target_period
            else ""
        )

        matched: list[MonthlySalesReport] = []

        for report in reports:
            report_symbol = re.sub(
                r"\s+",
                "",
                normalize_text(report.symbol),
            )

            searchable_symbol_text = normalize_text(
                f"{report.symbol} {report.company} {report.title}"
            )
            searchable_symbol_text = re.sub(
                r"\s+",
                "",
                searchable_symbol_text,
            )

            matched_symbol = next(
    (
        symbol
        for symbol in symbol_set
        if report_symbol == symbol
    ),
    None,
)

            if matched_symbol is None:
                continue

            if normalized_period:
                searchable = normalize_digits(
                    f"{report.title} {report.report_date}"
                )

                if normalized_period not in searchable:
                    continue

            matched.append(report)

        matched = self._deduplicate(matched)

        matched.sort(
            key=lambda report: (
                report.publish_date,
                report.publish_time,
                report.symbol,
            ),
            reverse=True,
        )

        logger.success(
            "Reports matched to AA-TSE symbols: {}",
            len(matched),
        )

        return matched

    # ------------------------------------------------------------------
    # Announcement list output
    # ------------------------------------------------------------------

    def export_reports_to_excel(
        self,
        reports: Iterable[MonthlySalesReport],
        output_path: str | Path | None = None,
    ) -> Path:
        """Create a consolidated announcement-list workbook."""

        report_list = list(reports)

        destination = (
            Path(output_path).expanduser().resolve()
            if output_path
            else self.output_dir
            / "AA-TSE-Monthly-Sales-Reports.xlsx"
        )

        destination.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Monthly Reports"
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.rightToLeft = True

        headers = [
            "No.",
            "Symbol",
            "Company",
            "Report Title",
            "Report Period",
            "Publish Date",
            "Publish Time",
            "Code",
            "Excel Available",
            "Excel URL",
            "Announcement URL",
            "PDF URL",
            "Attachment URL",
            "API Page",
        ]

        worksheet.append(headers)

        for index, report in enumerate(
            report_list,
            start=1,
        ):
            worksheet.append(
                [
                    index,
                    report.symbol,
                    report.company,
                    report.title,
                    report.report_date,
                    report.publish_date,
                    report.publish_time,
                    report.code,
                    "Yes" if report.excel_link else "No",
                    report.excel_link,
                    report.announcement_link,
                    report.pdf_link,
                    report.attachment_link,
                    report.source_page,
                ]
            )

        self._format_header(worksheet)

        self._set_widths(
            worksheet,
            {
                "A": 8,
                "B": 16,
                "C": 32,
                "D": 65,
                "E": 17,
                "F": 17,
                "G": 14,
                "H": 14,
                "I": 18,
                "J": 65,
                "K": 65,
                "L": 65,
                "M": 65,
                "N": 12,
            },
        )

        worksheet.auto_filter.ref = (
            f"A1:N{max(1, worksheet.max_row)}"
        )

        workbook.save(destination)

        logger.success(
            "Announcement workbook created: {}",
            destination,
        )

        return destination

    # ------------------------------------------------------------------
    # Downloads
    # ------------------------------------------------------------------

    def build_report_filename(
        self,
        report: MonthlySalesReport,
    ) -> str:
        """Build a stable local filename for one report."""

        report_date = (
            report.report_date.replace("/", "-")
            or report.publish_date.replace("/", "-")
            or "unknown-date"
        )

        code_part = (
            f"_{safe_filename(report.code)}"
            if report.code
            else ""
        )

        return (
            f"{safe_filename(report.symbol)}_"
            f"{safe_filename(report_date)}"
            f"{code_part}.xlsx"
        )

    def download_one_report(
        self,
        report: MonthlySalesReport,
        overwrite: bool = False,
    ) -> tuple[Path | None, str]:
        """
        Download one CODAL Excel file.

        Returns:
            (local_path, status)

        Status values:
            downloaded
            existing
            failed
        """

        if not report.excel_link:
            return None, "failed"

        destination = (
            self.download_dir
            / self.build_report_filename(report)
        )

        if destination.exists() and not overwrite:
            return destination, "existing"

        temporary_path = destination.with_suffix(".download")

        try:
            response = self.session.get(
                report.excel_link,
                timeout=self.request_timeout,
                allow_redirects=True,
            )

            response.raise_for_status()

            if not response.content:
                raise RuntimeError(
                    "Downloaded file is empty."
                )

            # XLSX files are ZIP containers and usually start with PK.
            if not response.content.startswith(b"PK"):
                preview = response.content[:100].decode(
                    "utf-8",
                    errors="ignore",
                )

                raise RuntimeError(
                    "The downloaded response is not an XLSX file. "
                    f"Response preview: {preview!r}"
                )

            temporary_path.write_bytes(response.content)
            temporary_path.replace(destination)

            return destination, "downloaded"

        except Exception as exc:
            if temporary_path.exists():
                temporary_path.unlink()

            logger.error(
                "Download failed | {} | {}",
                report.symbol,
                exc,
            )

            return None, "failed"

    def download_report_excels(
        self,
        reports: Iterable[MonthlySalesReport],
        overwrite: bool = False,
    ) -> dict[str, Any]:
        """Download all matched monthly-sales files."""

        downloaded: list[Path] = []
        skipped: list[Path] = []
        failed: list[str] = []

        report_files: list[
            tuple[MonthlySalesReport, Path]
        ] = []

        for report in reports:
            if not report.excel_link:
                failed.append(
                    f"{report.symbol}: Excel link missing"
                )
                continue

            local_path, status = self.download_one_report(
                report=report,
                overwrite=overwrite,
            )

            if status == "downloaded" and local_path:
                downloaded.append(local_path)
                report_files.append((report, local_path))

            elif status == "existing" and local_path:
                skipped.append(local_path)
                report_files.append((report, local_path))

            else:
                failed.append(
                    f"{report.symbol}: download failed"
                )

            if self.request_delay:
                time.sleep(self.request_delay)

        logger.success(
            "Excel download completed | "
            "Downloaded={} | Existing={} | Failed={}",
            len(downloaded),
            len(skipped),
            len(failed),
        )

        return {
            "downloaded": downloaded,
            "skipped": skipped,
            "failed": failed,
            "report_files": report_files,
        }

    # ------------------------------------------------------------------
    # Parse downloaded files
    # ------------------------------------------------------------------

    def parse_downloaded_reports(
        self,
        report_files: Iterable[
            tuple[MonthlySalesReport, Path]
        ],
    ) -> list[ParsedMonthlySalesReport]:
        """Parse every successfully available local report."""

        parsed_results: list[ParsedMonthlySalesReport] = []

        for report, local_path in report_files:
            logger.info(
                "Parsing monthly report | {} | {}",
                report.symbol,
                local_path.name,
            )

            try:
                values = parse_monthly_sales_workbook(
                    report_path=local_path,
                    symbol=report.symbol,
                )

            except Exception as exc:
                logger.exception(
                    "Unexpected parser failure | {}",
                    report.symbol,
                )

                values = self._error_values(
                    report=report,
                    local_file=local_path,
                    note=str(exc),
                )

            parsed_results.append(
                ParsedMonthlySalesReport(
                    report=report,
                    local_file=local_path,
                    values=values,
                )
            )

        ok_count = sum(
            item.values.status == "OK"
            for item in parsed_results
        )

        error_count = len(parsed_results) - ok_count

        logger.success(
            "Parsing completed | OK={} | Errors={}",
            ok_count,
            error_count,
        )

        return parsed_results

    @staticmethod
    def _error_values(
        report: MonthlySalesReport,
        local_file: Path | None,
        note: str,
    ) -> MonthlySalesValues:
        """Create a standard parser-error result."""

        return MonthlySalesValues(
            symbol=report.symbol,
            report_file=(
                local_file.name
                if local_file
                else ""
            ),
            sheet_name="",
            report_month=None,
            export_previous_original=None,
            export_previous_corrected=None,
            export_current_month=None,
            export_ytd_current=None,
            export_ytd_prior_year=None,
            total_previous_original=None,
            total_previous_corrected=None,
            total_current_month=None,
            total_ytd_current=None,
            total_ytd_prior_year=None,
            status="ERROR",
            note=note,
        )

    # ------------------------------------------------------------------
    # Extracted data output
    # ------------------------------------------------------------------

    def export_parsed_values_to_excel(
        self,
        parsed_results: Iterable[ParsedMonthlySalesReport],
        output_path: str | Path | None = None,
    ) -> Path:
        """Write extracted values to a new workbook."""

        results = list(parsed_results)

        destination = (
            Path(output_path).expanduser().resolve()
            if output_path
            else self.output_dir
            / "AA-TSE-Monthly-Sales-Extracted.xlsx"
        )

        destination.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()

        data_sheet = workbook.active
        data_sheet.title = "Extracted Data"
        data_sheet.freeze_panes = "A2"
        data_sheet.sheet_view.rightToLeft = True

        headers = [
            "No.",
            "Symbol",
            "Company",
            "Report Month",
            "Publish Date",
            "Publish Time",
            "Report Code",
            "Status",
            "Parser Note",
            "Local Report File",
            "Source Sheet",
            "Export - Previous Original",
            "Export - Previous Corrected",
            "Export - Current Month",
            "Export - YTD Current",
            "Export - YTD Prior Year",
            "Total - Previous Original",
            "Total - Previous Corrected",
            "Total - Current Month",
            "Total - YTD Current",
            "Total - YTD Prior Year",
            "Report Title",
            "Announcement URL",
            "Excel URL",
        ]

        data_sheet.append(headers)

        for index, item in enumerate(results, start=1):
            report = item.report
            values = item.values

            data_sheet.append(
                [
                    index,
                    values.symbol,
                    report.company,
                    values.report_month,
                    report.publish_date,
                    report.publish_time,
                    report.code,
                    values.status,
                    values.note,
                    (
                        str(item.local_file)
                        if item.local_file
                        else ""
                    ),
                    values.sheet_name,
                    values.export_previous_original,
                    values.export_previous_corrected,
                    values.export_current_month,
                    values.export_ytd_current,
                    values.export_ytd_prior_year,
                    values.total_previous_original,
                    values.total_previous_corrected,
                    values.total_current_month,
                    values.total_ytd_current,
                    values.total_ytd_prior_year,
                    report.title,
                    report.announcement_link,
                    report.excel_link,
                ]
            )

        self._format_header(data_sheet)

        for row in range(2, data_sheet.max_row + 1):
            status_cell = data_sheet.cell(row=row, column=8)

            if status_cell.value == "ERROR":
                status_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="F4CCCC",
                )

        self._set_widths(
            data_sheet,
            {
                "A": 8,
                "B": 16,
                "C": 32,
                "D": 18,
                "E": 18,
                "F": 14,
                "G": 16,
                "H": 12,
                "I": 45,
                "J": 55,
                "K": 24,
                "L": 24,
                "M": 24,
                "N": 22,
                "O": 22,
                "P": 24,
                "Q": 24,
                "R": 24,
                "S": 22,
                "T": 22,
                "U": 24,
                "V": 65,
                "W": 65,
                "X": 65,
            },
        )

        data_sheet.auto_filter.ref = (
            f"A1:X{max(1, data_sheet.max_row)}"
        )

        error_sheet = workbook.create_sheet(
            title="Errors"
        )
        error_sheet.sheet_view.rightToLeft = True

        error_headers = [
            "Symbol",
            "Report Month",
            "Publish Date",
            "Status",
            "Error",
            "Local File",
            "Report Title",
        ]

        error_sheet.append(error_headers)

        for item in results:
            if item.values.status == "OK":
                continue

            error_sheet.append(
                [
                    item.report.symbol,
                    item.values.report_month,
                    item.report.publish_date,
                    item.values.status,
                    item.values.note,
                    (
                        str(item.local_file)
                        if item.local_file
                        else ""
                    ),
                    item.report.title,
                ]
            )

        self._format_header(error_sheet)

        self._set_widths(
            error_sheet,
            {
                "A": 16,
                "B": 18,
                "C": 18,
                "D": 12,
                "E": 60,
                "F": 60,
                "G": 70,
            },
        )

        workbook.save(destination)

        logger.success(
            "Extracted-data workbook created: {}",
            destination,
        )

        return destination

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _format_header(worksheet: Any) -> None:
        """Apply a consistent header format."""

        fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[1].height = 35

    @staticmethod
    def _set_widths(
        worksheet: Any,
        widths: dict[str, float],
    ) -> None:
        """Set worksheet column widths."""

        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    # ------------------------------------------------------------------
    # Complete pipeline
    # ------------------------------------------------------------------

    def run(
        self,
        aa_tse_workbook: str | Path,
        date_start: str | None = None,
        date_end: str | None = None,
        target_period: str | None = None,
        max_pages: int | None = 50,
        download_original_excels: bool = True,
        overwrite_downloads: bool = False,
        consolidated_output_path: str | Path | None = None,
        extracted_output_path: str | Path | None = None,
    ) -> dict[str, Any]:
        """Execute the complete monthly-sales pipeline."""

        symbol_groups = self.read_aa_tse_symbols(
            aa_tse_workbook
        )

        all_reports = self.fetch_monthly_reports(
            date_start=date_start,
            date_end=date_end,
            max_pages=max_pages,
        )

        matched_reports = self.filter_reports_for_symbols(
            reports=all_reports,
            symbols=symbol_groups["all"],
            target_period=target_period,
        )

        consolidated_file = self.export_reports_to_excel(
            reports=matched_reports,
            output_path=consolidated_output_path,
        )

        download_result = self.download_report_excels(
            reports=matched_reports,
            overwrite=overwrite_downloads,
        )

        parsed_results = self.parse_downloaded_reports(
            report_files=download_result["report_files"],
        )

        extracted_file = self.export_parsed_values_to_excel(
            parsed_results=parsed_results,
            output_path=extracted_output_path,
        )

        parse_ok = sum(
            item.values.status == "OK"
            for item in parsed_results
        )

        parse_errors = len(parsed_results) - parse_ok

        return {
            "cement_symbols": symbol_groups["cement"],
            "mopfra_symbols": symbol_groups["mopfra"],
            "all_symbols": symbol_groups["all"],
            "all_reports": all_reports,
            "matched_reports": matched_reports,
            "consolidated_file": consolidated_file,
            "extracted_file": extracted_file,
            "download_result": download_result,
            "parsed_results": parsed_results,
            "parse_ok": parse_ok,
            "parse_errors": parse_errors,
        }