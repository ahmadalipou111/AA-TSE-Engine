from pathlib import Path
from openpyxl import load_workbook


WORKBOOK_PATH = Path("excel/TSE-Codal-Month-Sales-Extracted.xlsx")

MANUAL_SHEET = "Manual 1405 04 31"
AUTO_SHEET = "Auto 1405 04 31"

# Seven fields we are validating
FIELDS = {
    8: "Sales Last Year",          # H
    9: "Sales YTD",                # I
    10: "Sales Current Month",     # J
    11: "Sales Prior Month YTD",   # K
    12: "Export Last Year",        # L
    13: "Export YTD",              # M
    14: "Export Current Month",    # N
}

# Known exceptions. We still show them in the report,
# but they do not count as unexplained parser failures.
KNOWN_EXCEPTIONS = {
    "فن آوا": "IT format / prior-year fallback required",
    "فن‌آوا": "IT format / prior-year fallback required",
    "زفجر": "Agriculture report format",
}


def normalize_text(value):
    return (
        str(value or "")
        .replace("ي", "ی")
        .replace("ك", "ک")
        .replace("\u200c", " ")
        .replace("\u200f", "")
        .replace("\u200e", "")
        .replace("\xa0", " ")
        .strip()
    )


def normalize_number(value):
    """
    Excel may store the same number as int, float or string.
    Convert comparable numeric values to integers.

    Blank remains None.
    """

    if value is None:
        return None

    if isinstance(value, str):
        text = (
            value
            .replace(",", "")
            .replace("٬", "")
            .strip()
        )

        if text == "":
            return None

        try:
            return round(float(text))
        except ValueError:
            return value.strip()

    if isinstance(value, (int, float)):
        return round(value)

    return value


def get_symbol(ws, row):
    # Symbol is in column C
    return normalize_text(
        ws.cell(row=row, column=3).value
    )


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK_PATH}"
        )

    wb = load_workbook(
        WORKBOOK_PATH,
        data_only=False,
    )

    if MANUAL_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"Manual sheet not found: {MANUAL_SHEET}"
        )

    if AUTO_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"Auto sheet not found: {AUTO_SHEET}"
        )

    manual = wb[MANUAL_SHEET]
    auto = wb[AUTO_SHEET]

    symbol_count = 0
    company_exact_matches = 0
    company_differences = 0
    known_exception_count = 0

    field_match = 0
    field_difference = 0
    field_manual_blank = 0
    field_auto_blank = 0
    total_fields = 0

    differences = []
    exceptions = []

    max_row = max(
        manual.max_row,
        auto.max_row,
    )

    for row in range(1, max_row + 1):

        symbol = get_symbol(manual, row)

        if not symbol:
            continue

        if symbol.lower() == "symbol":
            continue

        if symbol == "نماد":
            continue

        symbol_count += 1

        normalized_symbol = normalize_text(symbol)

        known_reason = None

        for exception_symbol, reason in KNOWN_EXCEPTIONS.items():
            if normalize_text(exception_symbol) == normalized_symbol:
                known_reason = reason
                break

        company_has_difference = False
        company_field_differences = []

        for col, field_name in FIELDS.items():

            manual_raw = manual.cell(
                row=row,
                column=col,
            ).value

            auto_raw = auto.cell(
                row=row,
                column=col,
            ).value

            manual_value = normalize_number(
                manual_raw
            )

            auto_value = normalize_number(
                auto_raw
            )

            total_fields += 1

            if manual_value is None:
                field_manual_blank += 1

            if auto_value is None:
                field_auto_blank += 1

            if manual_value == auto_value:
                field_match += 1
                continue

            field_difference += 1
            company_has_difference = True

            difference_value = None

            if (
                isinstance(manual_value, (int, float))
                and isinstance(auto_value, (int, float))
            ):
                difference_value = (
                    auto_value - manual_value
                )

            item = {
                "row": row,
                "symbol": symbol,
                "field": field_name,
                "column": manual.cell(
                    row=row,
                    column=col,
                ).column_letter,
                "manual": manual_value,
                "auto": auto_value,
                "difference": difference_value,
            }

            company_field_differences.append(
                item
            )

        if company_has_difference:

            if known_reason:
                known_exception_count += 1

                exceptions.append({
                    "symbol": symbol,
                    "row": row,
                    "reason": known_reason,
                    "differences": company_field_differences,
                })

            else:
                company_differences += 1

                differences.extend(
                    company_field_differences
                )

        else:
            company_exact_matches += 1

    print()
    print("=" * 78)
    print("AA-TSE MONTHLY SALES VALIDATOR")
    print("=" * 78)

    print("Workbook :", WORKBOOK_PATH)
    print("Manual   :", MANUAL_SHEET)
    print("Auto     :", AUTO_SHEET)

    print()
    print("-" * 78)
    print("COMPANY SUMMARY")
    print("-" * 78)

    print(
        "Symbols checked             :",
        symbol_count,
    )

    print(
        "Companies exact match       :",
        company_exact_matches,
    )

    print(
        "Companies with differences  :",
        company_differences,
    )

    print(
        "Known exceptions            :",
        known_exception_count,
    )

    print()
    print("-" * 78)
    print("FIELD SUMMARY")
    print("-" * 78)

    print(
        "Fields compared             :",
        total_fields,
    )

    print(
        "Exact field matches         :",
        field_match,
    )

    print(
        "Field differences           :",
        field_difference,
    )

    print(
        "Manual blank fields         :",
        field_manual_blank,
    )

    print(
        "Auto blank fields           :",
        field_auto_blank,
    )

    if total_fields:
        match_percent = (
            field_match
            / total_fields
            * 100
        )

        print(
            "Exact match percentage      :",
            f"{match_percent:.2f}%",
        )

    print()
    print("=" * 78)
    print("UNEXPLAINED DIFFERENCES")
    print("=" * 78)

    if not differences:
        print(
            "NONE - all non-exception "
            "companies matched exactly."
        )
    else:
        for item in differences:
            print(
                f"{item['symbol']:12} | "
                f"Row {item['row']:>3} | "
                f"{item['column']} | "
                f"{item['field']:<24} | "
                f"Manual={item['manual']} | "
                f"Auto={item['auto']} | "
                f"Diff={item['difference']}"
            )

    print()
    print("=" * 78)
    print("KNOWN EXCEPTIONS")
    print("=" * 78)

    if not exceptions:
        print("NONE")
    else:
        for exception in exceptions:

            print()
            print(
                exception["symbol"],
                "|",
                exception["reason"],
            )

            for item in exception["differences"]:
                print(
                    f"   {item['column']} "
                    f"{item['field']} | "
                    f"Manual={item['manual']} | "
                    f"Auto={item['auto']}"
                )

    print()
    print("=" * 78)

    unexplained_company_count = (
        company_differences
    )

    if unexplained_company_count == 0:
        print(
            "VALIDATION RESULT: PASS"
        )
    else:
        print(
            "VALIDATION RESULT: CHECK REQUIRED"
        )

    print("=" * 78)
    print()


if __name__ == "__main__":
    main()