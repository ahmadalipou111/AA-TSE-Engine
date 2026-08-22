from copy import copy
from datetime import date, datetime
from email.utils import parsedate_to_datetime
from html import unescape
from html.parser import HTMLParser
import inspect
import json
import os
from pathlib import Path
import re
import shutil
import time
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

from api.codal_api import CodalAPI
from services.monthly_sales_html_parser import MonthlySalesHtmlParser


# ============================================================
# CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "excel" / "TSE-Codal-Month-Sales-Extracted.xlsx"
MASTER_PATH = BASE_DIR / "excel" / "AAI-TSE-Master.xlsx"

# Existing validated sheet used only as the layout/template
TEMPLATE_SHEET = "Manual 1405 04 31"

CATEGORY = 3

# BRSAPI documents a ceiling of 500 requests per five minutes.  A 0.75-second
# gap is deliberately more conservative than the theoretical 0.60-second gap.
BRSAPI_MIN_REQUEST_INTERVAL = 0.75
BRSAPI_MAX_429_RETRIES = 8
BRSAPI_INITIAL_BACKOFF = 5.0
BRSAPI_MAX_BACKOFF = 120.0

# CODAL report HTML can intermittently time out even after its announcement was
# found successfully.  Retry only that exact URL; never restart announcement
# pagination or the surrounding company batch.
CODAL_HTML_MAX_TIMEOUT_RETRIES = 3
CODAL_HTML_INITIAL_BACKOFF = 2.0
CODAL_HTML_MAX_BACKOFF = 15.0

# Recovery queries are symbol-filtered and normally finish on page one.  Keep a
# hard safety ceiling in case an upstream API ignores the symbol parameter.
TARGETED_RECOVERY_MAX_PAGES = 20
# An unfiltered Category 2 request is deliberately the last resort.  Keep its
# ceiling much lower than a company-targeted query so a broken upstream filter
# cannot turn one missing company into a long market-wide scan.
UNFILTERED_CATEGORY2_MAX_PAGES = 3
# Start the market-wide Category 2 fallback with small chronological windows.
# This avoids reading several doomed pages for a whole month before splitting.
UNFILTERED_CATEGORY2_WINDOW_DAYS = 3
REVISION_CHAIN_MAX_HOPS = 20
# Historical backfill searches the normal next-month publication window plus
# one additional late-publication month.  A miss then remains blank and the
# run advances to the next company instead of scanning distant dates.
LATE_PUBLICATION_RECOVERY_MONTHS = 1
HISTORICAL_TARGETED_MAX_PAGES = 2

# Prefer one symbol-filtered query path when a period has only a small number
# of gaps.  Either limit is sufficient: at most 10 rows, or less than 20% of
# the companies eligible for that period.  Larger gaps keep the batch path.
TARGETED_ONLY_MAX_ROWS = 10
TARGETED_ONLY_MAX_RATIO = 0.20

# Historical range we ultimately want to backfill
HISTORY_START_PERIOD = "1404/01/31"
HISTORY_END_PERIOD = "1405/04/31"

OUTPUT_HTML_DIR = Path("output/monthly_html")
OUTPUT_HTML_DIR.mkdir(parents=True, exist_ok=True)

LOG_SHEET = "_Report_Log"
SALES_TREND_SHEET = "Sales Trend"
SALES_TREND_AUDIT_SHEET = "_Sales_Trend_Audit"
SALES_TREND_MONTH_AUDIT_SHEET = "_Sales_Trend_Month_Audit"
SALES_TREND_TEMPLATE_PERIOD = "1405/02/31"
NEGATIVE_PERCENT_FONT_COLOR = "FFFF0000"
SALES_TREND_BAND_FILL_COLOR = "FFD9E1F2"
SALES_TREND_WHITE_FILL_COLOR = "FFFFFFFF"
MODE5_FLAG = "Fiscal Year Changed - Mode 5"
MODE5_COMMENT_PREFIX = "AAI-TSE MODE5 OVERRIDE\n"
MODE5_FILL_COLOR = "FFFFFF00"

FORECAST_COLUMNS_BEFORE_SYMBOL = (
    "FY_Sales_Forecast",
    "Month_of_Last_Sales_Rec",
    "Forecast_Confidence",
    "Forecast_Method",
)
FORECAST_DIAGNOSTIC_COLUMNS = (
    "Forecast_Fiscal_Month",
    "Forecast_History_Depth",
    "Manual_vs_Model_Gap",
    "Calendar_Regime_Flag",
)

AUTO_SCHEMA = (
    ("Company_ID", None),
    ("Company_Name", None),
    ("Symbol", None),
    ("Fiscal_Year_End", None),
    ("Report_Month", None),
    ("Reporting_Period_Months", None),
    ("Sales_Prior_Year_YTD", "sales_last_year"),
    ("Sales_YTD", "sales_ytd"),
    ("Sales_Month", "sales_month"),
    ("Sales_Prior_Month_YTD", "sales_prior_month_ytd"),
    ("Export_Prior_Year_YTD", "export_last_year"),
    ("Export_YTD", "export_ytd"),
    ("Export_Month", "export_month"),
)
AUTO_HEADERS = tuple(header for header, _ in AUTO_SCHEMA)
SALES_HEADERS = {
    parser_key: header
    for header, parser_key in AUTO_SCHEMA
    if parser_key is not None
}

# The workbook produced before AUTO_SCHEMA was introduced always stored the
# seven parser values in H:N.  Keep this mapping in one migration-only place;
# normal reads and all new writes remain header based.
LEGACY_SALES_COLUMNS = {
    "Sales_Prior_Year_YTD": 8,
    "Sales_YTD": 9,
    "Sales_Month": 10,
    "Sales_Prior_Month_YTD": 11,
    "Export_Prior_Year_YTD": 12,
    "Export_YTD": 13,
    "Export_Month": 14,
}


# ============================================================
# TEXT HELPERS
# ============================================================

def normalize_digits(value):
    return (
        str(value or "")
        .replace("۰", "0")
        .replace("۱", "1")
        .replace("۲", "2")
        .replace("۳", "3")
        .replace("۴", "4")
        .replace("۵", "5")
        .replace("۶", "6")
        .replace("۷", "7")
        .replace("۸", "8")
        .replace("۹", "9")
        .replace("٠", "0")
        .replace("١", "1")
        .replace("٢", "2")
        .replace("٣", "3")
        .replace("٤", "4")
        .replace("٥", "5")
        .replace("٦", "6")
        .replace("٧", "7")
        .replace("٨", "8")
        .replace("٩", "9")
    )


def normalize_text(value):
    return (
        str(value or "")
        .replace("ي", "ی")
        .replace("ك", "ک")
        .replace("\u200c", " ")
        .replace("\u200f", "")
        .replace("\u200e", "")
        .replace("\xa0", " ")
        .strip()
    )
# ============================================================
# PERIOD HELPERS
# ============================================================

def parse_period(period):
    """
    Convert a period such as 1404/01/31
    into integer year, month, day.
    """

    normalized = (
        str(period)
        .strip()
        .replace("-", "/")
    )

    parts = normalized.split("/")

    if len(parts) != 3:
        raise ValueError(
            f"Invalid period format: {period}"
        )

    year = int(parts[0])
    month = int(parts[1])
    day = int(parts[2])

    return year, month, day


def format_period(year, month, day):
    return f"{year:04d}/{month:02d}/{day:02d}"


def period_sheet_name(period):
    year, month, day = parse_period(period)

    return (
        f"Auto "
        f"{year:04d} "
        f"{month:02d} "
        f"{day:02d}"
    )


def period_html_suffix(period):
    year, month, day = parse_period(period)

    return (
        f"{year:04d}_"
        f"{month:02d}_"
        f"{day:02d}"
    )


def jalali_month_days(year, month):
    """
    Enough for CODAL publication search windows.

    Months 1-6  -> 31 days
    Months 7-11 -> 30 days
    Month 12    -> 29 days, or 30 in a Jalali leap year
    """

    if 1 <= month <= 6:
        return 31

    if 7 <= month <= 11:
        return 30

    if month == 12:
        # In the arithmetic Persian calendar, the leap day belongs to the
        # year immediately preceding an arithmetic leap-year boundary.
        next_year = year + 1
        ep_base = next_year - (474 if next_year >= 0 else 473)
        ep_year = 474 + (ep_base % 2820)
        is_leap = ((ep_year + 38) * 682) % 2816 < 682
        return 30 if is_leap else 29

    raise ValueError(
        f"Invalid Jalali month: {month}"
    )


def next_jalali_month(year, month):
    if month == 12:
        return year + 1, 1

    return year, month + 1


def gregorian_to_jalali(year, month, day):
    """Convert a Gregorian date to Jalali without external dependencies."""
    gregorian_days_in_month = [
        31, 28, 31, 30, 31, 30,
        31, 31, 30, 31, 30, 31,
    ]
    jalali_days_in_month = [
        31, 31, 31, 31, 31, 31,
        30, 30, 30, 30, 30, 29,
    ]

    gregorian_year = year - 1600
    gregorian_month = month - 1
    gregorian_day = day - 1

    day_number = (
        365 * gregorian_year
        + (gregorian_year + 3) // 4
        - (gregorian_year + 99) // 100
        + (gregorian_year + 399) // 400
    )
    for month_index in range(gregorian_month):
        day_number += gregorian_days_in_month[month_index]

    if (
        gregorian_month > 1
        and (
            gregorian_year % 4 == 0
            and (
                gregorian_year % 100 != 0
                or gregorian_year % 400 == 0
            )
        )
    ):
        day_number += 1

    day_number += gregorian_day
    jalali_day_number = day_number - 79

    jalali_cycle = jalali_day_number // 12053
    jalali_day_number %= 12053

    jalali_year = 979 + 33 * jalali_cycle + 4 * (jalali_day_number // 1461)
    jalali_day_number %= 1461

    if jalali_day_number >= 366:
        jalali_year += (jalali_day_number - 1) // 365
        jalali_day_number = (jalali_day_number - 1) % 365

    jalali_month = 0
    while (
        jalali_month < 11
        and jalali_day_number >= jalali_days_in_month[jalali_month]
    ):
        jalali_day_number -= jalali_days_in_month[jalali_month]
        jalali_month += 1

    return jalali_year, jalali_month + 1, jalali_day_number + 1


def title_contains_period(title, target_period):
    """Match a Jalali period even when CODAL omits leading zeroes."""
    target = parse_period(target_period)
    normalized_title = normalize_digits(normalize_text(title))
    for year, month, day in re.findall(
        r"(?<!\d)(1[34]\d{2})\s*[/-]\s*(\d{1,2})\s*[/-]\s*(\d{1,2})(?!\d)",
        normalized_title,
    ):
        if (int(year), int(month), int(day)) == target:
            return True
    return False


def prior_year_period(period):
    """Return the same reporting month/day in the preceding Jalali year."""
    year, month, day = parse_period(period)
    return format_period(year - 1, month, day)


def today_jalali():
    """Return today's local date as a Jalali (year, month, day) tuple."""
    today = date.today()
    return gregorian_to_jalali(today.year, today.month, today.day)


def current_jalali_month_period():
    """Return the current Jalali month-end period suggested by Live Mode."""
    year, month, _day = today_jalali()
    return format_period(year, month, jalali_month_days(year, month))


def live_publish_range(period):
    """Return the normal publication window clipped to today, or no window."""
    date_start, normal_end = publish_range_for_period(period)
    year, month, day = today_jalali()
    today_end = f"{year:04d}-{month:02d}-{day:02d}"
    if date_start > today_end:
        return None
    return date_start, min(normal_end, today_end)


def live_report_disposition(has_valid_existing, candidate, previous_link=""):
    """Classify a Live row without mutating its prior workbook values."""
    if has_valid_existing:
        if (
            candidate is None
            or not previous_link
            or candidate.get("link") == previous_link
        ):
            return "EXISTING"
        return "REVISED"
    return "NEW" if candidate is not None else "PENDING"


def should_recover_missing_report(report, api, live_mode):
    """Historical modes may recover misses; Live Mode never does."""
    return report is None and api is not None and not live_mode

def generate_periods(start_period, end_period):
    """
    Generate month-end periods inclusively.

    Example:
    1404/01/31 -> 1404/04/31
    """

    start_year, start_month, _ = parse_period(start_period)
    end_year, end_month, _ = parse_period(end_period)

    periods = []

    year = start_year
    month = start_month

    while (year, month) <= (end_year, end_month):
        day = jalali_month_days(year, month)

        periods.append(
            format_period(
                year,
                month,
                day,
            )
        )

        year, month = next_jalali_month(
            year,
            month,
        )

    return periods


def ignores_company_start_period(target_period):
    """Return True for backfill periods predating the current registry history.

    Company Start_Period records when a Company_ID entered AA-TSE; it is not
    evidence that the company had no earlier CODAL reports.  Only periods before
    the configured history boundary bypass that registry eligibility gate, so
    normal/current-period behavior remains unchanged.
    """
    return parse_period(target_period) < parse_period(HISTORY_START_PERIOD)

def publish_range_for_period(period):
    """
    Monthly CODAL reports are normally published
    during the following Jalali month.

    Example:
    1405/04/31 -> search 1405/05/01 through 1405/05/31
    """

    year, month, _ = parse_period(period)

    publish_year, publish_month = (
        next_jalali_month(year, month)
    )

    last_day = jalali_month_days(
        publish_year,
        publish_month,
    )

    date_start = (
        f"{publish_year:04d}-"
        f"{publish_month:02d}-01"
    )

    date_end = (
        f"{publish_year:04d}-"
        f"{publish_month:02d}-"
        f"{last_day:02d}"
    )

    return date_start, date_end


def shared_publish_range(periods):
    """
    Build one announcement range for all selected periods.

    It begins at the normal publication window for the first period and ends
    today, so late revisions are included and BRSAPI never receives a future
    Jalali date. BRSAPI date parameters use YYYY-MM-DD with zero padding.
    """
    first_start, _ = publish_range_for_period(periods[0])
    today_year, today_month, today_day = today_jalali()
    today_end = f"{today_year:04d}-{today_month:02d}-{today_day:02d}"

    if first_start > today_end:
        raise ValueError(
            "The selected period's publication window has not started yet "
            f"(start {first_start}, today {today_end})."
        )

    return first_start, today_end


def missing_period_publish_ranges(periods):
    """Return the smallest normal publish windows needed for backfill.

    Existing Company+Period data is deliberately excluded before this helper
    is called.  Adjacent windows are merged, while gaps remain separate so a
    sparse backfill does not fetch announcements for already-complete months.
    This is base/backfill behaviour only; a future revision sweep may search
    beyond these normal publication windows.
    """
    today_year, today_month, today_day = today_jalali()
    today_end = f"{today_year:04d}-{today_month:02d}-{today_day:02d}"
    windows = []

    for period in sorted(set(periods), key=parse_period):
        date_start, date_end = publish_range_for_period(period)
        if date_start > today_end:
            continue
        windows.append((date_start, min(date_end, today_end)))

    merged = []
    for date_start, date_end in windows:
        if not merged:
            merged.append([date_start, date_end])
            continue

        previous_end = parse_period(merged[-1][1])
        current_start = parse_period(date_start)
        previous_year, previous_month, previous_day = previous_end
        is_same_month = current_start[:2] == previous_end[:2]
        is_next_month = (
            current_start[2] == 1
            and current_start[:2] == next_jalali_month(
                previous_year, previous_month
            )
            and previous_day == jalali_month_days(
                previous_year, previous_month
            )
        )

        if is_same_month or is_next_month:
            merged[-1][1] = max(merged[-1][1], date_end)
        else:
            merged.append([date_start, date_end])

    return [tuple(window) for window in merged]


def monthly_publish_windows(date_start, date_end):
    """Split an inclusive Jalali date range at month boundaries only."""
    start_year, start_month, start_day = parse_period(date_start)
    end_year, end_month, end_day = parse_period(date_end)

    if (start_year, start_month, start_day) > (end_year, end_month, end_day):
        raise ValueError(
            f"Invalid announcement range: {date_start} through {date_end}."
        )

    year, month = start_year, start_month
    windows = []

    while (year, month) <= (end_year, end_month):
        window_start_day = start_day if (year, month) == (
            start_year,
            start_month,
        ) else 1
        window_end_day = end_day if (year, month) == (
            end_year,
            end_month,
        ) else jalali_month_days(year, month)

        windows.append((
            f"{year:04d}-{month:02d}-{window_start_day:02d}",
            f"{year:04d}-{month:02d}-{window_end_day:02d}",
        ))

        year, month = next_jalali_month(year, month)

    return windows


def split_publish_window(date_start, date_end):
    """Bisect one same-month inclusive Jalali window without losing a date."""
    start_year, start_month, start_day = parse_period(date_start)
    end_year, end_month, end_day = parse_period(date_end)
    if (start_year, start_month) != (end_year, end_month):
        raise ValueError("A pagination split must stay inside one Jalali month.")
    if start_day >= end_day:
        return None

    midpoint = (start_day + end_day) // 2
    return (
        (date_start, f"{start_year:04d}-{start_month:02d}-{midpoint:02d}"),
        (f"{start_year:04d}-{start_month:02d}-{midpoint + 1:02d}", date_end),
    )


def chunk_publish_windows(date_start, date_end, window_days):
    """Split an inclusive Jalali range into small chronological windows."""
    if window_days < 1:
        raise ValueError("window_days must be positive.")

    result = []
    for month_start, month_end in monthly_publish_windows(date_start, date_end):
        year, month, start_day = parse_period(month_start)
        _end_year, _end_month, end_day = parse_period(month_end)
        day = start_day
        while day <= end_day:
            chunk_end = min(day + window_days - 1, end_day)
            result.append((
                f"{year:04d}-{month:02d}-{day:02d}",
                f"{year:04d}-{month:02d}-{chunk_end:02d}",
            ))
            day = chunk_end + 1
    return result


def announcement_identity(report):
    """Return a stable, hashable identity used to remove API duplicates."""
    for field in (
        "tracing_no",
        "tracingNo",
        "tracking_number",
        "trackingNumber",
        "id",
    ):
        value = report.get(field)
        if value not in (None, ""):
            return field, str(value)

    return (
        "composite",
        str(report.get("link", "")),
        normalize_digits(report.get("date_publish", "")),
        str(report.get("time_publish", "")),
        normalize_text(report.get("l18", "")),
        normalize_digits(normalize_text(report.get("title", ""))),
    )

# ============================================================
# BRSAPI FETCH
# ============================================================

def _exception_chain(error):
    """Yield an exception and its wrapped causes without looping forever."""
    seen = set()
    current = error
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        yield current
        current = current.__cause__ or current.__context__


def _response_from_exception(error):
    """Find a requests-like response even when CodalAPI wraps the error."""
    for item in _exception_chain(error):
        response = getattr(item, "response", None)
        if response is not None:
            return response
        for arg in getattr(item, "args", ()):
            response = getattr(arg, "response", None)
            if response is not None:
                return response
    return None


def _is_rate_limit_error(error):
    response = _response_from_exception(error)
    if response is not None and getattr(response, "status_code", None) == 429:
        return True

    # CodalAPI may replace requests.HTTPError with RuntimeError and retain only
    # its message.  Match the status phrase/code, but avoid treating arbitrary
    # numbers containing 429 as an HTTP rate-limit response.
    message = " ".join(str(item) for item in _exception_chain(error))
    return bool(
        re.search(r"(?:HTTP(?:Error)?\s*)?429\b", message, re.IGNORECASE)
        or "too many requests" in message.lower()
    )


def _is_bad_request_error(error):
    """Return True when an exception chain represents HTTP 400."""
    response = _response_from_exception(error)
    if response is not None and getattr(response, "status_code", None) == 400:
        return True
    message = " ".join(str(item) for item in _exception_chain(error))
    return bool(
        re.search(r"(?:HTTP(?:Error)?\s*)?400\b", message, re.IGNORECASE)
        or "bad request" in message.lower()
    )


def _retry_after_seconds(error):
    """Return Retry-After as seconds (delta or HTTP date), when exposed."""
    response = _response_from_exception(error)
    headers = getattr(response, "headers", {}) if response is not None else {}
    value = headers.get("Retry-After") if hasattr(headers, "get") else None
    if value is None:
        return None

    try:
        return max(0.0, float(value))
    except (TypeError, ValueError):
        try:
            retry_at = parsedate_to_datetime(str(value))
            now = datetime.now(retry_at.tzinfo) if retry_at.tzinfo else datetime.now()
            return max(0.0, (retry_at - now).total_seconds())
        except (TypeError, ValueError, OverflowError):
            return None


def _get_announcements_with_retry(api, request_kwargs, rate_state):
    """Rate-limit one API call and handle narrow, documented API retries."""
    request_kwargs = dict(request_kwargs)
    retries = 0
    leap_day_fallback_used = False
    while True:
        elapsed = time.monotonic() - rate_state["last_request_at"]
        if elapsed < BRSAPI_MIN_REQUEST_INTERVAL:
            time.sleep(BRSAPI_MIN_REQUEST_INTERVAL - elapsed)

        # Record the attempt before calling so failed requests are also paced.
        rate_state["last_request_at"] = time.monotonic()
        try:
            return api.get_announcements(**request_kwargs)
        except Exception as error:
            date_end = normalize_digits(request_kwargs.get("date_end", ""))
            if (
                not leap_day_fallback_used
                and _is_bad_request_error(error)
                and re.fullmatch(r"\d{4}-12-30", date_end)
            ):
                fallback_end = date_end[:-2] + "29"
                date_start = normalize_digits(
                    request_kwargs.get("date_start", "")
                )
                if not date_start or date_start <= fallback_end:
                    request_kwargs["date_end"] = fallback_end
                    leap_day_fallback_used = True
                    print(
                        "  WARNING: BRSAPI rejected valid Jalali leap day "
                        f"date_end={date_end}; retrying this API request only "
                        f"with date_end={fallback_end}. Report periods and "
                        "workbook dates remain unchanged."
                    )
                    continue

            if not _is_rate_limit_error(error) or retries >= BRSAPI_MAX_429_RETRIES:
                raise

            retry_after = _retry_after_seconds(error)
            backoff = min(
                BRSAPI_INITIAL_BACKOFF * (2 ** retries),
                BRSAPI_MAX_BACKOFF,
            )
            wait_seconds = retry_after if retry_after is not None else backoff
            # Retry-After is authoritative, but retain the normal request gap.
            wait_seconds = max(wait_seconds, BRSAPI_MIN_REQUEST_INTERVAL)
            retries += 1
            print(
                f"  BRSAPI rate limit (429); waiting {wait_seconds:.1f}s "
                f"before retry {retries}/{BRSAPI_MAX_429_RETRIES} "
                "of the same page."
            )
            time.sleep(wait_seconds)

def fetch_all_announcements(
    api,
    date_start,
    date_end,
    company_count,
    rate_state=None,
):
    """
    Fetch all monthly-sales announcements for the publish-date range.

    Large ranges are split into Jalali calendar months.  The per-window page
    limit is dynamic (twice the Sales-enabled company count, with a floor of
    20).  A window that reaches the limit is bisected and retried recursively,
    preserving complete coverage without making every normal query small.
    """

    all_reports = []
    seen_reports = set()
    if rate_state is None:
        rate_state = {"last_request_at": float("-inf")}

    print("=" * 70)
    print("FETCHING CODAL ANNOUNCEMENTS")
    print("=" * 70)

    try:
        company_count = int(company_count)
    except (TypeError, ValueError):
        raise ValueError("company_count must be a positive integer.")
    if company_count < 1:
        raise ValueError("company_count must be a positive integer.")
    pagination_limit = max(20, 2 * company_count)

    windows = monthly_publish_windows(date_start, date_end)
    pending_windows = list(windows)
    completed_windows = 0

    while pending_windows:
        window_start, window_end = pending_windows.pop(0)
        print(
            f"Window {completed_windows + 1}: {window_start} through {window_end} "
            f"(pagination limit {pagination_limit})"
        )
        page = 1
        previous_page_signature = None
        window_reports = []
        window_seen = set()
        overflowed = False

        while True:
            data = _get_announcements_with_retry(
                api,
                {
                    "category": CATEGORY,
                    "date_start": window_start,
                    "date_end": window_end,
                    "page": page,
                },
                rate_state,
            )

            reports = data.get("announcement", []) or []
            print(f"  Page {page}: {len(reports)} announcement(s)")

            if not reports:
                break

            page_signature = tuple(announcement_identity(report) for report in reports)
            if page_signature == previous_page_signature:
                print("  Repeated page detected; this window is complete.")
                break
            previous_page_signature = page_signature

            for report in reports:
                identity = announcement_identity(report)
                if identity not in window_seen:
                    window_seen.add(identity)
                    window_reports.append(report)

            count_page = (
                data.get("count_page")
                or data.get("countPage")
                or data.get("total_pages")
                or data.get("totalPages")
            )

            if count_page is not None:
                try:
                    if page >= int(count_page):
                        break
                except (TypeError, ValueError):
                    pass

            if page >= pagination_limit:
                child_windows = split_publish_window(window_start, window_end)
                if child_windows is None:
                    raise RuntimeError(
                        "Pagination safety limit reached for the smallest "
                        f"possible window ({window_start})."
                    )
                print(
                    "  Pagination limit reached; splitting into "
                    f"{child_windows[0][0]} through {child_windows[0][1]} and "
                    f"{child_windows[1][0]} through {child_windows[1][1]}."
                )
                pending_windows[0:0] = child_windows
                overflowed = True
                break

            page += 1

        if overflowed:
            continue

        completed_windows += 1
        for report in window_reports:
            identity = announcement_identity(report)
            if identity not in seen_reports:
                seen_reports.add(identity)
                all_reports.append(report)

    print()
    print(
        "Total announcements fetched:",
        len(all_reports),
    )

    return all_reports


# ============================================================
# REPORT SELECTION
# ============================================================

def report_matches_symbol_and_period(
    report,
    symbol,
    target_period,
):
    report_symbol = (
        normalize_text(report.get("l18", ""))
        .replace("\u200c", "")
        .replace(" ", "")
    )

    target_symbol = (
        normalize_text(symbol)
        .replace("\u200c", "")
        .replace(" ", "")
    )

    title = normalize_digits(
        normalize_text(report.get("title", ""))
    )

    return (
        report_symbol == target_symbol
        and title_contains_period(title, target_period)
        and is_monthly_activity_report(report)
    )


def is_monthly_activity_report(report):
    """Positively identify CODAL monthly-activity announcements by title."""
    title = normalize_text(report.get("title", ""))
    compact_title = re.sub(r"[\s\u200c\-_]+", "", title).casefold()
    required_phrase = "\u06af\u0632\u0627\u0631\u0634\u0641\u0639\u0627\u0644\u06cc\u062a\u0645\u0627\u0647\u0627\u0646\u0647"
    return required_phrase in compact_title

def report_matches_company_and_period(
    report,
    company_name,
    target_period,
):
    """
    Fallback matcher:
    use Company Name only when Symbol matching finds nothing.
    """

    if not company_name:
        return False

    target_company = normalize_text(company_name)

    title = normalize_digits(
        normalize_text(report.get("title", ""))
    )

    record_text = normalize_text(
        str(report)
    )

    return (
        target_company in record_text
        and title_contains_period(title, target_period)
        and is_monthly_activity_report(report)
    )

def select_latest_report(
    all_reports,
    symbol,
    company_name=None,
    target_period=None,
):
    """
    Find all reports for symbol + target period and select
    the latest published one.

    This automatically handles revised reports when more than
    one announcement exists for the same period.
    """
        # CODAL symbol aliases

    candidates = [
    report
    for report in all_reports
    if report_matches_symbol_and_period(
        report,
        symbol,
        target_period,
    )
]

    used_company_fallback = False
    if not candidates and company_name:
        candidates = [
        report
        for report in all_reports
        if report_matches_company_and_period(
            report,
            company_name,
            target_period,
            )
        ]
        used_company_fallback = bool(candidates)

    if used_company_fallback:
        print(
            f"  COMPANY NAME FALLBACK: "
            f"{symbol} -> {company_name}"
        )

    if not candidates:
        return None, 0

    candidates.sort(
        key=lambda report: (
            normalize_digits(
                report.get("date_publish", "")
            ),
            str(
                report.get("time_publish", "")
            ),
        )
    )

    return candidates[-1], len(candidates)


def missing_report_recovery_range(target_period, include_normal_window=False):
    """Return a target-relative, bounded late-publication recovery range.

    In BATCH mode the normal next-month publication window has already been
    fetched, so recovery starts after it.  In TARGETED_ONLY mode there was no
    general fetch, so the symbol-filtered search may include the normal window.
    Historical recovery is capped to a small number of months after that normal
    window; it must not scan from an old target period through today's date.
    """
    normal_start, normal_end = publish_range_for_period(target_period)
    normal_year, normal_month, _ = parse_period(normal_end)
    end_year, end_month = normal_year, normal_month
    for _ in range(LATE_PUBLICATION_RECOVERY_MONTHS):
        end_year, end_month = next_jalali_month(end_year, end_month)

    bounded_end = (
        f"{end_year:04d}-{end_month:02d}-"
        f"{jalali_month_days(end_year, end_month):02d}"
    )
    today_year, today_month, today_day = today_jalali()
    today_end = f"{today_year:04d}-{today_month:02d}-{today_day:02d}"
    date_end = min(bounded_end, today_end)

    if include_normal_window:
        date_start = normal_start
    else:
        recovery_year, recovery_month = next_jalali_month(
            normal_year, normal_month
        )
        date_start = f"{recovery_year:04d}-{recovery_month:02d}-01"

    if date_start > date_end:
        return None
    return date_start, date_end


def recover_missing_report(
    api,
    symbol,
    company_name,
    target_period,
    rate_state,
    recovery_reports_cache,
    include_normal_window=False,
):
    """Find one missing Company+Period using only symbol-filtered API calls.

    Late publication months are searched chronologically after the normal
    batch window.  Every report is checked immediately and the function returns
    on the first valid monthly report for the target period.  Any newer version
    is resolved later through the explicit revision link in that report's HTML,
    rather than by continuing the announcement search.  This function never
    downloads category=3 results for unrelated companies.
    """
    recovery_range = missing_report_recovery_range(
        target_period,
        include_normal_window=include_normal_window,
    )
    if recovery_range is None:
        return None, 0

    cache_key = (normalize_text(symbol), target_period)
    if cache_key in recovery_reports_cache:
        print("  MISSING REPORT RECOVERY: using cached targeted result")
        return recovery_reports_cache[cache_key]

    date_start, date_end = recovery_range
    print(
        "  MISSING REPORT RECOVERY: targeted search | "
        f"symbol={symbol} | period={target_period}"
    )
    for category in (CATEGORY, 2):
        if category == 2:
            print(
                "  MISSING REPORT RECOVERY: no valid category=3 report; "
                "using targeted category=2 fallback"
            )
        for window_start, window_end in monthly_publish_windows(date_start, date_end):
            print(f"    Publish window: {window_start} through {window_end}")
            result = _fetch_targeted_window_report(
                api, symbol, company_name, target_period, window_start,
                window_end, category, rate_state,
            )
            if result[0] is not None:
                # Missing-report recovery historically reports one selected
                # announcement even when a page contains several revisions.
                result = (result[0], 1)
                recovery_reports_cache[cache_key] = result
                print(
                    f"      Valid target report found in category={category}; "
                    "recovery search stopped."
                )
                return result

    result = (None, 0)
    recovery_reports_cache[cache_key] = result
    return result


def fetch_targeted_period_reports(
    api,
    symbol,
    company_name,
    target_period,
    rate_state,
    targeted_reports_cache,
):
    """Return the latest report from one symbol-filtered period search.

    Unlike missing-report recovery, this scans the complete normal publication
    window so ``select_latest_report`` can preserve the existing revision
    selection behaviour when several announcements exist.  Results (including
    misses) are cached per symbol and period; unrelated companies are never
    fetched.
    """
    cache_key = (normalize_text(symbol), target_period)
    if cache_key in targeted_reports_cache:
        print("  PRIOR-YEAR FALLBACK: using cached targeted result")
        return targeted_reports_cache[cache_key]

    date_start, date_end = publish_range_for_period(target_period)
    print(
        "  PRIOR-YEAR FALLBACK: targeted search | "
        f"symbol={symbol} | period={target_period} | "
        f"{date_start} through {date_end}"
    )
    for category in (CATEGORY, 2):
        if category == 2:
            print(
                "  PRIOR-YEAR FALLBACK: no valid category=3 report; "
                "using targeted category=2 fallback"
            )
        result = _fetch_targeted_window_report(
            api, symbol, company_name, target_period, date_start, date_end,
            category, rate_state,
        )
        if result[0] is not None:
            targeted_reports_cache[cache_key] = result
            print(
                f"    Valid prior-year report found in category={category}; "
                "targeted search stopped."
            )
            return result

    result = (None, 0)
    targeted_reports_cache[cache_key] = result
    return result


def _fetch_targeted_window_report(
    api, symbol, company_name, target_period, date_start, date_end, category,
    rate_state,
):
    """Search one bounded window: symbol, then API company-name filter.

    ``CodalAPI`` versions have used either ``company`` or ``company_name`` for
    the company-name keyword.  The installed method signature decides which
    spelling is sent.  Only Category 2 retains a small unfiltered last resort.
    """
    result = _fetch_filtered_window_report(
        api, symbol, company_name, target_period, date_start, date_end,
        category, "symbol", symbol, rate_state,
    )
    if result[0] is not None:
        return result

    company_parameter = _company_filter_parameter(api)
    if company_name and company_parameter:
        print(
            f"      Category {category} symbol-filtered search found no valid "
            f"report; retrying with {company_parameter}={company_name}."
        )
        result = _fetch_filtered_window_report(
            api, symbol, company_name, target_period, date_start, date_end,
            category, company_parameter, company_name, rate_state,
        )
        if result[0] is not None:
            return result
    elif company_name:
        print(
            "      CodalAPI.get_announcements exposes neither 'company' nor "
            "'company_name'; company-targeted API recovery is unavailable."
        )

    if category == 2 and ignores_company_start_period(target_period):
        print(
            "      Historical backfill: filtered category 2 searches found no "
            "valid report; market-wide fallback skipped."
        )
        return None, 0

    if category == 2:
        print(
            "      Category 2 targeted symbol/company searches found no valid "
            "report; using the limited unfiltered last resort."
        )
        return _fetch_unfiltered_category2_window_report(
            api, symbol, company_name, target_period, date_start, date_end,
            rate_state,
        )
    return None, 0


def _company_filter_parameter(api):
    """Return the company-name keyword accepted by get_announcements."""
    try:
        parameters = inspect.signature(api.get_announcements).parameters
    except (TypeError, ValueError):
        return None
    if "company" in parameters:
        return "company"
    if "company_name" in parameters:
        return "company_name"
    # Test doubles and forwarding adapters often expose only **kwargs.  The
    # AA-TSE CodalAPI convention is ``company`` in that case.
    if any(item.kind == inspect.Parameter.VAR_KEYWORD for item in parameters.values()):
        return "company"
    return None


def _fetch_filtered_window_report(
    api, symbol, company_name, target_period, date_start, date_end, category,
    filter_name, filter_value, rate_state,
):
    """Page through one API-targeted symbol or company-name query."""
    seen = set()
    previous_page_signature = None
    page_limit = (
        HISTORICAL_TARGETED_MAX_PAGES
        if ignores_company_start_period(target_period)
        else TARGETED_RECOVERY_MAX_PAGES
    )
    for page in range(1, page_limit + 1):
        request = {
            "category": category,
            "date_start": date_start,
            "date_end": date_end,
            "page": page,
            filter_name: filter_value,
        }
        data = _get_announcements_with_retry(
            api, request, rate_state,
        )
        page_reports = data.get("announcement", []) or []
        print(
            f"      Category {category}, page {page}: "
            f"{len(page_reports)} {filter_name}-filtered result(s)"
        )
        if not page_reports:
            break

        page_signature = tuple(
            announcement_identity(item) for item in page_reports
        )
        if page_signature == previous_page_signature:
            print("      Repeated page detected; targeted window is complete.")
            break
        previous_page_signature = page_signature

        unique_page_reports = []
        for report in page_reports:
            identity = announcement_identity(report)
            if identity not in seen:
                seen.add(identity)
                unique_page_reports.append(report)

        # The explicit HTML revision chain resolves newer versions, so once a
        # correct symbol+period report is present there is no reason to scan
        # further announcement pages.  If the current page contains several
        # candidates, retain the latest one from that page.
        # Require the standard monthly-report title in every category so an
        # exact symbol/company + period match cannot select another report type.
        result = select_latest_report(
            unique_page_reports,
            symbol,
            company_name,
            target_period,
        )
        if result[0] is not None:
            return result

        count_page = (
            data.get("count_page") or data.get("countPage")
            or data.get("total_pages") or data.get("totalPages")
        )
        try:
            if count_page is not None and page >= int(count_page):
                break
        except (TypeError, ValueError):
            pass
    else:
        print("      WARNING: targeted pagination safety limit reached.")
    return None, 0


def _fetch_unfiltered_category2_window_report(
    api, symbol, company_name, target_period, date_start, date_end, rate_state,
):
    """Locally match category=2, splitting only when pagination overflows.

    Company-name filtering is not part of every ``CodalAPI`` signature.  When
    it is unavailable, a market-wide Category 2 result set can exceed the page
    safety limit before the target company's announcement is reached.  Bisect
    that same Jalali-month window and retry locally instead of silently turning
    the safety limit into a false ``MISSING_REPORT``.
    """
    windows = chunk_publish_windows(
        date_start, date_end, UNFILTERED_CATEGORY2_WINDOW_DAYS,
    )
    while windows:
        window_start, window_end = windows.pop(0)
        result, overflowed = _scan_unfiltered_category2_window(
            api, symbol, company_name, target_period, window_start, window_end,
            rate_state,
        )
        if result[0] is not None:
            return result
        if overflowed:
            child_windows = split_publish_window(window_start, window_end)
            if child_windows is None:
                print(
                    "      WARNING: one-day Category 2 window still exceeds "
                    "the pagination safety limit."
                )
            else:
                print(
                    "      Category 2 pagination limit reached; splitting into "
                    f"{child_windows[0][0]}..{child_windows[0][1]} and "
                    f"{child_windows[1][0]}..{child_windows[1][1]}."
                )
                windows[0:0] = child_windows
    return None, 0


def _scan_unfiltered_category2_window(
    api, symbol, company_name, target_period, date_start, date_end, rate_state,
):
    """Scan one Category 2 window and report whether its page cap overflowed."""
    seen = set()
    previous_page_signature = None

    for page in range(1, UNFILTERED_CATEGORY2_MAX_PAGES + 1):
        data = _get_announcements_with_retry(
            api,
            {
                "category": 2,
                "date_start": date_start,
                "date_end": date_end,
                "page": page,
            },
            rate_state,
        )
        page_reports = data.get("announcement", []) or []
        print(
            f"      Category 2, page {page}: {len(page_reports)} "
            "unfiltered result(s); applying local symbol/company/period match"
        )
        if not page_reports:
            break

        page_signature = tuple(
            announcement_identity(item) for item in page_reports
        )
        if page_signature == previous_page_signature:
            print(
                "      Repeated page detected; unfiltered category=2 "
                "window is complete."
            )
            break
        previous_page_signature = page_signature

        unique_page_reports = []
        for report in page_reports:
            identity = announcement_identity(report)
            if identity not in seen:
                seen.add(identity)
                unique_page_reports.append(report)

        result = select_latest_report(
            unique_page_reports,
            symbol,
            company_name,
            target_period,
        )
        if result[0] is not None:
            print(
                "      Valid category=2 report found by local matching; "
                "unfiltered scan stopped immediately."
            )
            return result, False

        count_page = (
            data.get("count_page") or data.get("countPage")
            or data.get("total_pages") or data.get("totalPages")
        )
        try:
            if count_page is not None and page >= int(count_page):
                print(
                    "      Category 2 window exhausted according to "
                    f"count_page={count_page}."
                )
                return (None, 0), False
        except (TypeError, ValueError):
            pass
    else:
        print(
            "      WARNING: unfiltered category=2 pagination safety limit "
            "reached for this bounded window."
        )
        return (None, 0), True

    return (None, 0), False


# ============================================================
# HTML DOWNLOAD / REVISION RESOLUTION
# ============================================================

class _RevisionLinkParser(HTMLParser):
    """Collect anchor text plus nearby text used by CODAL's revision banner."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.recent_text = ""
        self.current = None
        self.links = []

    def handle_starttag(self, tag, attrs):
        if tag.casefold() != "a":
            return
        attributes = dict(attrs)
        self.current = {
            "href": attributes.get("href", ""),
            "text": "",
            "context_before": self.recent_text[-240:],
            "attributes": " ".join(str(value) for _key, value in attrs if value),
        }

    def handle_data(self, data):
        self.recent_text = (self.recent_text + " " + data)[-800:]
        if self.current is not None:
            self.current["text"] += " " + data

    def handle_endtag(self, tag):
        if tag.casefold() == "a" and self.current is not None:
            self.links.append(self.current)
            self.current = None


def _extract_url_from_href(href):
    """Accept ordinary links and common javascript window/open wrappers."""
    href = unescape(str(href or "")).strip()
    if not href:
        return None
    if not href.casefold().startswith("javascript:"):
        return href
    match = re.search(r"['\"]([^'\"]+(?:LetterSerial|TracingNo|Report|Reports)[^'\"]*)['\"]", href, re.I)
    return match.group(1) if match else None


def find_newer_revision_link(html, current_url):
    """Return CODAL's explicit newer-version link, never its older link."""
    parser = _RevisionLinkParser()
    parser.feed(html)
    newer_hints = (
        "اطلاعیه جدیدتر", "نسخه جدیدتر", "گزارش جدیدتر",
        "اطلاعیه بعدی", "نسخه بعدی", "اصلاحیه جدید",
    )
    older_hints = (
        "اطلاعیه قبلی", "نسخه قبلی", "گزارش قبلی", "اطلاعیه پیشین",
        "نسخه قدیمی", "قدیمی تر", "قدیمی‌تر",
    )
    for link in parser.links:
        # The banner sentence may precede the clickable words, so include a
        # bounded amount of adjacent text and link attributes in the decision.
        context = normalize_text(" ".join((
            link.get("context_before", ""), link.get("text", ""),
            link.get("attributes", ""),
        )))
        newest_newer_hint = max((context.rfind(hint) for hint in newer_hints), default=-1)
        newest_older_hint = max((context.rfind(hint) for hint in older_hints), default=-1)
        if newest_newer_hint < 0 or newest_newer_hint < newest_older_hint:
            continue
        href = _extract_url_from_href(link.get("href"))
        if href:
            resolved = urljoin(current_url, href)
            if resolved != current_url:
                return resolved
    return None


def _get_html_with_retry(url, rate_state):
    """Fetch one CODAL URL with bounded retries for 429 and timeouts."""
    rate_limit_retries = 0
    timeout_retries = 0
    while True:
        elapsed = time.monotonic() - rate_state["last_request_at"]
        if elapsed < BRSAPI_MIN_REQUEST_INTERVAL:
            time.sleep(BRSAPI_MIN_REQUEST_INTERVAL - elapsed)
        rate_state["last_request_at"] = time.monotonic()
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            response.encoding = "utf-8"
            return response.text
        except requests.Timeout as error:
            if timeout_retries >= CODAL_HTML_MAX_TIMEOUT_RETRIES:
                raise
            wait_seconds = min(
                CODAL_HTML_INITIAL_BACKOFF * (2 ** timeout_retries),
                CODAL_HTML_MAX_BACKOFF,
            )
            timeout_retries += 1
            print(
                f"  CODAL HTML {type(error).__name__}; waiting "
                f"{wait_seconds:.1f}s before retry "
                f"{timeout_retries}/{CODAL_HTML_MAX_TIMEOUT_RETRIES} "
                "of the same URL."
            )
            time.sleep(wait_seconds)
        except requests.HTTPError as error:
            if (
                not _is_rate_limit_error(error)
                or rate_limit_retries >= BRSAPI_MAX_429_RETRIES
            ):
                raise
            retry_after = _retry_after_seconds(error)
            backoff = min(
                BRSAPI_INITIAL_BACKOFF * (2 ** rate_limit_retries),
                BRSAPI_MAX_BACKOFF,
            )
            wait_seconds = max(
                retry_after if retry_after is not None else backoff,
                BRSAPI_MIN_REQUEST_INTERVAL,
            )
            rate_limit_retries += 1
            print(
                f"  CODAL HTML rate limit (429); waiting {wait_seconds:.1f}s "
                f"before retry {rate_limit_retries}/"
                f"{BRSAPI_MAX_429_RETRIES} of the same URL."
            )
            time.sleep(wait_seconds)


def _write_report_html(html, symbol, target_period):
    safe_symbol = symbol.replace("/", "_").replace("\\", "_")
    path = OUTPUT_HTML_DIR / f"{safe_symbol}_{period_html_suffix(target_period)}.html"
    path.write_text(html, encoding="utf-8")
    return path


def resolve_latest_revision(report, symbol, target_period, rate_state):
    """Follow explicit CODAL newer-version links and return the terminal HTML."""
    current_report = dict(report)
    current_url = current_report.get("link")
    if not current_url:
        raise RuntimeError(f"No HTML link for {symbol}")

    visited = set()
    hops = 0
    while True:
        if current_url in visited:
            raise RuntimeError(f"Revision link cycle detected for {symbol}")
        visited.add(current_url)
        html = _get_html_with_retry(current_url, rate_state)
        newer_url = find_newer_revision_link(html, current_url)
        if not newer_url:
            path = _write_report_html(html, symbol, target_period)
            return current_report, html, path, hops
        hops += 1
        if hops > REVISION_CHAIN_MAX_HOPS:
            raise RuntimeError(
                f"Revision chain exceeded {REVISION_CHAIN_MAX_HOPS} hops for {symbol}"
            )
        print(f"  REVISION RESOLVER: following newer version ({hops})")
        current_url = newer_url
        current_report = dict(current_report)
        current_report["link"] = current_url

def sales_value_is_empty_or_zero(value):
    """Identify comparison values that require targeted validation."""
    if value is None or value == 0:
        return True
    if isinstance(value, str):
        normalized = normalize_digits(value).replace(",", "").strip()
        return normalized in {"", "0", "0.0"}
    return False


def apply_prior_year_sales_fallback(
    parsed,
    symbol,
    company_name,
    target_period,
    parser,
    api,
    company_count,
    rate_state,
    prior_year_reports_cache,
):
    """Validate an absent/zero current comparison against prior-period YTD.

    This function must only be called after the current report parsed
    successfully.  The lookup is symbol-filtered and cached by symbol+period,
    so one missing comparison never downloads the prior-year market batch.
    """
    current_value = parsed.get("sales_last_year")
    if not sales_value_is_empty_or_zero(current_value):
        return "NOT_NEEDED", None, None

    fallback_period = prior_year_period(target_period)
    # Retain ``company_count`` in the signature for call-site compatibility;
    # targeted lookup deliberately does not use it.
    _ = company_count
    prior_report, prior_report_count = fetch_targeted_period_reports(
        api,
        symbol,
        company_name,
        fallback_period,
        rate_state,
        prior_year_reports_cache,
    )
    if prior_report is None:
        return "MISSING_PRIOR_REPORT", None, (
            f"No valid monthly-activity report found for {fallback_period}; "
            f"sales_last_year retains {current_value!r}."
        )

    try:
        prior_report, prior_html, prior_html_path, prior_revision_hops = (
            resolve_latest_revision(
                prior_report, symbol, fallback_period, rate_state
            )
        )
        prior_parsed = parser.parse(prior_html)
    except Exception as error:
        return "PRIOR_PARSE_FAILED", prior_report, (
            f"Prior-year report could not be parsed ({type(error).__name__}: "
            f"{error}); sales_last_year retains {current_value!r}."
        )

    prior_sales_ytd = prior_parsed.get("sales_ytd")
    if prior_sales_ytd is None or (
        isinstance(prior_sales_ytd, str) and not prior_sales_ytd.strip()
    ):
        return "MISSING_PRIOR_SALES_YTD", prior_report, (
            f"Prior-year report for {fallback_period} has no sales_ytd; "
            f"sales_last_year retains {current_value!r}."
        )

    parsed["sales_last_year"] = prior_sales_ytd
    revision_note = (
        f"; latest revision selected ({prior_report_count} search candidate(s), "
        f"{prior_revision_hops} HTML chain hop(s))"
        if prior_report_count > 1 or prior_revision_hops else ""
    )
    status = (
        "CONFIRMED_ZERO"
        if current_value == 0 and prior_sales_ytd == 0
        else "FILLED"
    )
    action = "confirmed as zero" if status == "CONFIRMED_ZERO" else "filled"
    return status, prior_report, (
        f"sales_last_year {action} from {fallback_period} sales_ytd"
        f"{revision_note}; HTML: {prior_html_path}"
    )


# ============================================================
# EXCEL HELPERS
# ============================================================

def reset_worksheet_view(ws):
    """Give Auto sheets a valid, unfrozen Excel worksheet view."""
    ws.freeze_panes = None
    ws.sheet_view.pane = None
    ws.sheet_view.selection = [
        Selection(pane=None, activeCell="A1", sqref="A1")
    ]


def copy_sheet_layout(source_ws, target_ws):
    """
    Copy the visible/manual sheet layout so Auto has the same
    structure and formatting.
    """

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[
                source_cell.coordinate
            ]

            target_cell.value = source_cell.value

            if source_cell.has_style:
                target_cell._style = copy(
                    source_cell._style
                )

            if source_cell.number_format:
                target_cell.number_format = (
                    source_cell.number_format
                )

            if source_cell.font:
                target_cell.font = copy(
                    source_cell.font
                )

            if source_cell.fill:
                target_cell.fill = copy(
                    source_cell.fill
                )

            if source_cell.border:
                target_cell.border = copy(
                    source_cell.border
                )

            if source_cell.alignment:
                target_cell.alignment = copy(
                    source_cell.alignment
                )

            if source_cell.protection:
                target_cell.protection = copy(
                    source_cell.protection
                )

    for key, dimension in (
        source_ws.column_dimensions.items()
    ):
        target_ws.column_dimensions[
            key
        ].width = dimension.width

        target_ws.column_dimensions[
            key
        ].hidden = dimension.hidden

    for key, dimension in (
        source_ws.row_dimensions.items()
    ):
        target_ws.row_dimensions[
            key
        ].height = dimension.height

        target_ws.row_dimensions[
            key
        ].hidden = dimension.hidden

    for merged_range in (
        source_ws.merged_cells.ranges
    ):
        target_ws.merge_cells(
            str(merged_range)
        )

    target_ws.sheet_view.showGridLines = (
        source_ws.sheet_view.showGridLines
    )

    reset_worksheet_view(target_ws)


def _header_token(value):
    return re.sub(
        r"[\s_\-]+", "", normalize_digits(normalize_text(value))
    ).casefold()


HEADER_ALIASES = {
    "Company_ID": ("Company_ID", "Company ID", "کد شرکت"),
    "Company_Name": ("Company_Name", "Company Name", "نام شرکت"),
    "Symbol": ("Symbol", "نماد"),
    "Fiscal_Year_End": (
        "Fiscal_Year_End", "Fiscal Year End", "ماه سال مالی", "پایان سال مالی",
    ),
    "Report_Month": ("Report_Month", "Report Month", "ماه گزارش"),
    "Reporting_Period_Months": (
        "Reporting_Period_Months", "Reporting Period Months",
        "تعداد ماه دوره گزارش", "دوره گزارش",
    ),
    "Sales_Prior_Year_YTD": (
        "Sales_Prior_Year_YTD", "Sales Last Year", "sales_last_year",
        "فروش دوره مشابه سال قبل",
    ),
    "Sales_YTD": ("Sales_YTD", "Sales YTD", "sales_ytd", "فروش تجمعی"),
    "Sales_Month": ("Sales_Month", "Sales Month", "sales_month", "فروش ماه"),
    "Sales_Prior_Month_YTD": (
        "Sales_Prior_Month_YTD", "Sales Prior Month YTD",
        "sales_prior_month_ytd", "فروش تجمعی ماه قبل",
    ),
    "Export_Prior_Year_YTD": (
        "Export_Prior_Year_YTD", "Export Last Year", "export_last_year",
        "صادرات دوره مشابه سال قبل",
    ),
    "Export_YTD": ("Export_YTD", "Export YTD", "export_ytd", "صادرات تجمعی"),
    "Export_Month": ("Export_Month", "Export Month", "export_month", "صادرات ماه"),
    **{
        header: (header, header.replace("_", " "))
        for header in (
            *FORECAST_COLUMNS_BEFORE_SYMBOL,
            *FORECAST_DIAGNOSTIC_COLUMNS,
        )
    },
    "Month_of_Last_Sales_Rec": (
        "Month_of_Last_Sales_Rec", "Month of Last Sales Rec",
        "Forecast_As_Of", "Forecast As Of",
    ),
}


def find_header_map(ws, required=()):
    """Return (header row, canonical header -> column) for an Auto-like sheet."""
    aliases = {
        canonical: {_header_token(label) for label in labels}
        for canonical, labels in HEADER_ALIASES.items()
    }
    best = None
    for row in range(1, min(ws.max_row, 40) + 1):
        found = {}
        for column in range(1, min(max(ws.max_column, len(AUTO_HEADERS)), 40) + 1):
            token = _header_token(ws.cell(row=row, column=column).value)
            if not token:
                continue
            for canonical, accepted in aliases.items():
                if token in accepted and canonical not in found:
                    found[canonical] = column
        candidate = (len(found), row, found)
        if best is None or candidate[0] > best[0]:
            best = candidate
    available = best[2] if best else {}
    missing = set(required) - set(available)
    if missing:
        raise RuntimeError("Auto header row is missing: " + ", ".join(sorted(missing)))
    return best[1], available


def clear_auto_data(ws, header_map, header_row):
    """Clear parser outputs by stable header name, never by Excel letter."""
    for column in (header_map[header] for header in SALES_HEADERS.values()):
        for row in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = None


def standardize_auto_schema(ws, company_map, company_by_name, target_period):
    """Rewrite the copied Template to the fixed schema while preserving formulas."""
    header_row, old_map = find_header_map(
        ws,
        required=(
            "Company_Name", "Symbol", "Report_Month",
            "Reporting_Period_Months",
        ),
    )
    required_sales = set(SALES_HEADERS.values())
    if required_sales.issubset(old_map):
        sales_map = {header: old_map[header] for header in required_sales}
    else:
        legacy = legacy_auto_header_map(ws, header_row=header_row)
        if legacy is None:
            missing = sorted(required_sales - set(old_map))
            raise RuntimeError(
                "Auto sales columns are neither canonical nor a valid legacy "
                "H:N layout; missing: " + ", ".join(missing)
            )
        _legacy_header_row, legacy_map = legacy
        sales_map = {
            header: legacy_map[header]
            for header in required_sales
        }

    # Metadata is discovered from its aliases; parser values come from either
    # the complete canonical block or the validated historical H:N block.
    # Snapshot everything before overwriting columns because the two layouts
    # overlap in different positions.
    source_by_target = {
        header: sales_map.get(header, old_map.get(header))
        for header in AUTO_HEADERS
        if header != "Company_ID"
    }
    snapshots = {}
    original_max_row = ws.max_row
    for header, source_column in source_by_target.items():
        if source_column is None:
            continue
        snapshots[header] = [
            (ws.cell(row=row, column=source_column).value,
             copy(ws.cell(row=row, column=source_column)._style))
            for row in range(1, original_max_row + 1)
        ]

    old_max_column = ws.max_column
    for target_column, header in enumerate(AUTO_HEADERS, start=1):
        source_column = source_by_target.get(header)
        snapshot = snapshots.get(header)
        for row in range(1, original_max_row + 1):
            cell = ws.cell(row=row, column=target_column)
            if snapshot is not None:
                cell.value, cell._style = snapshot[row - 1]
            elif header == "Company_ID":
                cell._style = copy(
                    ws.cell(row=row, column=old_map["Company_Name"])._style
                )
                cell.value = None
        ws.cell(row=header_row, column=target_column).value = header
        if source_column is not None:
            source_letter = get_column_letter(source_column)
            target_letter = get_column_letter(target_column)
            ws.column_dimensions[target_letter].width = (
                ws.column_dimensions[source_letter].width
            )

    if old_max_column > len(AUTO_HEADERS):
        ws.delete_cols(len(AUTO_HEADERS) + 1, old_max_column - len(AUTO_HEADERS))

    header_map = {header: column for column, header in enumerate(AUTO_HEADERS, 1)}
    symbol_rows = get_symbol_rows(ws, company_map, company_by_name, header_map)
    report_month = parse_period(target_period)[1]
    fiscal_year_column = get_column_letter(header_map["Fiscal_Year_End"])
    report_month_column = get_column_letter(header_map["Report_Month"])
    for row, _symbol, _name, company_id, _start_period in symbol_rows:
        ws.cell(row=row, column=header_map["Company_ID"]).value = company_id
        ws.cell(row=row, column=header_map["Report_Month"]).value = report_month
        ws.cell(
            row=row,
            column=header_map["Reporting_Period_Months"],
        ).value = (
            f"=IF({report_month_column}{row}>{fiscal_year_column}{row},"
            f"({report_month_column}{row}-{fiscal_year_column}{row}),"
            f"((12-{fiscal_year_column}{row})+{report_month_column}{row}))"
        )
    return header_row, header_map, symbol_rows


def row_has_valid_sales_data(ws, row, header_map):
    return all(
        ws.cell(row=row, column=header_map[header]).value is not None
        for header in SALES_HEADERS.values()
    )


def row_has_valid_current_sales_data(ws, row, header_map):
    """Return True when only the prior-year comparison may need repair."""
    comparison_header = SALES_HEADERS["sales_last_year"]
    return all(
        ws.cell(row=row, column=header_map[header]).value is not None
        for header in SALES_HEADERS.values()
        if header != comparison_header
    )


def legacy_auto_header_map(ws, header_row=None):
    """Return a safe H:N migration map, or None for an unknown old layout."""
    try:
        discovered_row, identity_map = find_header_map(
            ws, required=("Company_Name", "Symbol")
        )
    except RuntimeError:
        return None
    if header_row is not None and discovered_row != header_row:
        return None
    header_row = discovered_row

    # H:N is only safe when it actually looks like the historical Auto block.
    # Some early Auto sheets lost the seven visible H:N captions while retaining
    # complete parser data underneath them.  Accept that exact legacy case only
    # for an Auto sheet with the original B:G identity layout and several fully
    # populated H:N data rows.  This keeps migration strict for unknown sheets.
    if ws.max_column < 14:
        return None
    legacy_headers = [ws.cell(header_row, column).value for column in range(8, 15)]

    recognized = sum(
        _header_token(ws.cell(header_row, column).value)
        in {_header_token(label) for label in HEADER_ALIASES[canonical]}
        for canonical, column in LEGACY_SALES_COLUMNS.items()
    )
    headers_populated = all(normalize_text(value) for value in legacy_headers)
    if not (headers_populated and recognized >= 2):
        blank_headers = all(not normalize_text(value) for value in legacy_headers)
        consistent_header_state = blank_headers or headers_populated
        original_identity_layout = (
            ws.title.startswith("Auto ")
            and identity_map.get("Company_Name") == 2
            and identity_map.get("Symbol") == 3
            and identity_map.get("Report_Month") == 6
            and identity_map.get("Reporting_Period_Months") == 7
        )
        complete_legacy_rows = sum(
            all(ws.cell(row=row, column=column).value is not None
                for column in range(8, 15))
            for row in range(header_row + 1, ws.max_row + 1)
        )
        if not (
            consistent_header_state
            and original_identity_layout
            and complete_legacy_rows >= 3
        ):
            return None

    return header_row, {
        **identity_map,
        **LEGACY_SALES_COLUMNS,
    }


def existing_sales_by_company(ws, company_map, company_by_name):
    """Capture every non-empty sales row before the Auto sheet is rebuilt."""
    _header_row, discovered = find_header_map(ws)
    required_sales = set(SALES_HEADERS.values())
    if {"Company_Name", "Symbol", *required_sales}.issubset(discovered):
        header_map = discovered
        source_schema = "canonical"
    else:
        legacy = legacy_auto_header_map(ws)
        if legacy is None:
            print(
                f"WARNING: {ws.title}: existing Auto schema is not recognized; "
                "old values will not be preserved and the sheet will be rebuilt."
            )
            return {}
        _header_row, header_map = legacy
        source_schema = "legacy H:N"

    existing = {}
    for row, _symbol, _name, company_id, _start_period in get_symbol_rows(
        ws, company_map, company_by_name, header_map
    ):
        values = {
            header: ws.cell(row=row, column=header_map[header]).value
            for header in SALES_HEADERS.values()
        }
        # Preserve even a partially populated row.  Its missing fields can be
        # fetched below, but a failed lookup must never erase known-good cells.
        if any(value is not None for value in values.values()):
            existing[company_id] = values
    print(
        f"Migration: {ws.title}: preserved {len(existing)} non-empty sales row(s) "
        f"from {source_schema} schema."
    )
    return existing


def restore_existing_sales(ws, symbol_rows, existing, header_map):
    restored = set()
    prior_year_pending = set()
    for row, _symbol, _name, company_id, _start_period in symbol_rows:
        values = existing.get(company_id)
        if values is None:
            continue
        for header, value in values.items():
            ws.cell(row=row, column=header_map[header]).value = value
        current_values_complete = all(
            values[header] is not None
            for header in SALES_HEADERS.values()
            if header != SALES_HEADERS["sales_last_year"]
        )
        if not current_values_complete:
            # Leave the company pending so missing fields are fetched; restored
            # values remain in place if discovery or parsing fails.
            continue
        if sales_value_is_empty_or_zero(
            values[SALES_HEADERS["sales_last_year"]]
        ):
            prior_year_pending.add(company_id)
        else:
            restored.add(company_id)
    return restored, prior_year_pending


def get_symbol_rows(ws, company_map, company_by_name, header_map=None):
    """
    Columns are located by header. Company_ID is authoritative when present;
    legacy sheets can still be migrated via registry name/symbol lookup.

    Skip empty rows and header rows automatically.
    """

    if header_map is None:
        _header_row, header_map = find_header_map(
            ws, required=("Company_Name", "Symbol")
        )
    company_name_column = header_map["Company_Name"]
    symbol_column = header_map["Symbol"]
    company_id_column = header_map.get("Company_ID")
    companies_by_id = {
        company["company_id"]: company for company in company_map.values()
    }
    rows = []

    for row in range(1, ws.max_row + 1):
        raw_company_name = ws.cell(
            row=row,
            column=company_name_column,
        ).value

        raw_symbol = ws.cell(
            row=row,
            column=symbol_column,
        ).value

        company_name = normalize_text(raw_company_name)
        sheet_symbol = normalize_text(raw_symbol)

        if not company_name or not sheet_symbol:
            continue

        if sheet_symbol in (
            "نماد",
            "Symbol",
            "symbol",
        ):
            continue

        company = None
        if company_id_column is not None:
            company = companies_by_id.get(normalize_text(
                ws.cell(row=row, column=company_id_column).value
            ))
        company = company or company_by_name.get(company_name)
        if company is None:
            company = company_map.get(sheet_symbol)

        # Only companies enabled for both general and Sales monitoring are
        # present in these Registry maps.
        if company is None:
            continue

        rows.append(
            (
                row,
                company["symbol"],
                company["company_name"],
                company["company_id"],
                company["start_period"],
            )
        )

    return rows

# ============================================================
# LOG
# ============================================================

def prepare_log_sheet(wb):
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET)

        headers = [
            "Period",
            "Symbol",
            "Status",
            "Report Count",
            "Publish Date",
            "Publish Time",
            "Title",
            "Letter Serial",
            "HTML Link",
            "HTML File",
            "Extracted At",
            "Message",
        ]

        for col, value in enumerate(
            headers,
            start=1,
        ):
            ws.cell(
                row=1,
                column=col,
                value=value,
            )

    return ws


def latest_logged_report_links(log_ws, target_period):
    """Return the last successfully extracted link per symbol for a period."""
    successful = {"OK", "REVISION_SELECTED", "RECOVERED_REPORT", "NEW", "REVISED"}
    links = {}
    for row in range(2, log_ws.max_row + 1):
        if log_ws.cell(row=row, column=1).value != target_period:
            continue
        symbol = normalize_text(log_ws.cell(row=row, column=2).value)
        status = str(log_ws.cell(row=row, column=3).value or "")
        link = str(log_ws.cell(row=row, column=9).value or "").strip()
        if symbol and status in successful and link:
            links[symbol] = link
    return links

def log_result(
    log_ws,
    symbol,
    status,
    target_period,
    report=None,
    report_count=0,
    html_path=None,
    message="",
):
    
    report = report or {}

    next_row = log_ws.max_row + 1

    values = [
        target_period,
        symbol,
        status,
        report_count,
        report.get("date_publish"),
        report.get("time_publish"),
        report.get("title"),
        (
            report.get("letter_serial")
            or report.get("LetterSerial")
            or report.get("letterSerial")
            or ""
        ),
        report.get("link"),
        str(html_path or ""),
        datetime.now().isoformat(
            timespec="seconds"
        ),
        message,
    ]

    for col, value in enumerate(
        values,
        start=1,
    ):
        log_ws.cell(
            row=next_row,
            column=col,
            value=value,
        )


# ============================================================
# EXCEL FIELD MAPPING
# ============================================================

def write_parser_result(ws, row, result, header_map):
    """Write parser fields through the stable public header contract."""
    for parser_key, header in SALES_HEADERS.items():
        ws.cell(row=row, column=header_map[header]).value = result[parser_key]


def write_parser_result_preserving_valid(ws, row, result, header_map):
    """Apply a revision without replacing prior values with blank/zero values."""
    for parser_key, header in SALES_HEADERS.items():
        value = result.get(parser_key)
        cell = ws.cell(row=row, column=header_map[header])
        if value not in (None, "", 0) or cell.value in (None, "", 0):
            cell.value = value


def _period_key(period):
    year, month, _day = parse_period(period)
    return year, month


def _previous_period(period):
    year, month, _day = parse_period(period)
    if month == 1:
        year -= 1
        month = 12
    else:
        month -= 1
    return format_period(year, month, jalali_month_days(year, month))


def _auto_period_key(ws):
    """Return (year, month) for an Auto sheet, or None for another sheet."""
    if not ws.title.startswith("Auto "):
        return None
    try:
        year, month, _day = parse_period(ws.title[5:].replace(" ", "/"))
    except (TypeError, ValueError):
        return None
    return year, month


def _next_period_key(period_key):
    year, month = period_key
    return (year + 1, 1) if month == 12 else (year, month + 1)


def _forecast_number(value):
    """Return a positive numeric sales value, otherwise None."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    if isinstance(value, str):
        normalized = normalize_digits(value).replace(",", "").strip()
        try:
            number = float(normalized)
        except ValueError:
            return None
        return number if number > 0 else None
    return None


def fiscal_year_and_month(calendar_year, calendar_month, fiscal_year_end):
    """Map one Jalali calendar month to the company's fiscal year/month."""
    fiscal_month = ((calendar_month - fiscal_year_end - 1) % 12) + 1
    fiscal_year = calendar_year if calendar_month <= fiscal_year_end else calendar_year + 1
    return fiscal_year, fiscal_month


def calculate_sales_forecast_v1(
    current_ytd,
    fiscal_month,
    prior_same_period,
    prior_full_year,
    historical_shares,
):
    """Calculate the frozen Forecast V1 and its auditable metadata."""
    current_ytd = _forecast_number(current_ytd)
    prior_same_period = _forecast_number(prior_same_period)
    prior_full_year = _forecast_number(prior_full_year)
    shares = [
        float(share) for share in historical_shares
        if isinstance(share, (int, float)) and share > 0
    ][:3]
    if current_ytd is None:
        return None
    if fiscal_month == 12:
        return {
            "forecast": current_ytd,
            "confidence": "Actual",
            "method": "Actual",
            "history_depth": len(shares),
            "manual_gap": 0.0,
            "calendar_flag": "Diagnostic only",
        }

    if len(shares) >= 3:
        seasonality = 0.5 * shares[0] + 0.3 * shares[1] + 0.2 * shares[2]
        machine_method = "3Y 50/30/20"
    elif len(shares) == 2:
        seasonality = 0.5 * shares[0] + 0.5 * shares[1]
        machine_method = "2Y 50/50"
    elif len(shares) == 1:
        seasonality = shares[0]
        machine_method = "1Y recent"
    else:
        seasonality = None
        machine_method = "No machine history"
    machine = current_ytd / seasonality if seasonality else None

    manual = None
    if prior_same_period and prior_full_year and fiscal_month:
        annualized = current_ytd * 12 / fiscal_month
        growth_based = prior_full_year * (current_ytd / prior_same_period)
        amount_change = prior_full_year + current_ytd - prior_same_period
        average = (annualized + 2 * growth_based + amount_change) / 4
        manual = (
            amount_change
            if current_ytd >= prior_same_period and average < prior_full_year
            else average
        )

    if fiscal_month == 1 and manual is not None:
        forecast, method = manual, "Manual"
    elif fiscal_month == 2 and manual is not None and machine is not None:
        forecast, method = 0.75 * manual + 0.25 * machine, "Hybrid 75M/25A"
    elif fiscal_month == 3 and manual is not None and machine is not None:
        forecast, method = 0.25 * manual + 0.75 * machine, "Hybrid 25M/75A"
    elif machine is not None:
        forecast, method = machine, machine_method
    elif manual is not None:
        forecast, method = manual, "Manual fallback"
    else:
        return None

    stability_gap = None
    stability_band = "N/A"
    if len(shares) >= 2:
        stability_gap = abs(shares[0] - shares[1]) / ((shares[0] + shares[1]) / 2)
        stability_band = (
            "High" if stability_gap < 0.10
            else "Medium" if stability_gap < 0.25
            else "Low"
        )
    depth = len(shares)
    confidence = (
        "High"
        if depth >= 3 and fiscal_month >= 8 and stability_band == "High"
        else "Medium"
        if depth >= 2 and fiscal_month >= 5 and stability_band != "Low"
        else "Low"
    )
    manual_gap = (
        abs(machine / manual - 1)
        if machine is not None and manual not in (None, 0)
        else None
    )
    return {
        "forecast": forecast,
        "confidence": confidence,
        "method": method,
        "history_depth": depth,
        "manual_gap": manual_gap,
        "calendar_flag": "Diagnostic only",
    }


def ensure_sales_trend_forecast_columns(ws):
    """Migrate Sales Trend once, preserving its established visual style."""
    _header_row, headers = find_header_map(
        ws, required=("Company_ID", "Company_Name", "Symbol", "Fiscal_Year_End")
    )
    if set(FORECAST_COLUMNS_BEFORE_SYMBOL + FORECAST_DIAGNOSTIC_COLUMNS).issubset(headers):
        return False

    symbol_column = headers["Symbol"]
    ws.insert_cols(symbol_column, len(FORECAST_COLUMNS_BEFORE_SYMBOL))
    for offset, header in enumerate(FORECAST_COLUMNS_BEFORE_SYMBOL):
        column = symbol_column + offset
        ws.cell(row=1, column=column).value = header
        for row in range(1, ws.max_row + 1):
            source = ws.cell(row=row, column=symbol_column - 1)
            target = ws.cell(row=row, column=column)
            if row > 1:
                target.value = None
            target._style = copy(source._style)
            target.alignment = copy(source.alignment)
        ws.column_dimensions[get_column_letter(column)].width = 18

    _header_row, headers = find_header_map(
        ws, required=("Company_ID", "Company_Name", "Symbol", "Fiscal_Year_End")
    )
    diagnostic_start = headers["Fiscal_Year_End"] + 1
    ws.insert_cols(diagnostic_start, len(FORECAST_DIAGNOSTIC_COLUMNS))
    for offset, header in enumerate(FORECAST_DIAGNOSTIC_COLUMNS):
        column = diagnostic_start + offset
        ws.cell(row=1, column=column).value = header
        for row in range(1, ws.max_row + 1):
            source = ws.cell(row=row, column=diagnostic_start - 1)
            target = ws.cell(row=row, column=column)
            if row > 1:
                target.value = None
            target._style = copy(source._style)
            target.alignment = copy(source.alignment)
        ws.column_dimensions[get_column_letter(column)].width = 18
    return True


def normalize_sales_trend_forecast_layout(ws):
    """Keep Symbol beside Company_Name and restore row-band formatting."""
    _header_row, headers = find_header_map(
        ws,
        required=(
            "Company_ID", "Company_Name", "Symbol", "Fiscal_Year_End",
            *FORECAST_COLUMNS_BEFORE_SYMBOL, *FORECAST_DIAGNOSTIC_COLUMNS,
        ),
    )
    desired = (
        "Company_ID",
        "Company_Name",
        "Symbol",
        "FY_Sales_Forecast",
        "Month_of_Last_Sales_Rec",
        "Forecast_Confidence",
        "Forecast_Method",
        "Fiscal_Year_End",
        *FORECAST_DIAGNOSTIC_COLUMNS,
    )
    metadata_columns = {headers[header] for header in desired}
    source_order = [headers[header] for header in desired]
    source_order.extend(
        column for column in range(1, ws.max_column + 1)
        if column not in metadata_columns
    )
    snapshots = []
    widths = []
    for source_column in source_order:
        snapshots.append([
            (
                ws.cell(row, source_column).value,
                copy(ws.cell(row, source_column)._style),
                copy(ws.cell(row, source_column).alignment),
                copy(ws.cell(row, source_column).protection),
                copy(ws.cell(row, source_column).comment),
            )
            for row in range(1, ws.max_row + 1)
        ])
        widths.append(
            ws.column_dimensions[get_column_letter(source_column)].width
        )
    for target_column, column_snapshot in enumerate(snapshots, 1):
        for row, (value, style, alignment, protection, comment) in enumerate(
            column_snapshot, 1
        ):
            cell = ws.cell(row, target_column)
            cell.value = value
            cell._style = style
            cell.alignment = alignment
            cell.protection = protection
            cell.comment = comment
        ws.column_dimensions[get_column_letter(target_column)].width = (
            widths[target_column - 1]
        )
    for column, header in enumerate(desired, 1):
        ws.cell(1, column).value = header

    # The source Symbol/diagnostic columns may have had plain white fills.
    # Company_Name is the authoritative alternating-row visual template.
    # Apply one authoritative band across the entire used row.  Historical
    # monthly blocks inherited a few white-filled metric columns from older
    # templates, which otherwise interrupt the alternating gray/white stripe.
    for row in range(4, ws.max_row + 1):
        band_fill = PatternFill(
            fill_type="solid",
            fgColor=(
                SALES_TREND_BAND_FILL_COLOR
                if row % 2 == 0
                else SALES_TREND_WHITE_FILL_COLOR
            ),
        )
        for column in range(1, ws.max_column + 1):
            ws.cell(row, column).fill = copy(band_fill)

    _ensure_sales_trend_row_band_rule(ws)


def _is_sales_trend_row_band_rule(rule):
    """Identify an older alternating-row rule without touching font rules."""
    if getattr(rule, "type", None) != "expression":
        return False
    formula = "".join(str(item) for item in (getattr(rule, "formula", None) or []))
    compact = re.sub(r"\s+", "", formula).upper().replace("$", "")
    return (
        "MOD(ROW(),2)" in compact
        or "ISEVEN(ROW())" in compact
        or "ROW()=EVEN(ROW())" in compact
        or "EVEN(ROW())=ROW()" in compact
    )


def _ensure_sales_trend_row_band_rule(ws):
    """Apply one authoritative gray/white band across the full used table."""
    if ws.max_row < 4 or ws.max_column < 1:
        return

    # Remove only superseded alternating-fill rules.  Percentage/font rules and
    # all other conditional formatting are preserved exactly as they are.
    retained = []
    for conditional_formatting in list(ws.conditional_formatting):
        sqref = str(conditional_formatting.sqref)
        rules = list(ws.conditional_formatting[conditional_formatting])
        kept = [rule for rule in rules if not _is_sales_trend_row_band_rule(rule)]
        if len(kept) != len(rules):
            del ws.conditional_formatting[sqref]
            retained.extend((sqref, rule) for rule in kept)
    for sqref, rule in retained:
        ws.conditional_formatting.add(sqref, rule)

    full_range = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(
            formula=["MOD(ROW(),2)=0"],
            fill=PatternFill(
                fill_type="solid",
                fgColor=SALES_TREND_BAND_FILL_COLOR,
            ),
        ),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(
            formula=["MOD(ROW(),2)=1"],
            fill=PatternFill(
                fill_type="solid",
                fgColor=SALES_TREND_WHITE_FILL_COLOR,
            ),
        ),
    )


def refresh_sales_trend_growth_formulas(ws):
    """Rebuild internal growth references after metadata column migration."""
    _header_row, headers = find_header_map(ws, required=("Company_ID",))
    company_column = headers["Company_ID"]
    period_columns = sales_trend_period_columns(ws)
    for period_key, columns in sorted(period_columns.items()):
        growth_column = columns.get("month_growth")
        if growth_column is None:
            continue
        previous = period_columns.get(
            (period_key[0] - 1, 12) if period_key[1] == 1
            else (period_key[0], period_key[1] - 1),
            {},
        )
        previous_month_column = previous.get("month")
        current_month_column = columns.get("month")
        for row in range(_header_row + 1, ws.max_row + 1):
            if ws.cell(row, company_column).value in (None, ""):
                continue
            target = ws.cell(row, growth_column)
            if current_month_column is None or previous_month_column is None:
                target.value = ""
            else:
                current_ref = ws.cell(row, current_month_column).coordinate
                previous_ref = ws.cell(row, previous_month_column).coordinate
                target.value = (
                    f'=IFERROR(IF(OR({current_ref}="",{previous_ref}="",'
                    f'{previous_ref}=0),"",{current_ref}/{previous_ref}-1),"")'
                )
            target.number_format = "0.0%"


def collect_auto_sales_history(wb):
    """Collect positive YTD observations from immutable monthly Auto sheets."""
    history = {}
    for ws in wb.worksheets:
        period_key = _auto_period_key(ws)
        if period_key is None:
            continue
        try:
            header_row, headers = find_header_map(
                ws, required=("Company_ID", "Sales_YTD")
            )
        except RuntimeError:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            company_id = ws.cell(row, headers["Company_ID"]).value
            ytd = _forecast_number(ws.cell(row, headers["Sales_YTD"]).value)
            if company_id in (None, "") or ytd is None:
                continue
            history.setdefault(str(company_id), {})[period_key] = ytd
    return history


def collect_auto_monthly_sales_history(wb):
    """Return reported monthly sales by company without changing Auto sheets."""
    history = {}
    for ws in wb.worksheets:
        period_key = _auto_period_key(ws)
        if period_key is None:
            continue
        try:
            header_row, headers = find_header_map(
                ws, required=("Company_ID", "Sales_Month")
            )
        except RuntimeError:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            company_id = ws.cell(row, headers["Company_ID"]).value
            monthly = _forecast_number(ws.cell(row, headers["Sales_Month"]).value)
            if company_id in (None, "") or monthly is None:
                continue
            history.setdefault(str(company_id), {})[period_key] = monthly
    return history


def _mode5_override_from_cell(cell):
    comment = cell.comment
    if comment is None or not comment.text.startswith(MODE5_COMMENT_PREFIX):
        return None
    try:
        payload = json.loads(comment.text[len(MODE5_COMMENT_PREFIX):])
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict) or not payload.get("active", True):
        return None
    return payload


def load_mode5_overrides(ws):
    """Read persistent fiscal-calendar overrides from visible Sales Trend cells."""
    try:
        header_row, headers = find_header_map(
            ws, required=("Company_ID", "Calendar_Regime_Flag")
        )
    except RuntimeError:
        return {}
    overrides = {}
    for row in range(header_row + 1, ws.max_row + 1):
        company_id = normalize_text(ws.cell(row, headers["Company_ID"]).value)
        if not company_id:
            continue
        payload = _mode5_override_from_cell(
            ws.cell(row, headers["Calendar_Regime_Flag"])
        )
        if payload:
            payload = dict(payload)
            payload["row"] = row
            overrides[company_id] = payload
    return overrides


def _mode5_adjusted_history(monthly_history, new_fiscal_year_end, opening_base=None):
    """Rebuild YTD history from unchanged monthly sales under the new calendar."""
    result = {}
    running = 0.0
    active_fiscal_year = None
    segment_available = True
    first_period = min(monthly_history) if monthly_history else None
    for period_key in sorted(monthly_history):
        fiscal_year, fiscal_month = fiscal_year_and_month(
            period_key[0], period_key[1], new_fiscal_year_end
        )
        if fiscal_year != active_fiscal_year:
            active_fiscal_year = fiscal_year
            running = 0.0
            segment_available = True
            if period_key == first_period and fiscal_month > 1:
                if opening_base is None:
                    segment_available = False
                else:
                    running = float(opening_base)
        monthly = monthly_history[period_key]
        if monthly is None:
            segment_available = False
        if segment_available:
            running += float(monthly)
            result[period_key] = running
    return result


def _mode5_formula_for_accumulated(ws, row, period_keys, period_columns, base):
    month_cells = [
        ws.cell(row, period_columns[key]["month"]).coordinate
        for key in period_keys
        if "month" in period_columns[key]
    ]
    if len(month_cells) != len(period_keys) or not month_cells:
        return ""
    refs = ",".join(month_cells)
    base_number = float(base or 0)
    return (
        f'=IF(COUNT({refs})<>{len(month_cells)},"",'
        f'SUM({refs})+{base_number:.15g})'
    )


def _mode5_corrected_monthly_history(wb, company_id):
    """Reproduce Sales Trend's validated monthly sales without circular formulas."""
    auto_by_period = {
        key: ws
        for ws in wb.worksheets
        if (key := _auto_period_key(ws)) is not None
    }
    source = {}
    for key, ws in auto_by_period.items():
        try:
            header_row, headers = find_header_map(
                ws,
                required=(
                    "Company_ID", "Fiscal_Year_End", "Sales_YTD",
                    "Sales_Month", "Sales_Prior_Month_YTD",
                ),
            )
        except RuntimeError:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            if normalize_text(ws.cell(row, headers["Company_ID"]).value) != company_id:
                continue
            source[key] = {
                "fiscal_year_end": _forecast_number(
                    ws.cell(row, headers["Fiscal_Year_End"]).value
                ),
                "ytd": _forecast_number(ws.cell(row, headers["Sales_YTD"]).value),
                "month": _forecast_number(ws.cell(row, headers["Sales_Month"]).value),
                "prior_month_ytd": _forecast_number(
                    ws.cell(row, headers["Sales_Prior_Month_YTD"]).value
                ),
            }
            break

    result = {}
    for key, current in sorted(source.items()):
        if current["ytd"] is None or current["month"] is None:
            continue
        corrected_ytd = current["ytd"]
        next_key = _next_period_key(key)
        later = source.get(next_key)
        if later is not None and later["prior_month_ytd"] is not None:
            next_fye = later["fiscal_year_end"]
            fiscal_reset = (
                next_fye is not None
                and (int(next_fye) % 12) + 1 == next_key[1]
            )
            if not fiscal_reset:
                corrected_ytd = later["prior_month_ytd"]
        result[key] = current["month"] + corrected_ytd - current["ytd"]
    return result


def _ensure_mode5_highlights(ws, addresses):
    """Keep Mode 5 cells visibly yellow above the alternating-row rules."""
    retained = []
    for conditional_formatting in list(ws.conditional_formatting):
        sqref = str(conditional_formatting.sqref)
        rules = list(ws.conditional_formatting[conditional_formatting])
        kept = [
            rule for rule in rules
            if not (
                getattr(rule, "type", None) == "expression"
                and [str(item).upper() for item in (rule.formula or [])] == ["1=1"]
            )
        ]
        if len(kept) != len(rules):
            del ws.conditional_formatting[sqref]
            retained.extend((sqref, rule) for rule in kept)
    for sqref, rule in retained:
        ws.conditional_formatting.add(sqref, rule)
    if not addresses:
        return
    highlight = FormulaRule(
        formula=["1=1"],
        fill=PatternFill(fill_type="solid", fgColor=MODE5_FILL_COLOR),
        stopIfTrue=True,
    )
    highlight.priority = 1
    ws.conditional_formatting.add(" ".join(sorted(set(addresses))), highlight)
    priority = 2
    for item in ws.conditional_formatting:
        for rule in ws.conditional_formatting[item]:
            if rule is highlight:
                continue
            rule.priority = priority
            priority += 1


def apply_mode5_overrides(wb):
    """Reapply persistent Mode 5 rules after every ordinary Sales Trend refresh."""
    ws = wb[SALES_TREND_SHEET]
    overrides = load_mode5_overrides(ws)
    if not overrides:
        return 0
    header_row, headers = find_header_map(
        ws,
        required=(
            "Company_ID", "Fiscal_Year_End", "Calendar_Regime_Flag",
        ),
    )
    period_columns = sales_trend_period_columns(ws)
    highlighted = []
    for company_id, override in overrides.items():
        row = override["row"]
        new_fye = int(override["new_fiscal_year_end"])
        opening_base = override.get("opening_base")
        ws.cell(row, headers["Fiscal_Year_End"]).value = new_fye
        corrected_monthly = _mode5_corrected_monthly_history(wb, company_id)
        adjusted_ytd = _mode5_adjusted_history(
            corrected_monthly, new_fye, opening_base
        )
        for key, columns in sorted(period_columns.items()):
            month_column = columns.get("month")
            accumulated_column = columns.get("accumulated")
            if month_column is not None and key in corrected_monthly:
                month_cell = ws.cell(row, month_column)
                month_cell.value = corrected_monthly[key]
                month_cell.number_format = "#,##0"
            if accumulated_column is None:
                continue
            cell = ws.cell(row, accumulated_column)
            cell.value = adjusted_ytd.get(key, "")
            cell.number_format = "#,##0"
            cell.fill = PatternFill(fill_type="solid", fgColor=MODE5_FILL_COLOR)
            highlighted.append(cell.coordinate)

        # Recalculate accumulated-sales YoY growth from the restated series.
        for key, columns in sorted(period_columns.items()):
            growth_column = columns.get("yoy_growth")
            current_column = columns.get("accumulated")
            prior_columns = period_columns.get((key[0] - 1, key[1]), {})
            prior_column = prior_columns.get("accumulated")
            if growth_column is None or current_column is None:
                continue
            growth = ws.cell(row, growth_column)
            if prior_column is None:
                growth.value = ""
            else:
                current = ws.cell(row, current_column).coordinate
                prior = ws.cell(row, prior_column).coordinate
                growth.value = (
                    f'=IFERROR(IF(OR({current}="",{prior}="",{prior}=0),"",'
                    f'{current}/{prior}-1),"")'
                )
            growth.number_format = "0.0%"

        flag_cell = ws.cell(row, headers["Calendar_Regime_Flag"])
        flag_cell.value = MODE5_FLAG
        flag_cell.fill = PatternFill(fill_type="solid", fgColor=MODE5_FILL_COLOR)
        fiscal_cell = ws.cell(row, headers["Fiscal_Year_End"])
        fiscal_cell.fill = PatternFill(
            fill_type="solid", fgColor=MODE5_FILL_COLOR
        )
        highlighted.extend((flag_cell.coordinate, fiscal_cell.coordinate))
    _ensure_mode5_highlights(ws, highlighted)
    return len(overrides)


def refresh_sales_forecast_metadata(wb):
    """Refresh Forecast V1 from the latest successful Auto month per company."""
    ws = wb[SALES_TREND_SHEET]
    ensure_sales_trend_forecast_columns(ws)
    normalize_sales_trend_forecast_layout(ws)
    refresh_sales_trend_growth_formulas(ws)
    apply_mode5_overrides(wb)
    header_row, headers = find_header_map(
        ws,
        required=(
            "Company_ID", "Fiscal_Year_End", "FY_Sales_Forecast",
            "Month_of_Last_Sales_Rec", "Forecast_Confidence", "Forecast_Method",
            "Forecast_Fiscal_Month", "Forecast_History_Depth",
            "Manual_vs_Model_Gap", "Calendar_Regime_Flag",
        ),
    )
    history = collect_auto_sales_history(wb)
    overrides = load_mode5_overrides(ws)
    for company_id, override in overrides.items():
        history[company_id] = _mode5_adjusted_history(
            _mode5_corrected_monthly_history(wb, company_id),
            int(override["new_fiscal_year_end"]),
            override.get("opening_base"),
        )
    updated = 0
    unavailable = 0
    for row in range(header_row + 1, ws.max_row + 1):
        company_id = str(ws.cell(row, headers["Company_ID"]).value or "")
        if not company_id:
            continue
        observations = history.get(company_id, {})
        if not observations:
            unavailable += 1
            continue
        fiscal_year_end = int(ws.cell(row, headers["Fiscal_Year_End"]).value)
        latest_period = max(observations)
        calendar_year, calendar_month = latest_period
        fiscal_year, fiscal_month = fiscal_year_and_month(
            calendar_year, calendar_month, fiscal_year_end
        )
        current_ytd = observations[latest_period]
        shares = []
        prior_same_period = None
        prior_full_year = None
        for lag in (1, 2, 3):
            prior_fy = fiscal_year - lag
            same_ytd = None
            full_year = None
            for period_key, ytd in observations.items():
                fy, fm = fiscal_year_and_month(
                    period_key[0], period_key[1], fiscal_year_end
                )
                if fy == prior_fy and fm == fiscal_month:
                    same_ytd = ytd
                if fy == prior_fy and fm == 12:
                    full_year = ytd
            if lag == 1:
                prior_same_period, prior_full_year = same_ytd, full_year
            if same_ytd and full_year:
                shares.append(same_ytd / full_year)

        result = calculate_sales_forecast_v1(
            current_ytd, fiscal_month, prior_same_period, prior_full_year, shares
        )
        if result is None:
            unavailable += 1
            continue
        values = {
            "FY_Sales_Forecast": result["forecast"],
            "Month_of_Last_Sales_Rec": calendar_month,
            "Forecast_Confidence": result["confidence"],
            "Forecast_Method": result["method"],
            "Forecast_Fiscal_Month": fiscal_month,
            "Forecast_History_Depth": result["history_depth"],
            "Manual_vs_Model_Gap": result["manual_gap"],
            "Calendar_Regime_Flag": (
                MODE5_FLAG if company_id in overrides else result["calendar_flag"]
            ),
        }
        for header, value in values.items():
            ws.cell(row, headers[header]).value = value
        ws.cell(row, headers["FY_Sales_Forecast"]).number_format = "#,##0"
        ws.cell(row, headers["Manual_vs_Model_Gap"]).number_format = "0.0%"
        updated += 1
    print(
        f"Sales Forecast V1 refreshed: {updated} company row(s); "
        f"{unavailable} unavailable."
    )
    return updated, unavailable


def sales_trend_period_columns(ws):
    """Map (year, month) to the metric columns declared in Sales Trend row 2."""
    result = {}
    patterns = (
        ("accumulated", r"^AS-(\d{4})/(\d{1,2})$"),
        ("month", r"^MS-(\d{4})/(\d{1,2})$"),
        ("month_growth", r"^MSG-(\d{4})/(\d{1,2})(?:\s|$)"),
        ("yoy_growth", r"^ASG-(\d{4})/(\d{1,2})(?:\s|-)"),
    )
    for column in range(1, ws.max_column + 1):
        label = normalize_digits(normalize_text(ws.cell(2, column).value))
        for metric, pattern in patterns:
            match = re.match(pattern, label, re.I)
            if match:
                key = (int(match.group(1)), int(match.group(2)))
                result.setdefault(key, {})[metric] = column
                break
    return result


def _copy_sales_trend_column(ws, source_column, target_column):
    source_letter = get_column_letter(source_column)
    target_letter = get_column_letter(target_column)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    ws.column_dimensions[target_letter].hidden = ws.column_dimensions[source_letter].hidden
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row=row, column=source_column)
        target = ws.cell(row=row, column=target_column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def reorder_sales_trend_period_columns(ws):
    """Sort complete Sales Trend period blocks oldest-to-newest, left-to-right."""
    period_columns = sales_trend_period_columns(ws)
    if len(period_columns) < 2:
        return False

    physical_order = sorted(
        period_columns, key=lambda key: min(period_columns[key].values())
    )
    chronological_order = sorted(period_columns)
    if physical_order == chronological_order:
        return False

    starts = {
        key: min(columns.values()) for key, columns in period_columns.items()
    }
    first_period_column = min(starts.values())
    ordered_by_position = sorted(starts, key=starts.get)
    block_ranges = {}
    for index, key in enumerate(ordered_by_position):
        start = starts[key]
        end = (
            starts[ordered_by_position[index + 1]] - 1
            if index + 1 < len(ordered_by_position)
            else ws.max_column
        )
        block_ranges[key] = (start, end)

    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col >= first_period_column:
            raise RuntimeError(
                "Sales Trend period-column reorder cannot safely move merged "
                f"range {merged_range}."
            )

    snapshots = {}
    for key, (start, end) in block_ranges.items():
        columns = []
        for source_column in range(start, end + 1):
            source_letter = get_column_letter(source_column)
            dimension = ws.column_dimensions[source_letter]
            cells = []
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=source_column)
                cells.append({
                    "source_coordinate": cell.coordinate,
                    "value": cell.value,
                    "style": copy(cell._style),
                    "number_format": cell.number_format,
                    "alignment": copy(cell.alignment),
                    "protection": copy(cell.protection),
                    "comment": copy(cell.comment),
                    "hyperlink": copy(cell.hyperlink),
                })
            columns.append({
                "width": dimension.width,
                "hidden": dimension.hidden,
                "best_fit": dimension.bestFit,
                "outline_level": dimension.outlineLevel,
                "collapsed": dimension.collapsed,
                "cells": cells,
            })
        snapshots[key] = columns

    target_column = first_period_column
    for key in chronological_order:
        for column_snapshot in snapshots[key]:
            target_letter = get_column_letter(target_column)
            target_dimension = ws.column_dimensions[target_letter]
            target_dimension.width = column_snapshot["width"]
            target_dimension.hidden = column_snapshot["hidden"]
            target_dimension.bestFit = column_snapshot["best_fit"]
            target_dimension.outlineLevel = column_snapshot["outline_level"]
            target_dimension.collapsed = column_snapshot["collapsed"]
            for row, cell_snapshot in enumerate(column_snapshot["cells"], start=1):
                target = ws.cell(row=row, column=target_column)
                value = cell_snapshot["value"]
                if isinstance(value, str) and value.startswith("="):
                    value = Translator(
                        value, origin=cell_snapshot["source_coordinate"]
                    ).translate_formula(target.coordinate)
                target.value = value
                target._style = copy(cell_snapshot["style"])
                target.number_format = cell_snapshot["number_format"]
                target.alignment = copy(cell_snapshot["alignment"])
                target.protection = copy(cell_snapshot["protection"])
                target.comment = copy(cell_snapshot["comment"])
                target._hyperlink = copy(cell_snapshot["hyperlink"])
            target_column += 1

    return True


def ensure_sales_trend_period_columns(ws, target_period):
    """Return a period block, creating future blocks from the 1405/02 template."""
    period_columns = sales_trend_period_columns(ws)
    target_key = _period_key(target_period)
    if target_key in period_columns:
        reorder_sales_trend_period_columns(ws)
        return sales_trend_period_columns(ws)[target_key]

    template_key = _period_key(SALES_TREND_TEMPLATE_PERIOD)
    template = period_columns.get(template_key)
    required = {"accumulated", "month", "month_growth", "yoy_growth"}
    if template is None or not required.issubset(template):
        raise RuntimeError(
            f"Sales Trend template block {SALES_TREND_TEMPLATE_PERIOD} is incomplete."
        )

    # Keep one empty separator column between monthly blocks.
    separator_column = ws.max_column + 1
    block_start = separator_column + 1
    template_start = min(template.values())
    template_separator = template_start - 1
    _copy_sales_trend_column(ws, template_separator, separator_column)
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=separator_column).value = None

    ordered_metrics = ("accumulated", "month", "month_growth", "yoy_growth")
    created = {}
    for offset, metric in enumerate(ordered_metrics):
        target_column = block_start + offset
        _copy_sales_trend_column(ws, template[metric], target_column)
        created[metric] = target_column

    year, month, day = parse_period(target_period)
    previous = _previous_period(target_period)
    prior_year = prior_year_period(target_period)
    headers = {
        "accumulated": (
            f"Accumulated Sales at {target_period}", f"AS-{year:04d}/{month:02d}"
        ),
        "month": (
            f"Sales at Month Ended {target_period}", f"MS-{year:04d}/{month:02d}"
        ),
        "month_growth": (
            "Month Sales Growth to Previous Month",
            f"MSG-{year:04d}/{month:02d} to {parse_period(previous)[1]:02d}",
        ),
        "yoy_growth": (
            "Accumulated Sales Growth to Same Period of Last Year",
            f"ASG-{year:04d}/{month:02d} to "
            f"{parse_period(prior_year)[0]:04d}/{month:02d}",
        ),
    }
    for metric, column in created.items():
        ws.cell(1, column).value = headers[metric][0]
        ws.cell(2, column).value = headers[metric][1]
        ws.cell(3, column).value = month
        for row in range(4, ws.max_row + 1):
            ws.cell(row, column).value = None
    reorder_sales_trend_period_columns(ws)
    return sales_trend_period_columns(ws)[target_key]


def _apply_sales_trend_font_template(ws):
    """Use the 1405/02 metric font colors for every monthly metric block."""
    period_columns = sales_trend_period_columns(ws)
    template = period_columns.get(_period_key(SALES_TREND_TEMPLATE_PERIOD))
    if not template:
        raise RuntimeError(
            f"Sales Trend font template not found: {SALES_TREND_TEMPLATE_PERIOD}"
        )
    for columns in period_columns.values():
        for metric, column in columns.items():
            template_column = template.get(metric)
            if template_column is None:
                continue
            for row in range(1, ws.max_row + 1):
                font = copy(ws.cell(row, column).font)
                font.color = copy(ws.cell(row, template_column).font.color)
                ws.cell(row, column).font = font


def _ensure_negative_percent_rule(ws):
    period_columns = sales_trend_period_columns(ws)
    percentage_columns = sorted({
        columns[metric]
        for columns in period_columns.values()
        for metric in ("month_growth", "yoy_growth")
        if metric in columns
    })
    ranges = " ".join(
        f"{get_column_letter(column)}4:{get_column_letter(column)}{ws.max_row}"
        for column in percentage_columns
    )
    if not ranges:
        return
    for conditional_formatting in list(ws.conditional_formatting):
        rules = ws.conditional_formatting[conditional_formatting]
        if any(
            rule.type == "cellIs"
            and rule.operator == "lessThan"
            and list(rule.formula or []) == ["0"]
            for rule in rules
        ):
            del ws.conditional_formatting[str(conditional_formatting.sqref)]
    if any(str(item.sqref) == ranges for item in ws.conditional_formatting):
        return
    ws.conditional_formatting.add(
        ranges,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=Font(color=NEGATIVE_PERCENT_FONT_COLOR),
        ),
    )


def update_sales_trend(wb, target_period, auto_ws, auto_header_row, auto_header_map):
    """Connect one Sales Trend month to its Auto sheet with auditable formulas."""
    trend_ws = wb[SALES_TREND_SHEET]
    columns = ensure_sales_trend_period_columns(trend_ws, target_period)
    _trend_header_row, trend_headers = find_header_map(
        trend_ws, required=("Company_ID", "Fiscal_Year_End")
    )
    trend_company_column = trend_headers["Company_ID"]
    trend_fiscal_year_end_column = trend_headers["Fiscal_Year_End"]
    auto_company_column = auto_header_map["Company_ID"]
    auto_accumulated_column = auto_header_map["Sales_YTD"]
    auto_month_column = auto_header_map["Sales_Month"]
    auto_prior_year_column = auto_header_map["Sales_Prior_Year_YTD"]
    auto_first_row = auto_header_row + 1
    auto_last_row = auto_ws.max_row
    auto_name = auto_ws.title.replace("'", "''")

    previous_columns = sales_trend_period_columns(trend_ws).get(
        _period_key(_previous_period(target_period)), {}
    )
    for row in range(_trend_header_row + 1, trend_ws.max_row + 1):
        if trend_ws.cell(row, trend_company_column).value in (None, ""):
            continue
        trend_id = f"${get_column_letter(trend_company_column)}{row}"
        id_range = (
            f"'{auto_name}'!${get_column_letter(auto_company_column)}"
            f"${auto_first_row}:${get_column_letter(auto_company_column)}${auto_last_row}"
        )

        def lookup_formula(source_column):
            source_range = (
                f"'{auto_name}'!${get_column_letter(source_column)}${auto_first_row}:"
                f"${get_column_letter(source_column)}${auto_last_row}"
            )
            lookup = f"INDEX({source_range},MATCH({trend_id},{id_range},0))"
            return f'=IFERROR(IF({lookup}="","",{lookup}),"")'

        accumulated_cell = trend_ws.cell(row, columns["accumulated"])
        month_cell = trend_ws.cell(row, columns["month"])
        month_growth_column = columns.get("month_growth")
        yoy_growth_column = columns.get("yoy_growth")
        accumulated_cell.value = lookup_formula(auto_accumulated_column)
        month_cell.value = lookup_formula(auto_month_column)

        if month_growth_column is not None:
            month_growth_cell = trend_ws.cell(row, month_growth_column)
            previous_month_column = previous_columns.get("month")
            if previous_month_column is None:
                month_growth_cell.value = ""
            else:
                current_ref = month_cell.coordinate
                previous_ref = trend_ws.cell(row, previous_month_column).coordinate
                month_growth_cell.value = (
                    f'=IF(OR({current_ref}="",{previous_ref}="",'
                    f'{previous_ref}=0),"",{current_ref}/{previous_ref}-1)'
                )
            month_growth_cell.number_format = "0.0%"

        if yoy_growth_column is not None:
            yoy_growth_cell = trend_ws.cell(row, yoy_growth_column)
            current_ytd = get_column_letter(auto_accumulated_column)
            prior_ytd = get_column_letter(auto_prior_year_column)
            yoy_growth_cell.value = (
                f'=IFERROR(IF(OR(INDEX(\'{auto_name}\'!${current_ytd}${auto_first_row}:'
                f'${current_ytd}${auto_last_row},MATCH({trend_id},{id_range},0))="",'
                f'INDEX(\'{auto_name}\'!${prior_ytd}${auto_first_row}:'
                f'${prior_ytd}${auto_last_row},MATCH({trend_id},{id_range},0))=0),"",'
                f'INDEX(\'{auto_name}\'!${current_ytd}${auto_first_row}:'
                f'${current_ytd}${auto_last_row},MATCH({trend_id},{id_range},0))/'
                f'INDEX(\'{auto_name}\'!${prior_ytd}${auto_first_row}:'
                f'${prior_ytd}${auto_last_row},MATCH({trend_id},{id_range},0))-1),"")'
            )
            yoy_growth_cell.number_format = "0.0%"
        accumulated_cell.number_format = "#,##0"
        month_cell.number_format = "#,##0"

    _apply_sales_trend_font_template(trend_ws)
    _ensure_negative_percent_rule(trend_ws)


def _excel_lookup(sheet_name, value_column, id_column, first_row, last_row, trend_id):
    """Build one bounded, auditable Company_ID lookup formula fragment."""
    safe_name = sheet_name.replace("'", "''")
    value_letter = get_column_letter(value_column)
    id_letter = get_column_letter(id_column)
    value_range = (
        f"'{safe_name}'!${value_letter}${first_row}:${value_letter}${last_row}"
    )
    id_range = f"'{safe_name}'!${id_letter}${first_row}:${id_letter}${last_row}"
    return f"INDEX({value_range},MATCH({trend_id},{id_range},0))"


def refresh_sales_trend_prior_corrections(wb, affected_periods=None):
    """
    Restate each Sales Trend accumulated value from the following Auto report.

    Auto sheets remain immutable.  For month t, the normal source is Sales_YTD
    in Auto month t.  When Auto month t+1 exists for the same Company_ID, its
    Sales_Prior_Month_YTD becomes the preferred source.  A row whose next
    report has Reporting_Period_Months == 1 is a fiscal-year reset and keeps
    the original month-t Sales_YTD.  Monthly sales is then recalculated from
    the corrected accumulated series so the analytical trend reconciles.

    The audit sheet is rebuilt on every run, making this operation idempotent.
    When affected_periods is provided, the full audit is still rebuilt but only
    those Sales Trend periods are rewritten.  Refresh-only mode omits the
    argument and deliberately retains the existing global refresh behaviour.
    """
    trend_ws = wb[SALES_TREND_SHEET]
    _trend_header_row, trend_headers = find_header_map(
        trend_ws, required=("Company_ID", "Fiscal_Year_End")
    )
    trend_company_column = trend_headers["Company_ID"]
    trend_fiscal_year_end_column = trend_headers["Fiscal_Year_End"]
    trend_columns = sales_trend_period_columns(trend_ws)

    auto_by_period = {
        key: ws
        for ws in wb.worksheets
        if (key := _auto_period_key(ws)) is not None
    }
    if affected_periods is not None:
        affected_periods = set(affected_periods)
    details = []

    for period_key, current_ws in sorted(auto_by_period.items()):
        rewrite_period = (
            affected_periods is None or period_key in affected_periods
        )
        columns = trend_columns.get(period_key)
        if not columns or "accumulated" not in columns:
            continue

        current_header_row, current_headers = find_header_map(
            current_ws,
            required=("Company_ID", "Sales_YTD"),
        )
        current_first = current_header_row + 1
        current_last = current_ws.max_row
        current_id_col = current_headers["Company_ID"]
        current_ytd_col = current_headers["Sales_YTD"]
        next_key = _next_period_key(period_key)
        next_ws = auto_by_period.get(next_key)

        next_lookup = None
        next_values = {}
        if next_ws is not None:
            next_header_row, next_headers = find_header_map(
                next_ws,
                required=(
                    "Company_ID",
                    "Fiscal_Year_End",
                    "Sales_Prior_Month_YTD",
                ),
            )
            next_first = next_header_row + 1
            next_last = next_ws.max_row
            next_id_col = next_headers["Company_ID"]
            next_fiscal_year_end_col = next_headers["Fiscal_Year_End"]
            next_prior_col = next_headers["Sales_Prior_Month_YTD"]
            for source_row in range(next_first, next_last + 1):
                company_id = normalize_text(next_ws.cell(source_row, next_id_col).value)
                if company_id:
                    next_values[company_id] = {
                        "row": source_row,
                        "fiscal_year_end": next_ws.cell(source_row, next_fiscal_year_end_col).value,
                        "prior_ytd": next_ws.cell(source_row, next_prior_col).value,
                    }
            next_lookup = (next_first, next_last, next_id_col, next_fiscal_year_end_col, next_prior_col)

        current_values = {}
        for source_row in range(current_first, current_last + 1):
            company_id = normalize_text(current_ws.cell(source_row, current_id_col).value)
            if company_id:
                current_values[company_id] = {
                    "row": source_row,
                    "sales_ytd": current_ws.cell(source_row, current_ytd_col).value,
                    "company_name": current_ws.cell(
                        source_row, current_headers.get("Company_Name", 2)
                    ).value,
                    "symbol": current_ws.cell(
                        source_row, current_headers.get("Symbol", 3)
                    ).value,
                }

        for trend_row in range(_trend_header_row + 1, trend_ws.max_row + 1):
            raw_id = trend_ws.cell(trend_row, trend_company_column).value
            company_id = normalize_text(raw_id)
            if not company_id:
                continue
            trend_id = f"${get_column_letter(trend_company_column)}{trend_row}"
            base = _excel_lookup(
                current_ws.title,
                current_ytd_col,
                current_id_col,
                current_first,
                current_last,
                trend_id,
            )
            target = trend_ws.cell(trend_row, columns["accumulated"])

            if next_ws is None:
                if rewrite_period:
                    target.value = f'=IFERROR(IF({base}="","",{base}),"")'
                continue

            next_first, next_last, next_id_col, next_fiscal_year_end_col, next_prior_col = next_lookup
            fiscal_year_end = _excel_lookup(
                next_ws.title, next_fiscal_year_end_col, next_id_col,
                next_first, next_last, trend_id,
            )
            corrected = _excel_lookup(
                next_ws.title, next_prior_col, next_id_col,
                next_first, next_last, trend_id,
            )
            if rewrite_period:
                target.value = (
                    f'=IFERROR(IF({base}="","",'
                    f'IF(OR(MOD({fiscal_year_end},12)+1={next_key[1]},'
                    f'{corrected}=""),{base},{corrected})),"")'
                )
                target.number_format = "#,##0"

            original = current_values.get(company_id)
            later = next_values.get(company_id)
            if original is None or later is None:
                details.append([
                    f"{period_key[0]:04d}/{period_key[1]:02d}", company_id,
                    original.get("company_name") if original else None,
                    original.get("symbol") if original else None,
                    current_ws.title, next_ws.title,
                    original.get("sales_ytd") if original else None,
                    later.get("prior_ytd") if later else None,
                    None, "Missing Company_ID", "Review",
                ])
                continue

            original_ytd = original["sales_ytd"]
            corrected_ytd = later["prior_ytd"]
            fiscal_year_end_value = later["fiscal_year_end"]

            # Historical backfills can reach a period before the available
            # Sales_YTD series.  There is no valid comparison in that case;
            # leave the worksheet formula intact and skip only the audit diff.
            if original_ytd in (None, "") or corrected_ytd in (None, ""):
                continue

            is_fiscal_reset = (
                isinstance(fiscal_year_end_value, (int, float))
                and int(fiscal_year_end_value) % 12 + 1 == next_key[1]
            )
            if is_fiscal_reset:
                if original_ytd != corrected_ytd:
                    details.append([
                        f"{period_key[0]:04d}/{period_key[1]:02d}", company_id,
                        original["company_name"], original["symbol"],
                        current_ws.title, next_ws.title, original_ytd, corrected_ytd,
                        None if original_ytd is None or corrected_ytd is None else corrected_ytd - original_ytd,
                        "Fiscal Reset - Not Applicable", "Original YTD retained",
                    ])
            elif corrected_ytd not in (None, "") and corrected_ytd != original_ytd:
                details.append([
                    f"{period_key[0]:04d}/{period_key[1]:02d}", company_id,
                    original["company_name"], original["symbol"],
                    current_ws.title, next_ws.title, original_ytd, corrected_ytd,
                    corrected_ytd - original_ytd,
                    "Correction Applied - OK", "Sales Trend uses next report",
                ])

    if SALES_TREND_AUDIT_SHEET in wb.sheetnames:
        del wb[SALES_TREND_AUDIT_SHEET]
    audit_ws = wb.create_sheet(SALES_TREND_AUDIT_SHEET)
    headers = [
        "Period", "Company_ID", "Company_Name", "Symbol", "Original_Auto_Sheet",
        "Next_Auto_Sheet", "Original_Sales_YTD", "Corrected_Prior_Month_YTD",
        "Difference", "Status", "Action",
    ]
    audit_ws.append(headers)
    for item in details:
        audit_ws.append(item)
    audit_ws.freeze_panes = "A2"
    audit_ws.auto_filter.ref = audit_ws.dimensions
    for cell in audit_ws[1]:
        cell.font = Font(bold=True)
    for column, width in enumerate((12, 14, 32, 14, 24, 24, 20, 28, 16, 34, 28), 1):
        audit_ws.column_dimensions[get_column_letter(column)].width = width

    # Apply the later accumulated correction to the same month's reported
    # Sales_Month.  Using the delta preserves CODAL's reported rounding:
    # corrected month = reported month + (corrected YTD - original YTD).
    period_data = {}
    for period_key, auto_ws in auto_by_period.items():
        header_row, headers = find_header_map(
            auto_ws,
            required=(
                "Company_ID", "Fiscal_Year_End", "Sales_YTD", "Sales_Month",
                "Sales_Prior_Month_YTD",
            ),
        )
        records = {}
        for row in range(header_row + 1, auto_ws.max_row + 1):
            company_id = normalize_text(auto_ws.cell(row, headers["Company_ID"]).value)
            if not company_id:
                continue
            records[company_id] = {
                "fiscal_year_end": auto_ws.cell(row, headers["Fiscal_Year_End"]).value,
                "sales_ytd": auto_ws.cell(row, headers["Sales_YTD"]).value,
                "sales_month": auto_ws.cell(row, headers["Sales_Month"]).value,
                "prior_month_ytd": auto_ws.cell(row, headers["Sales_Prior_Month_YTD"]).value,
                "company_name": auto_ws.cell(row, headers.get("Company_Name", 2)).value,
                "symbol": auto_ws.cell(row, headers.get("Symbol", 3)).value,
            }
        period_data[period_key] = records

    effective_ytd = {}
    for period_key, records in period_data.items():
        next_key = _next_period_key(period_key)
        next_records = period_data.get(next_key, {})
        for company_id, record in records.items():
            effective = record["sales_ytd"]
            later = next_records.get(company_id)
            if later is not None:
                fye = later["fiscal_year_end"]
                is_reset = (
                    isinstance(fye, (int, float))
                    and int(fye) % 12 + 1 == next_key[1]
                )
                if not is_reset and later["prior_month_ytd"] not in (None, ""):
                    effective = later["prior_month_ytd"]
            effective_ytd[(period_key, company_id)] = effective

    month_details = []
    row_by_id = {
        normalize_text(trend_ws.cell(row, trend_company_column).value): row
        for row in range(_trend_header_row + 1, trend_ws.max_row + 1)
        if normalize_text(trend_ws.cell(row, trend_company_column).value)
    }
    for period_key, records in sorted(period_data.items()):
        rewrite_period = (
            affected_periods is None or period_key in affected_periods
        )
        columns = trend_columns.get(period_key, {})
        if "accumulated" not in columns or "month" not in columns:
            continue
        auto_ws = auto_by_period[period_key]
        auto_header_row, auto_headers = find_header_map(
            auto_ws, required=("Company_ID", "Sales_YTD", "Sales_Month")
        )
        auto_first = auto_header_row + 1
        auto_last = auto_ws.max_row
        for company_id, record in records.items():
            trend_row = row_by_id.get(company_id)
            if trend_row is None:
                continue
            month_cell = trend_ws.cell(trend_row, columns["month"])
            current_accumulated = trend_ws.cell(
                trend_row, columns["accumulated"]
            ).coordinate
            trend_id = f"${get_column_letter(trend_company_column)}{trend_row}"
            original_ytd_formula = _excel_lookup(
                auto_ws.title, auto_headers["Sales_YTD"], auto_headers["Company_ID"],
                auto_first, auto_last, trend_id,
            )
            original_month_formula = _excel_lookup(
                auto_ws.title, auto_headers["Sales_Month"], auto_headers["Company_ID"],
                auto_first, auto_last, trend_id,
            )
            if rewrite_period:
                month_cell.value = (
                    f'=IFERROR(IF(OR({original_month_formula}="",'
                    f'{original_ytd_formula}="",{current_accumulated}=""),"",'
                    f'{original_month_formula}+'
                    f'({current_accumulated}-{original_ytd_formula})),"")'
                )
                month_cell.number_format = "#,##0"

            current_effective = effective_ytd.get((period_key, company_id))
            original_ytd = record["sales_ytd"]
            original_month = record["sales_month"]
            if (
                current_effective in (None, "")
                or original_ytd in (None, "")
                or original_month in (None, "")
            ):
                continue
            corrected_month = original_month + (current_effective - original_ytd)
            if corrected_month != original_month:
                month_details.append([
                    f"{period_key[0]:04d}/{period_key[1]:02d}", company_id,
                    record["company_name"], record["symbol"], original_month,
                    corrected_month,
                    None if original_month in (None, "") else corrected_month - original_month,
                    "Monthly Correction Applied - OK",
                    "Reported month plus accumulated correction delta",
                ])

    if SALES_TREND_MONTH_AUDIT_SHEET in wb.sheetnames:
        del wb[SALES_TREND_MONTH_AUDIT_SHEET]
    month_audit_ws = wb.create_sheet(SALES_TREND_MONTH_AUDIT_SHEET)
    month_headers = [
        "Period", "Company_ID", "Company_Name", "Symbol",
        "Original_Sales_Month", "Corrected_Sales_Month", "Difference",
        "Status", "Method",
    ]
    month_audit_ws.append(month_headers)
    for item in month_details:
        month_audit_ws.append(item)
    month_audit_ws.freeze_panes = "A2"
    month_audit_ws.auto_filter.ref = month_audit_ws.dimensions
    for cell in month_audit_ws[1]:
        cell.font = Font(bold=True)
    for column, width in enumerate((12, 14, 32, 14, 22, 24, 16, 34, 46), 1):
        month_audit_ws.column_dimensions[get_column_letter(column)].width = width
    if affected_periods is None:
        return details, month_details

    affected_labels = {
        f"{year:04d}/{month:02d}" for year, month in affected_periods
    }
    return (
        [item for item in details if item[0] in affected_labels],
        [item for item in month_details if item[0] in affected_labels],
    )


def reorder_workbook_sheets(wb):
    """Keep fixed sheets logical and Auto sheets newest-to-oldest."""
    sales_trend = [wb[SALES_TREND_SHEET]] if SALES_TREND_SHEET in wb.sheetnames else []
    log_sheets = [wb[LOG_SHEET]] if LOG_SHEET in wb.sheetnames else []
    auto_sheets = [ws for ws in wb.worksheets if ws.title.startswith("Auto ")]
    auto_sheets.sort(
        key=lambda ws: parse_period(ws.title[5:].replace(" ", "/")),
        reverse=True,
    )
    fixed_middle = [
        ws for ws in wb.worksheets
        if ws not in sales_trend and ws not in log_sheets and ws not in auto_sheets
    ]
    wb._sheets = sales_trend + fixed_middle + auto_sheets + log_sheets

def load_company_name_map():
    """
    Load Sales-enabled companies from AAI-TSE-Master / Companies.

    The header is detected by name so the registry can tolerate leading title
    rows and column moves. Rows are streamed without relying on ws.max_row,
    which may be None for read-only workbooks with an incomplete dimension.
    """

    if not MASTER_PATH.exists():
        raise FileNotFoundError(
            f"Master workbook not found: {MASTER_PATH}"
        )

    master_wb = load_workbook(
        MASTER_PATH,
        data_only=True,
        read_only=True,
    )

    try:
        if "Companies" not in master_wb.sheetnames:
            raise RuntimeError(
                "Registry sheet not found: Companies"
            )

        company_map = {}
        company_by_name = {}
        ws = master_wb["Companies"]

        required_headers = {
            "company_id",
            "symbol",
            "company_name",
            "monitor",
            "monitor_sales",
            "start_period",
        }
        header_indexes = None
        header_row_number = None
        rows_scanned = 0
        eligible_rows = 0
        skipped_incomplete = 0

        for row_number, values in enumerate(
            ws.iter_rows(values_only=True),
            start=1,
        ):
            rows_scanned += 1

            if header_indexes is None:
                normalized_headers = {
                    re.sub(r"[\s\-]+", "_", normalize_text(value).casefold()): index
                    for index, value in enumerate(values)
                    if normalize_text(value)
                }
                if required_headers.issubset(normalized_headers):
                    header_indexes = normalized_headers
                    header_row_number = row_number
                continue

            def registry_value(column_name):
                index = header_indexes[column_name]
                return values[index] if index < len(values) else None

            monitor = normalize_text(registry_value("monitor"))
            monitor_sales = normalize_text(registry_value("monitor_sales"))

            if monitor.casefold() != "yes" or monitor_sales.casefold() != "yes":
                continue

            eligible_rows += 1
            company_id = normalize_text(registry_value("company_id"))
            symbol = normalize_text(registry_value("symbol"))
            company_name = normalize_text(registry_value("company_name"))
            start_period = normalize_text(registry_value("start_period"))
            start_period = normalize_digits(start_period) or HISTORY_START_PERIOD
            parse_period(start_period)

            if not company_id or not symbol or not company_name:
                skipped_incomplete += 1
                continue

            company = {
                "company_id": company_id,
                "symbol": symbol,
                "company_name": company_name,
                "start_period": start_period,
            }
            company_map[symbol] = company
            company_by_name[company_name] = company

        if header_indexes is None:
            raise RuntimeError(
                "Could not find the Companies header row. Required columns: "
                + ", ".join(sorted(required_headers))
            )

        print(
            "Registry summary: "
            f"header row={header_row_number}, rows scanned={rows_scanned}, "
            f"Sales-enabled rows={eligible_rows}, companies loaded={len(company_map)}, "
            f"incomplete rows skipped={skipped_incomplete}."
        )
    finally:
        master_wb.close()

    return company_map, company_by_name

# ============================================================
# MAIN
# ============================================================

def save_workbook_safely(wb, path):
    """Save beside the workbook, then atomically replace the old file."""
    temporary_path = path.with_name(
        f".{path.stem}.tmp{path.suffix}"
    )

    try:
        wb.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def create_workbook_backup(path):
    """Create one timestamped, recoverable backup before workbook mutation."""
    backup_dir = path.parent.parent / "backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{path.stem}_before_refresh_{timestamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def assert_no_external_formulas(wb):
    """Refuse to save dangling [book]Sheet! formulas that trigger Excel repair."""
    problems = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and re.search(
                    r"\[[^\]]+\]", value
                ):
                    problems.append(f"{ws.title}!{cell.coordinate}")
                    if len(problems) >= 20:
                        break
            if len(problems) >= 20:
                break
        if len(problems) >= 20:
            break
    if problems:
        raise RuntimeError(
            "External workbook formulas remain; workbook was not saved. "
            "First cells: " + ", ".join(problems)
        )


def run_sales_trend_refresh_only(workbook_path=WORKBOOK_PATH):
    """Refresh Sales Trend and its audit without network or Auto-sheet writes."""
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    backup_path = create_workbook_backup(workbook_path)
    print("Workbook backup:", backup_path)
    wb = load_workbook(workbook_path, keep_links=False)
    try:
        assert_no_external_formulas(wb)
        if SALES_TREND_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Required sheet not found: {SALES_TREND_SHEET}")

        forecast_columns_added = ensure_sales_trend_forecast_columns(
            wb[SALES_TREND_SHEET]
        )
        if forecast_columns_added:
            print("Sales Trend Forecast V1 columns added.")
            refresh_sales_trend_prior_corrections(wb)

        auto_snapshot = {
            ws.title: tuple(
                tuple(cell.value for cell in row)
                for row in ws.iter_rows()
            )
            for ws in wb.worksheets
            if ws.title.startswith("Auto ")
        }
        audit_rows, month_audit_rows = refresh_sales_trend_prior_corrections(wb)
        forecast_updated, forecast_unavailable = refresh_sales_forecast_metadata(wb)
        auto_after = {
            ws.title: tuple(
                tuple(cell.value for cell in row)
                for row in ws.iter_rows()
            )
            for ws in wb.worksheets
            if ws.title.startswith("Auto ")
        }
        if auto_snapshot != auto_after:
            raise RuntimeError(
                "Refresh-only safety check failed: at least one Auto sheet changed."
            )

        reorder_workbook_sheets(wb)
        assert_no_external_formulas(wb)
        save_workbook_safely(wb, workbook_path)

        correction_count = sum(
            1 for item in audit_rows if item[9] == "Correction Applied - OK"
        )
        fiscal_reset_count = sum(
            1 for item in audit_rows if item[9] == "Fiscal Reset - Not Applicable"
        )
        print("Sales Trend refresh completed.")
        print("Corrections applied :", correction_count)
        print("Fiscal resets ignored:", fiscal_reset_count)
        print("Monthly sales corrected:", len(month_audit_rows))
        print("Forecast rows updated:", forecast_updated)
        print("Forecast rows unavailable:", forecast_unavailable)
        print("Auto sheets changed : 0")
        print("Workbook saved      :", workbook_path)
        return {
            "backup_path": backup_path,
            "corrections_applied": correction_count,
            "fiscal_resets_ignored": fiscal_reset_count,
            "monthly_sales_corrected": len(month_audit_rows),
            "auto_sheets_changed": 0,
        }
    finally:
        wb.close()


def _mode5_missing_opening_periods(first_period, new_fiscal_year_end):
    """List months before the first available month in its restated fiscal year."""
    fiscal_month = fiscal_year_and_month(
        first_period[0], first_period[1], new_fiscal_year_end
    )[1]
    if fiscal_month <= 1:
        return []
    start_month = (new_fiscal_year_end % 12) + 1
    start_year = first_period[0] if start_month <= first_period[1] else first_period[0] - 1
    result = []
    key = (start_year, start_month)
    while key != first_period:
        result.append(key)
        key = _next_period_key(key)
    return result


def _mode5_parse_amount(raw):
    raw = normalize_digits(normalize_text(raw)).replace(",", "").replace("٬", "")
    if not raw:
        return None
    amount = float(raw)
    if amount < 0:
        raise ValueError("Opening monthly-sales sum cannot be negative.")
    return amount


def run_fiscal_year_change_mode(workbook_path=WORKBOOK_PATH):
    """Register and apply one persistent Sales Trend fiscal-calendar change."""
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    backup_path = create_workbook_backup(workbook_path)
    print("Workbook backup:", backup_path)
    wb = load_workbook(workbook_path, keep_links=False)
    try:
        assert_no_external_formulas(wb)
        ws = wb[SALES_TREND_SHEET]
        ensure_sales_trend_forecast_columns(ws)
        normalize_sales_trend_forecast_layout(ws)
        header_row, headers = find_header_map(
            ws,
            required=(
                "Company_ID", "Company_Name", "Symbol", "Fiscal_Year_End",
                "Calendar_Regime_Flag",
            ),
        )
        query = normalize_text(input("Company name or symbol: ")).casefold()
        matches = []
        for row in range(header_row + 1, ws.max_row + 1):
            name = normalize_text(ws.cell(row, headers["Company_Name"]).value)
            symbol = normalize_text(ws.cell(row, headers["Symbol"]).value)
            if query and query in {name.casefold(), symbol.casefold()}:
                matches.append((row, name, symbol))
        if len(matches) != 1:
            raise ValueError(
                "Enter one exact Company_Name or Symbol; matching rows: "
                + str(len(matches))
            )
        row, company_name, symbol = matches[0]
        company_id = normalize_text(ws.cell(row, headers["Company_ID"]).value)
        stored_fye = int(ws.cell(row, headers["Fiscal_Year_End"]).value)
        old_fye = int(input(f"Previous fiscal year-end month [{stored_fye}]: ").strip() or stored_fye)
        if not 1 <= old_fye <= 12:
            raise ValueError("Previous fiscal year-end month must be between 1 and 12.")
        if old_fye != stored_fye:
            print(
                "Note: Sales Trend currently shows fiscal month "
                f"{stored_fye}; the operator-entered previous month {old_fye} "
                "will be retained in the Mode 5 audit comment."
            )
        new_fye = int(input("New fiscal year-end month (1-12): ").strip())
        if not 1 <= new_fye <= 12:
            raise ValueError("New fiscal year-end month must be between 1 and 12.")

        period_columns = sales_trend_period_columns(ws)
        if not period_columns:
            raise RuntimeError("No Sales Trend monthly history was found.")
        missing = _mode5_missing_opening_periods(min(period_columns), new_fye)
        opening_base = 0.0
        if missing:
            first_missing = f"{missing[0][0]:04d}/{missing[0][1]:02d}"
            last_missing = f"{missing[-1][0]:04d}/{missing[-1][1]:02d}"
            print()
            print("IMPORTANT WARNING")
            print("Do NOT enter Codal cumulative sales.")
            print(
                "Enter only the SUM of monthly sales for the missing opening "
                f"months: {first_missing} through {last_missing}."
            )
            print(
                "Press Enter if unavailable; that incomplete fiscal segment "
                "will be excluded from forecasting."
            )
            opening_base = _mode5_parse_amount(
                input("Missing opening monthly-sales sum: ")
            )
        print()
        print("Company                 :", company_name, f"({symbol})")
        print("Old fiscal year-end     :", old_fye)
        print("New fiscal year-end     :", new_fye)
        print("Opening monthly sum     :", opening_base if opening_base is not None else "Unavailable / Exclude")
        if input("Apply Mode 5 change? [Y/N]: ").strip().casefold() not in {"y", "yes"}:
            print("Mode 5 cancelled; workbook was not changed.")
            return {"cancelled": True, "backup_path": backup_path}

        auto_snapshot = {
            sheet.title: tuple(tuple(cell.value for cell in cells) for cells in sheet.iter_rows())
            for sheet in wb.worksheets if sheet.title.startswith("Auto ")
        }
        payload = {
            "active": True,
            "company_id": company_id,
            "company_name": company_name,
            "symbol": symbol,
            "old_fiscal_year_end": old_fye,
            "new_fiscal_year_end": new_fye,
            "opening_base": opening_base,
            "missing_start": (
                f"{missing[0][0]:04d}/{missing[0][1]:02d}" if missing else None
            ),
            "missing_end": (
                f"{missing[-1][0]:04d}/{missing[-1][1]:02d}" if missing else None
            ),
            "recorded_at": datetime.now().isoformat(timespec="seconds"),
        }
        flag_cell = ws.cell(row, headers["Calendar_Regime_Flag"])
        flag_cell.value = MODE5_FLAG
        flag_cell.comment = Comment(
            MODE5_COMMENT_PREFIX + json.dumps(payload, ensure_ascii=False, indent=2),
            "AAI-TSE Engine",
        )
        applied = apply_mode5_overrides(wb)
        forecast_updated, forecast_unavailable = refresh_sales_forecast_metadata(wb)
        auto_after = {
            sheet.title: tuple(tuple(cell.value for cell in cells) for cells in sheet.iter_rows())
            for sheet in wb.worksheets if sheet.title.startswith("Auto ")
        }
        if auto_snapshot != auto_after:
            raise RuntimeError("Mode 5 safety check failed: an Auto sheet changed.")
        reorder_workbook_sheets(wb)
        assert_no_external_formulas(wb)
        save_workbook_safely(wb, workbook_path)
        print("Mode 5 overrides applied :", applied)
        print("Forecast rows updated    :", forecast_updated)
        print("Forecast rows unavailable:", forecast_unavailable)
        print("Auto sheets changed      : 0")
        print("Workbook saved           :", workbook_path)
        return {
            "cancelled": False,
            "backup_path": backup_path,
            "company_id": company_id,
            "overrides_applied": applied,
            "auto_sheets_changed": 0,
        }
    finally:
        wb.close()


def main():
    print()
    print("=" * 70)
    print("AAI-TSE MONTHLY SALES ENGINE")
    print("=" * 70)
    print()
    print("Select mode:")
    print("1 - Single month")
    print("2 - Historical backfill")
    print("3 - Current Month Live Update")
    print("4 - Refresh Sales Trend only (no network, Auto sheets unchanged)")
    print("5 - Fiscal Year-End Change / Rebuild Sales Trend")
    print()

    mode = input("Mode: ").strip()

    if mode == "4":
        run_sales_trend_refresh_only()
        return
    if mode == "5":
        run_fiscal_year_change_mode()
        return
    live_mode = mode == "3"
    if mode == "1":
        target_period = input(
            "Enter target period (example 1404/01/31): "
        ).strip()
        parse_period(target_period)
        periods = [target_period]
    elif mode == "2":
        start_period = input(
            f"Start period [{HISTORY_START_PERIOD}]: "
        ).strip() or HISTORY_START_PERIOD
        end_period = input(
            f"End period [{HISTORY_END_PERIOD}]: "
        ).strip() or HISTORY_END_PERIOD
        parse_period(start_period)
        parse_period(end_period)
        periods = generate_periods(start_period, end_period)
    elif live_mode:
        suggested_period = current_jalali_month_period()
        target_period = input(
            f"Current target period [{suggested_period}]: "
        ).strip() or suggested_period
        parse_period(target_period)
        periods = [target_period]
    else:
        raise ValueError("Mode must be 1, 2, 3, 4, or 5.")

    if not periods:
        raise ValueError("No periods selected; check the historical range.")

    print()
    print(f"{len(periods)} period(s) selected:")
    for period in periods:
        print("  ", period)
    print()

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    backup_path = create_workbook_backup(WORKBOOK_PATH)
    print("Workbook backup:", backup_path)

    wb = load_workbook(
        WORKBOOK_PATH,
        keep_links=False,
    )
    try:
        # The user-cleaned baseline must stay link-free.  Check before any
        # mutation and again immediately before the atomic save.
        assert_no_external_formulas(wb)
        if SALES_TREND_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Required sheet not found: {SALES_TREND_SHEET}")
        if TEMPLATE_SHEET not in wb.sheetnames:
            raise RuntimeError(
                f"Template sheet not found: {TEMPLATE_SHEET}"
            )

        forecast_columns_added = ensure_sales_trend_forecast_columns(
            wb[SALES_TREND_SHEET]
        )
        if forecast_columns_added:
            print("Sales Trend Forecast V1 columns added.")
            refresh_sales_trend_prior_corrections(wb)

        manual_ws = wb[TEMPLATE_SHEET]
        log_ws = prepare_log_sheet(wb)
        company_map, company_by_name = load_company_name_map()
        parser = MonthlySalesHtmlParser()

        print(
            f"{len(company_map)} company names loaded "
            "from AAI-TSE-Master / Companies."
        )

        period_work = {}
        batch_periods = set()

        # Refresh from the current template while retaining complete named
        # sales fields by authoritative Company_ID. New Registry companies are
        # therefore added without re-fetching already valid company-periods.
        for target_period in periods:
            historical_start_bypass = ignores_company_start_period(target_period)
            auto_sheet_name = period_sheet_name(target_period)
            existing = {}
            if auto_sheet_name in wb.sheetnames:
                existing_ws = wb[auto_sheet_name]
                existing = existing_sales_by_company(
                    existing_ws, company_map, company_by_name
                )
                del wb[auto_sheet_name]

            auto_ws = wb.create_sheet(auto_sheet_name)
            copy_sheet_layout(manual_ws, auto_ws)
            header_row, header_map, symbol_rows = standardize_auto_schema(
                auto_ws, company_map, company_by_name, target_period
            )
            reset_worksheet_view(auto_ws)
            clear_auto_data(auto_ws, header_map, header_row)
            restored, prior_year_pending = restore_existing_sales(
                auto_ws, symbol_rows, existing, header_map
            )
            pending_rows = [
                item for item in symbol_rows
                if item[3] not in restored
                and item[3] not in prior_year_pending
                and (
                    historical_start_bypass
                    or parse_period(item[4]) <= parse_period(target_period)
                )
            ]
            eligible_count = sum(
                1
                for item in symbol_rows
                if (
                    historical_start_bypass
                    or parse_period(item[4]) <= parse_period(target_period)
                )
            )
            missing_count = len(pending_rows)
            missing_ratio = (
                missing_count / eligible_count if eligible_count else 0.0
            )
            if live_mode:
                fetch_strategy = "LIVE_BATCH"
                batch_periods.add(target_period)
            elif not pending_rows:
                fetch_strategy = "NONE"
            elif (
                missing_count <= TARGETED_ONLY_MAX_ROWS
                or missing_ratio < TARGETED_ONLY_MAX_RATIO
            ):
                fetch_strategy = "TARGETED_ONLY"
            else:
                fetch_strategy = "BATCH"
                batch_periods.add(target_period)

            print(
                f"PRE-FETCH STRATEGY: {target_period} | {fetch_strategy} | "
                f"missing={missing_count}/{eligible_count} "
                f"({missing_ratio:.1%}) | thresholds: "
                f"rows<={TARGETED_ONLY_MAX_ROWS} OR "
                f"ratio<{TARGETED_ONLY_MAX_RATIO:.0%}"
            )
            period_work[target_period] = {
                "sheet": auto_ws,
                "sheet_name": auto_sheet_name,
                "symbol_rows": symbol_rows,
                "restored": restored,
                "prior_year_pending": prior_year_pending,
                "pending_rows": pending_rows,
                "fetch_strategy": fetch_strategy,
                "header_row": header_row,
                "header_map": header_map,
                "historical_start_bypass": historical_start_bypass,
            }

        if live_mode:
            live_range = live_publish_range(periods[0])
            publish_ranges = [live_range] if live_range else []
        else:
            publish_ranges = missing_period_publish_ranges(batch_periods)
        all_reports = []
        seen_reports = set()
        api = None
        rate_state = {"last_request_at": float("-inf")}
        prior_year_reports_cache = {}
        targeted_reports_cache = {}
        recovery_reports_cache = {}
        has_pending_rows = live_mode or any(
            work["pending_rows"] or work["prior_year_pending"]
            for work in period_work.values()
        )
        if has_pending_rows:
            api = CodalAPI()
        if publish_ranges:
            print()
            print("=" * 70)
            print(
                "LIVE ANNOUNCEMENT RANGE (CLIPPED TO TODAY)"
                if live_mode else "INCREMENTAL BACKFILL ANNOUNCEMENT RANGES"
            )
            for date_start, date_end in publish_ranges:
                print(f"{date_start} through {date_end}")
                reports = fetch_all_announcements(
                    api,
                    date_start,
                    date_end,
                    company_count=len(company_map),
                    rate_state=rate_state,
                )
                for report in reports:
                    identity = announcement_identity(report)
                    if identity not in seen_reports:
                        seen_reports.add(identity)
                        all_reports.append(report)
            if live_mode:
                print("Existing rows change only for a newly detected revision.")
                print("Missing Live reports remain NOT_YET_PUBLISHED.")
            else:
                print("Existing complete Company+Period rows were not fetched.")
                print("Revision sweep is not part of this backfill mode.")
            print("=" * 70)
        else:
            if has_pending_rows:
                print("No general BRSAPI fetch is required; all gaps use TARGETED_ONLY.")
            else:
                print("All eligible Company+Period rows already contain valid data.")
                print("No BRSAPI announcement request is required.")

        print(
            "\nDEBUG: total reports fetched =",
            len(all_reports),
        )
        print(
            "DEBUG: first report =",
            all_reports[0] if all_reports else None,
        )

        for target_period in periods:
            work = period_work[target_period]
            auto_sheet_name = work["sheet_name"]
            auto_ws = work["sheet"]
            symbol_rows = work["symbol_rows"]
            restored = work["restored"]
            prior_year_pending = work["prior_year_pending"]
            header_map = work["header_map"]
            historical_start_bypass = work["historical_start_bypass"]
            pending_company_ids = {item[3] for item in work["pending_rows"]}
            fetch_strategy = work["fetch_strategy"]
            logged_links = latest_logged_report_links(log_ws, target_period)

            print()
            print("=" * 70)
            print(f"PERIOD: {target_period}")
            print(f"PRE-FETCH STRATEGY: {fetch_strategy}")
            print("=" * 70)

            print(f"{len(symbol_rows)} symbol rows found.")

            success = 0
            missing = 0
            failed = 0
            revised = 0
            live_revised_count = 0
            skipped = 0
            existing_skipped = 0
            new_count = 0
            pending_count = 0
            prior_year_filled = 0
            prior_year_zero_confirmed = 0
            prior_year_unavailable = 0
            total = len(symbol_rows)

            print()
            print("=" * 70)
            print(f"PROCESSING SYMBOLS FOR {target_period}")
            print("=" * 70)

            for index, (
                row,
                symbol,
                company_name,
                company_id,
                start_period,
            ) in enumerate(
                symbol_rows,
                start=1,
            ):
                print()
                print(f"[{index}/{total}] {symbol}")

                existing_live_row = live_mode and company_id in restored
                if company_id in restored and not live_mode:
                    print("  SKIPPED: valid existing Company+Period data")
                    existing_skipped += 1
                    continue
                if existing_live_row:
                    candidate, candidate_count = select_latest_report(
                        all_reports, symbol, company_name, target_period
                    )
                    previous_link = logged_links.get(normalize_text(symbol), "")
                    if live_report_disposition(
                        True, candidate, previous_link
                    ) == "EXISTING":
                        print("  EXISTING: valid data retained; no new revision")
                        existing_skipped += 1
                        continue

                if company_id in prior_year_pending:
                    if live_mode:
                        print(
                            "  EXISTING: current report retained; prior-year "
                            "recovery deferred in Live Mode"
                        )
                        existing_skipped += 1
                        continue
                    print(
                        "  CURRENT REPORT VALID: prior-year comparison "
                        "is empty or zero"
                    )
                    parsed = {
                        parser_key: auto_ws.cell(
                            row=row, column=header_map[header]
                        ).value
                        for parser_key, header in SALES_HEADERS.items()
                    }
                    try:
                        fallback_status, prior_report, fallback_message = (
                            apply_prior_year_sales_fallback(
                                parsed,
                                symbol,
                                company_name,
                                target_period,
                                parser,
                                api,
                                len(company_map),
                                rate_state,
                                prior_year_reports_cache,
                            )
                        )
                    except Exception as error:
                        fallback_status = "LOOKUP_FAILED"
                        prior_report = None
                        fallback_message = (
                            "Targeted prior-year lookup failed "
                            f"({type(error).__name__}: {error}); current-report "
                            "values were preserved."
                        )
                    if fallback_status == "FILLED":
                        prior_year_filled += 1
                        print("  PRIOR-YEAR FALLBACK FILLED:", fallback_message)
                    elif fallback_status == "CONFIRMED_ZERO":
                        prior_year_zero_confirmed += 1
                        print("  PRIOR-YEAR ZERO CONFIRMED:", fallback_message)
                    else:
                        prior_year_unavailable += 1
                        print("  PRIOR-YEAR FALLBACK UNAVAILABLE:", fallback_message)
                    if live_mode and existing_live_row:
                        write_parser_result_preserving_valid(
                            auto_ws, row, parsed, header_map
                        )
                    else:
                        write_parser_result(auto_ws, row, parsed, header_map)
                    log_result(
                        log_ws,
                        symbol,
                        "PRIOR_YEAR_" + fallback_status,
                        target_period,
                        report=prior_report,
                        report_count=1 if prior_report is not None else 0,
                        message=fallback_message or "",
                    )
                    success += 1
                    continue

                if (
                    not historical_start_bypass
                    and parse_period(start_period) > parse_period(target_period)
                ):
                    print(
                        f"  SKIPPED: Company_ID {company_id} starts at "
                        f"{start_period}, after target period {target_period}."
                    )
                    skipped += 1
                    continue

                if company_id not in pending_company_ids and not existing_live_row:
                    continue

                if existing_live_row:
                    report, report_count = candidate, candidate_count
                elif fetch_strategy == "TARGETED_ONLY":
                    # Search the report's normal publication window first.
                    # Revision discovery remains the responsibility of the
                    # explicit CODAL HTML revision chain.  Late-publication
                    # recovery is used only when this normal lookup misses.
                    report, report_count = fetch_targeted_period_reports(
                        api,
                        symbol,
                        company_name,
                        target_period,
                        rate_state,
                        targeted_reports_cache,
                    )
                else:
                    report, report_count = select_latest_report(
                        all_reports,
                        symbol,
                        company_name,
                        target_period,
                    )
                recovered_report = False
                if should_recover_missing_report(report, api, live_mode):
                    report, report_count = recover_missing_report(
                        api,
                        symbol,
                        company_name,
                        target_period,
                        rate_state,
                        recovery_reports_cache,
                        include_normal_window=False,
                    )
                    recovered_report = report is not None

                if report is None:
                    existing_data_preserved = any(
                        auto_ws.cell(row=row, column=header_map[header]).value
                        is not None
                        for header in SALES_HEADERS.values()
                    )
                    if live_mode:
                        status = (
                            "NOT_YET_PUBLISHED_EXISTING_DATA_PRESERVED"
                            if existing_data_preserved else "NOT_YET_PUBLISHED"
                        )
                    else:
                        status = (
                            "MISSING_REPORT_EXISTING_DATA_PRESERVED"
                            if existing_data_preserved else "MISSING_REPORT"
                        )
                    print(
                        ("  PENDING / NOT_YET_PUBLISHED" if live_mode else "  MISSING REPORT")
                        + (" — EXISTING DATA PRESERVED" if existing_data_preserved else "")
                    )
                    log_result(
                        log_ws,
                        symbol,
                        status,
                        target_period,
                        report_count=0,
                        message=(
                            (
                                "No report published through today's Live window."
                                if live_mode else
                                "No matching report found for target period."
                            )
                            + (
                                " Existing sales values were preserved."
                                if existing_data_preserved else ""
                            )
                        ),
                    )
                    missing += 1
                    if live_mode:
                        pending_count += 1
                    continue

                if report_count > 1:
                    print(
                        f"  {report_count} reports found "
                        "-> latest selected"
                    )
                    revised += 1

                try:
                    report, html, html_path, revision_hops = (
                        resolve_latest_revision(
                            report,
                            symbol,
                            target_period,
                            rate_state,
                        )
                    )
                    resolved_report_count = report_count + revision_hops
                    if revision_hops and report_count <= 1:
                        revised += 1
                    parsed = parser.parse(html)
                    fallback_status, prior_report, fallback_message = (
                        apply_prior_year_sales_fallback(
                            parsed,
                            symbol,
                            company_name,
                            target_period,
                            parser,
                            api,
                            len(company_map),
                            rate_state,
                            prior_year_reports_cache,
                        )
                    )
                    if fallback_status == "FILLED":
                        prior_year_filled += 1
                        print("  PRIOR-YEAR FALLBACK FILLED:", fallback_message)
                    elif fallback_status == "CONFIRMED_ZERO":
                        prior_year_zero_confirmed += 1
                        print("  PRIOR-YEAR ZERO CONFIRMED:", fallback_message)
                    elif fallback_status != "NOT_NEEDED":
                        prior_year_unavailable += 1
                        print("  PRIOR-YEAR FALLBACK UNAVAILABLE:", fallback_message)
                    if live_mode and existing_live_row:
                        write_parser_result_preserving_valid(
                            auto_ws, row, parsed, header_map
                        )
                    else:
                        write_parser_result(auto_ws, row, parsed, header_map)

                    if live_mode and existing_live_row:
                        status = "REVISED"
                    elif live_mode:
                        status = "NEW"
                    elif recovered_report:
                        status = "RECOVERED_REPORT"
                    elif report_count > 1 or revision_hops:
                        status = "REVISION_SELECTED"
                    else:
                        status = "OK"
                    log_result(
                        log_ws,
                        symbol,
                        status,
                        target_period,
                        report=report,
                        report_count=resolved_report_count,
                        html_path=html_path,
                        message=(
                            f"Revision chain hops: {revision_hops}. "
                            + (fallback_message or "")
                        ).strip(),
                    )
                    success += 1
                    if live_mode and existing_live_row:
                        live_revised_count += 1
                    elif live_mode:
                        new_count += 1

                    print(
                        "  SUCCESS |",
                        parsed["sales_last_year"],
                        parsed["sales_ytd"],
                        parsed["sales_month"],
                        parsed["sales_prior_month_ytd"],
                        parsed["export_last_year"],
                        parsed["export_ytd"],
                        parsed["export_month"],
                    )
                except Exception as exc:
                    failed += 1
                    print("  FAILED:", type(exc).__name__, exc)
                    log_result(
                        log_ws,
                        symbol,
                        "PARSE_FAILED",
                        target_period,
                        report=report,
                        report_count=report_count,
                        message=f"{type(exc).__name__}: {exc}",
                    )

            period_has_sales_ytd = any(
                auto_ws.cell(
                    row=item[0], column=header_map["Sales_YTD"]
                ).value not in (None, "")
                for item in symbol_rows
            )
            if period_has_sales_ytd:
                update_sales_trend(
                    wb,
                    target_period,
                    auto_ws,
                    work["header_row"],
                    header_map,
                )
                correction_audit, month_correction_audit = (
                    refresh_sales_trend_prior_corrections(
                        wb,
                        affected_periods={
                            _period_key(target_period),
                            _period_key(_previous_period(target_period)),
                        },
                    )
                )
            else:
                correction_audit, month_correction_audit = [], []
                print(
                    "Sales Trend update/refresh skipped: this period has no "
                    "valid Sales_YTD data."
                )
            correction_count = sum(
                1 for item in correction_audit
                if item[9] == "Correction Applied - OK"
            )
            fiscal_reset_count = sum(
                1 for item in correction_audit
                if item[9] == "Fiscal Reset - Not Applicable"
            )
            forecast_updated, forecast_unavailable = (
                refresh_sales_forecast_metadata(wb)
            )
            reorder_workbook_sheets(wb)

            # Existing Auto sheets may contain pane-specific selections left
            # behind by an older openpyxl save.  With no pane element those
            # selections make Excel repair the worksheet view on every open.
            for worksheet in wb.worksheets:
                if worksheet.title.startswith("Auto "):
                    reset_worksheet_view(worksheet)

            assert_no_external_formulas(wb)
            save_workbook_safely(wb, WORKBOOK_PATH)

            print()
            print("=" * 70)
            print(f"AAI-TSE {'LIVE' if live_mode else 'BATCH'} SUMMARY: {target_period}")
            print("=" * 70)
            print("Symbols        :", total)
            print("Success        :", success)
            print("Missing report :", missing)
            print("Before start    :", skipped)
            print("Existing valid :", existing_skipped)
            print("Parse failed   :", failed)
            print("Prior-year fallback filled     :", prior_year_filled)
            print("Prior-year zero confirmed      :", prior_year_zero_confirmed)
            print("Prior-year fallback unavailable:", prior_year_unavailable)
            print("Multiple/revised candidates:", revised)
            if live_mode:
                print("New            :", new_count)
                print("Revised        :", live_revised_count)
                print("Existing       :", existing_skipped)
                print("Pending        :", pending_count)
            print("Workbook saved :", WORKBOOK_PATH)
            print("Auto sheet     :", auto_sheet_name)
            print("Sales Trend corrections applied:", correction_count)
            print("Fiscal-year resets excluded     :", fiscal_reset_count)
            print("Monthly sales corrected         :", len(month_correction_audit))
            print("Forecast rows updated           :", forecast_updated)
            print("Forecast rows unavailable       :", forecast_unavailable)
            print("=" * 70)

        print()
        print("=" * 70)
        print("ALL SELECTED PERIODS COMPLETED")
        print("=" * 70)
    finally:
        wb.close()


if __name__ == "__main__":
    main()
