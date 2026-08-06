from __future__ import annotations

import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# USER SETTINGS
# ============================================================
# Put the MTSE workbook and downloaded monthly-report workbooks anywhere.
# The program will ask you to choose them when it starts.

MASTER_SHEET_NAMES = ["Cement", "MOPFRA"]
SYMBOL_HEADER_CANDIDATES = ["نماد", "symbol", "ticker"]
OUTPUT_SHEET_NAME = "Monthly Sales Extract"

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
SKIP_TEMP_FILES = True


# ============================================================
# TEXT / NUMBER HELPERS
# ============================================================
def normalize_text(value: Any) -> str:
    """Normalize Persian/Arabic text for reliable comparisons."""
    if value is None:
        return ""

    text = str(value)
    replacements = {
        "ي": "ی",
        "ك": "ک",
        "ۀ": "ه",
        "ة": "ه",
        "ؤ": "و",
        "إ": "ا",
        "أ": "ا",
        "ٱ": "ا",
        "ـ": "",
        "\u200c": " ",  # zero-width non-joiner
        "\u200f": "",
        "\u200e": "",
        "\ufeff": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def compact_text(value: Any) -> str:
    """Normalized text without spaces/punctuation, useful for matching."""
    text = normalize_text(value).lower()
    return re.sub(r"[^0-9a-zآ-ی]+", "", text)


def numeric_value(value: Any) -> int | float | None:
    """Convert Excel values or formatted text to a numeric value."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value

    text = normalize_text(value)
    persian_digits = "۰۱۲۳۴۵۶۷۸۹"
    arabic_digits = "٠١٢٣٤٥٦٧٨٩"
    for i, digit in enumerate(persian_digits):
        text = text.replace(digit, str(i))
    for i, digit in enumerate(arabic_digits):
        text = text.replace(digit, str(i))

    text = text.replace(",", "").replace("٬", "").replace(" ", "")
    text = text.replace("−", "-")
    if text in {"", "-", "—"}:
        return None

    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return None


# ============================================================
# FILE/FOLDER PICKERS
# ============================================================
def choose_file(title: str, filetypes: list[tuple[str, str]]) -> Path:
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askopenfilename(title=title, filetypes=filetypes)
        root.destroy()
        if selected:
            return Path(selected)
    except Exception:
        pass

    while True:
        raw = input(f"{title}\nPaste the full path: ").strip().strip('"').strip("'")
        path = Path(raw).expanduser()
        if path.is_file():
            return path
        print("File not found. Please try again.\n")


def choose_folder(title: str) -> Path:
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askdirectory(title=title)
        root.destroy()
        if selected:
            return Path(selected)
    except Exception:
        pass

    while True:
        raw = input(f"{title}\nPaste the full folder path: ").strip().strip('"').strip("'")
        path = Path(raw).expanduser()
        if path.is_dir():
            return path
        print("Folder not found. Please try again.\n")


def choose_output_file(default_name: str) -> Path:
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.asksaveasfilename(
            title="Save extracted output workbook",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        root.destroy()
        if selected:
            return Path(selected)
    except Exception:
        pass

    raw = input(
        f"Output file path (press Enter for ./{default_name}): "
    ).strip().strip('"').strip("'")
    return Path(raw).expanduser() if raw else Path.cwd() / default_name


# ============================================================
# MASTER SYMBOL LIST
# ============================================================
def find_symbol_column(ws) -> tuple[int, int] | None:
    """Find the column whose header is نماد / Symbol / Ticker."""
    candidates = {compact_text(x) for x in SYMBOL_HEADER_CANDIDATES}
    max_scan_row = min(ws.max_row, 30)
    max_scan_col = min(ws.max_column, 100)

    for row in range(1, max_scan_row + 1):
        for col in range(1, max_scan_col + 1):
            if compact_text(ws.cell(row, col).value) in candidates:
                return row, col
    return None


def read_master_symbols(master_path: Path) -> list[str]:
    wb = load_workbook(master_path, read_only=True, data_only=True)
    symbols: list[str] = []

    requested_sheets = [name for name in MASTER_SHEET_NAMES if name in wb.sheetnames]
    sheets_to_scan = requested_sheets or wb.sheetnames

    for sheet_name in sheets_to_scan:
        ws = wb[sheet_name]
        symbol_location = find_symbol_column(ws)
        if symbol_location is None:
            print(f"Warning: symbol header not found in sheet '{sheet_name}'.")
            continue

        header_row, symbol_col = symbol_location
        for row in range(header_row + 1, ws.max_row + 1):
            value = normalize_text(ws.cell(row, symbol_col).value)
            if not value:
                continue
            if compact_text(value) in {compact_text(x) for x in SYMBOL_HEADER_CANDIDATES}:
                continue
            if value.startswith("="):
                continue
            symbols.append(value)

    wb.close()

    # Preserve order while removing duplicates.
    unique: list[str] = []
    seen: set[str] = set()
    for symbol in symbols:
        key = compact_text(symbol)
        if key and key not in seen:
            seen.add(key)
            unique.append(symbol)

    if not unique:
        raise RuntimeError(
            "No symbols were found in the MTSE workbook. "
            "Check MASTER_SHEET_NAMES and SYMBOL_HEADER_CANDIDATES at the top of the script."
        )

    return unique


# ============================================================
# REPORT PARSING
# ============================================================
@dataclass
class ReportResult:
    symbol: str
    report_file: str
    sheet_name: str
    report_month: str | None

    export_previous_original: int | float | None
    export_previous_corrected: int | float | None
    export_current_month: int | float | None
    export_ytd_current: int | float | None
    export_ytd_prior_year: int | float | None

    total_previous_original: int | float | None
    total_previous_corrected: int | float | None
    total_current_month: int | float | None
    total_ytd_current: int | float | None
    total_ytd_prior_year: int | float | None

    status: str
    note: str


def merged_anchor_value(ws, row: int, col: int) -> Any:
    """Return the visible value even when the cell belongs to a merged range."""
    value = ws.cell(row, col).value
    if value is not None:
        return value

    coordinate = ws.cell(row, col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def find_row(ws, exact_labels: Iterable[str], start_row: int = 1) -> int | None:
    targets = {compact_text(x) for x in exact_labels}
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            if compact_text(ws.cell(row, col).value) in targets:
                return row
    return None


def find_sales_header_row(ws) -> int | None:
    """Find the row containing sales-period group headers."""
    for row in range(1, min(ws.max_row, 40) + 1):
        values = [normalize_text(merged_anchor_value(ws, row, col)) for col in range(1, ws.max_column + 1)]
        joined = " | ".join(v for v in values if v)
        if "دوره یک ماهه" in joined and "از ابتدای سال مالی" in joined:
            # Confirm this is the production/sales section, not energy.
            next_row = " | ".join(
                normalize_text(ws.cell(row + 1, col).value)
                for col in range(1, ws.max_column + 1)
                if ws.cell(row + 1, col).value is not None
            ) if row < ws.max_row else ""
            if "مبلغ فروش" in next_row:
                return row
    return None


def detect_period_columns(ws, header_row: int) -> dict[str, int]:
    """Find the amount column for each desired sales period."""
    amount_columns: list[int] = []
    subheader_row = header_row + 1
    for col in range(1, ws.max_column + 1):
        if "مبلغ فروش" in normalize_text(ws.cell(subheader_row, col).value):
            amount_columns.append(col)

    period_columns: dict[str, int] = {}
    ytd_headers: list[tuple[int, str]] = []

    for col in amount_columns:
        header = normalize_text(merged_anchor_value(ws, header_row, col))
        compact = compact_text(header)

        if "اصلاحات" in header and "اصلاح شده" not in header:
            # Not one of the requested output fields.
            continue
        if "اصلاح شده" in header:
            period_columns["previous_corrected"] = col
        elif "دوره یک ماهه" in header or "دوره یکماهه" in header:
            period_columns["current_month"] = col
        elif "از ابتدای سال مالی" in header:
            ytd_headers.append((col, compact))

    # YTD groups are normally ordered: previous original, current YTD, prior-year YTD.
    if len(ytd_headers) >= 1:
        period_columns["previous_original"] = ytd_headers[0][0]
    if len(ytd_headers) >= 2:
        period_columns["ytd_current"] = ytd_headers[1][0]
    if len(ytd_headers) >= 3:
        period_columns["ytd_prior_year"] = ytd_headers[2][0]

    required = {
        "previous_original",
        "previous_corrected",
        "current_month",
        "ytd_current",
        "ytd_prior_year",
    }
    missing = required - period_columns.keys()
    if missing:
        raise ValueError(f"Could not identify sales period columns: {sorted(missing)}")

    return period_columns


def extract_report_month(ws, header_row: int, current_month_col: int) -> str | None:
    text = normalize_text(merged_anchor_value(ws, header_row, current_month_col))
    persian_digits = "۰۱۲۳۴۵۶۷۸۹"
    arabic_digits = "٠١٢٣٤٥٦٧٨٩"
    for i, digit in enumerate(persian_digits):
        text = text.replace(digit, str(i))
    for i, digit in enumerate(arabic_digits):
        text = text.replace(digit, str(i))
    match = re.search(r"(1[34]\d{2}/\d{2}/\d{2})", text)
    return match.group(1) if match else None


def identify_symbol(report_path: Path, symbols: list[str], workbook=None) -> str | None:
    """Match a master symbol first in filename, then in workbook cells."""
    filename_key = compact_text(report_path.stem)
    matches = [s for s in symbols if compact_text(s) and compact_text(s) in filename_key]
    if matches:
        return max(matches, key=lambda x: len(compact_text(x)))

    if workbook is not None:
        symbol_map = {compact_text(s): s for s in symbols}
        for ws in workbook.worksheets:
            max_rows = min(ws.max_row, 25)
            max_cols = min(ws.max_column, 30)
            for row in range(1, max_rows + 1):
                for col in range(1, max_cols + 1):
                    key = compact_text(ws.cell(row, col).value)
                    if key in symbol_map:
                        return symbol_map[key]
    return None


def parse_report(report_path: Path, symbols: list[str]) -> ReportResult | None:
    wb = load_workbook(report_path, read_only=False, data_only=True)
    symbol = identify_symbol(report_path, symbols, wb)
    if symbol is None:
        wb.close()
        return None  # Report does not belong to an MTSE symbol.

    last_error = "Production and sales table not found."

    for ws in wb.worksheets:
        try:
            header_row = find_sales_header_row(ws)
            if header_row is None:
                continue

            columns = detect_period_columns(ws, header_row)
            export_row = find_row(ws, ["جمع فروش صادراتی"], start_row=header_row + 1)

            # The required grand-total row is the first exact "جمع" after the sales rows
            # and before the "خرید مواد اولیه" section.
            raw_total_row = find_row(ws, ["جمع"], start_row=header_row + 1)
            purchase_row = find_row(ws, ["خرید مواد اولیه"], start_row=header_row + 1)
            total_row = raw_total_row if raw_total_row and (not purchase_row or raw_total_row < purchase_row) else None

            if export_row is None:
                raise ValueError("Row 'جمع فروش صادراتی' not found.")
            if total_row is None:
                raise ValueError("Grand-total row 'جمع' not found before 'خرید مواد اولیه'.")

            def value_at(row: int, key: str) -> int | float | None:
                return numeric_value(ws.cell(row, columns[key]).value)

            result = ReportResult(
                symbol=symbol,
                report_file=report_path.name,
                sheet_name=ws.title,
                report_month=extract_report_month(ws, header_row, columns["current_month"]),

                export_previous_original=value_at(export_row, "previous_original"),
                export_previous_corrected=value_at(export_row, "previous_corrected"),
                export_current_month=value_at(export_row, "current_month"),
                export_ytd_current=value_at(export_row, "ytd_current"),
                export_ytd_prior_year=value_at(export_row, "ytd_prior_year"),

                total_previous_original=value_at(total_row, "previous_original"),
                total_previous_corrected=value_at(total_row, "previous_corrected"),
                total_current_month=value_at(total_row, "current_month"),
                total_ytd_current=value_at(total_row, "ytd_current"),
                total_ytd_prior_year=value_at(total_row, "ytd_prior_year"),

                status="OK",
                note="",
            )
            wb.close()
            return result

        except Exception as exc:
            last_error = f"{ws.title}: {exc}"

    wb.close()
    return ReportResult(
        symbol=symbol,
        report_file=report_path.name,
        sheet_name="",
        report_month=None,
        export_previous_original=None,
        export_previous_corrected=None,
        export_current_month=None,
        export_ytd_current=None,
        export_ytd_prior_year=None,
        total_previous_original=None,
        total_previous_corrected=None,
        total_current_month=None,
        total_ytd_current=None,
        total_ytd_prior_year=None,
        status="ERROR",
        note=last_error,
    )


# ============================================================
# OUTPUT WORKBOOK
# ============================================================
OUTPUT_HEADERS = [
    "Symbol",
    "Report Month",
    "Report File",
    "Source Sheet",
    "Export Sales - Previous Original",
    "Export Sales - Previous Corrected",
    "Export Sales - Current Month",
    "Export Sales - YTD Current",
    "Export Sales - YTD Prior Year",
    "Total Sales - Previous Original",
    "Total Sales - Previous Corrected",
    "Total Sales - Current Month",
    "Total Sales - YTD Current",
    "Total Sales - YTD Prior Year",
    "Status",
    "Note",
]


def result_to_row(result: ReportResult) -> list[Any]:
    return [
        result.symbol,
        result.report_month,
        result.report_file,
        result.sheet_name,
        result.export_previous_original,
        result.export_previous_corrected,
        result.export_current_month,
        result.export_ytd_current,
        result.export_ytd_prior_year,
        result.total_previous_original,
        result.total_previous_corrected,
        result.total_current_month,
        result.total_ytd_current,
        result.total_ytd_prior_year,
        result.status,
        result.note,
    ]


def write_output(results: list[ReportResult], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_NAME
    ws.sheet_view.rightToLeft = False
    ws.append(OUTPUT_HEADERS)

    # One report per symbol/month. Keep latest duplicate by file modification time
    # before calling this function; rows here are already sorted.
    for result in results:
        ws.append(result_to_row(result))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    error_fill = PatternFill("solid", fgColor="FCE4D6")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 15).alignment = Alignment(horizontal="center")
        if ws.cell(row, 15).value != "OK":
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = error_fill

    for col in range(5, 15):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = "#,##0"

    widths = {
        1: 14,
        2: 14,
        3: 42,
        4: 24,
        15: 12,
        16: 55,
    }
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 22)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 38

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================
def main() -> int:
    print("\nAA-TSE Monthly Sales Extractor\n")

    master_path = choose_file(
        "Select the latest MTSE / AA-TSE master workbook",
        [("Excel Workbooks", "*.xlsx *.xlsm"), ("All Files", "*.*")],
    )
    reports_folder = choose_folder("Select the folder containing downloaded monthly reports")
    output_path = choose_output_file("AA-TSE-Monthly-Sales-Extract.xlsx")

    print("\nReading MTSE symbols...")
    symbols = read_master_symbols(master_path)
    print(f"Found {len(symbols)} unique symbols.")

    report_files = sorted(
        [
            path
            for path in reports_folder.rglob("*")
            if path.is_file()
            and path.suffix.lower() in SUPPORTED_EXTENSIONS
            and (not SKIP_TEMP_FILES or not path.name.startswith("~$"))
            and path.resolve() != master_path.resolve()
            and path.resolve() != output_path.resolve()
        ],
        key=lambda p: p.stat().st_mtime,
    )

    print(f"Found {len(report_files)} Excel report files.\n")

    results: list[ReportResult] = []
    ignored = 0

    for index, report_path in enumerate(report_files, start=1):
        try:
            result = parse_report(report_path, symbols)
            if result is None:
                ignored += 1
                print(f"[{index}/{len(report_files)}] IGNORE  {report_path.name}")
                continue

            results.append(result)
            print(
                f"[{index}/{len(report_files)}] {result.status:<5}   "
                f"{result.symbol:<12} {report_path.name}"
            )
        except Exception as exc:
            print(f"[{index}/{len(report_files)}] FAILED  {report_path.name}: {exc}")

    # Remove duplicate symbol/month records, retaining the last processed file.
    deduplicated: dict[tuple[str, str], ReportResult] = {}
    for result in results:
        key = (compact_text(result.symbol), result.report_month or result.report_file)
        deduplicated[key] = result

    final_results = sorted(
        deduplicated.values(),
        key=lambda r: (compact_text(r.symbol), r.report_month or "", r.report_file),
    )

    write_output(final_results, output_path)

    ok_count = sum(r.status == "OK" for r in final_results)
    error_count = sum(r.status != "OK" for r in final_results)

    print("\nCompleted.")
    print(f"Extracted successfully: {ok_count}")
    print(f"Reports with parsing errors: {error_count}")
    print(f"Ignored because symbol was not in MTSE: {ignored}")
    print(f"Output: {output_path.resolve()}\n")

    return 0 if error_count == 0 else 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nCancelled by user.")
        raise SystemExit(130)
    except Exception as exc:
        print(f"\nERROR: {exc}")
        raise SystemExit(1)
